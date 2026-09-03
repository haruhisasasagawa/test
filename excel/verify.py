#!/usr/bin/env python3
"""
再計算済みのxlsxを読み、Excelの数式とは別にPythonで同じ計算をやり直して突き合わせる。

  python3 verify.py <再計算済みのxlsx>

見るもの
  ・商品の並び（商品マスタの順 → 当日CSVだけ → 1ヶ月前だけ → 2ヶ月前だけ）と区分（仕分け）
  ・区分ごとの納品日フラグと発注から納品までの日数
  ・在庫（CSVの総数、数えた在庫の上書き）、定数（マスタ → CSVの規定数）
  ・14日の在庫予測（減り数＝100人あたりの減り数×動員÷100、0で止める、入庫は届く日の列）
  ・切れる日、次の便、目安、発注期限、納品日でない入庫・定数超え
  ・「このままだと」「入庫を入れると」の記号（● △ ○ ？ － ⚠）
  ・③発注書の明細（商品順→日付順）と合計、区分シートの「発注予定額」
"""
import datetime
import math
import sys
from collections import OrderedDict

import openpyxl

DAYS = 14
NDT = 21
ORD_FIRST = 10
KINDS = ['冷凍', '飲料', '常温']
IN_FIRST = 13
PJ_FIRST = 27
COL = {'name': 3, 'par': 4, 'stock': 5, 'use': 6, 'cut': 7, 'status': 8, 'next': 9, 'guide': 10,
       'unit': 11, 'after': 12, 'counted': 41, 'counted_date': 42, 'code': 43, 'vendor': 44}
SORT_DELIV_ROW = {'冷凍': 6, '飲料': 7, '常温': 8}
SORT_CAT = (12, 31)
SORT_OV_FIRST = 36


def as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    return v


def num(v, default=0):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else default


def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    st = wb['設定']
    theater = str(st['C4'].value)
    base = as_date(st['C6'].value)
    lt_default = int(num(st['C7'].value, 0))
    std = num(st['C8'].value, 0)
    assert isinstance(base, datetime.date), f'当日基準日が取れていません: {base!r}'

    def csv_rows(name):
        ws = wb[name]
        out = []
        for r in range(6, ws.max_row + 1):
            th = ws.cell(r, 2).value
            if th in (None, ''):
                continue
            th = str(th).zfill(4) if isinstance(th, int) else str(th)
            code = ws.cell(r, 9).value
            code = str(code) if not isinstance(code, float) else str(int(code))
            out.append({'th': th, 'code': code, 'name': ws.cell(r, 10).value, 'cat': ws.cell(r, 8).value,
                        'vendor': ws.cell(r, 5).value, 'pack': num(ws.cell(r, 11).value),
                        'total': num(ws.cell(r, 14).value), 'kitei': num(ws.cell(r, 15).value),
                        'price': num(ws.cell(r, 17).value)})
        return out

    sheets = {'cur': csv_rows('①在庫を貼る'), 'mid': csv_rows('月1回_1ヶ月前の在庫'), 'prv': csv_rows('月1回_2ヶ月前の在庫')}
    mine = {k: [x for x in v if x['th'] == theater] for k, v in sheets.items()}
    seen_any = {x['code'] for v in mine.values() for x in v}

    # 商品マスタ
    im = wb['商品マスタ']
    master = OrderedDict()
    master_row = {}
    for r in range(6, im.max_row + 1):
        code = im.cell(r, 2).value
        if code in (None, ''):
            continue
        master[str(code)] = {'name': im.cell(r, 3).value, 'cat': im.cell(r, 4).value, 'vendor': im.cell(r, 5).value,
                             'pack': im.cell(r, 6).value, 'price': im.cell(r, 7).value,
                             'par_case': im.cell(r, 8).value, 'pi': im.cell(r, 10).value}
        master_row[str(code)] = r

    # 仕分け
    so = wb['仕分け']
    deliv = {}
    lt = {}
    for k in KINDS:
        rr = SORT_DELIV_ROW[k]
        marks = [so.cell(rr, 3 + i).value == '○' for i in range(7)]      # 月..日
        deliv[k] = marks if any(marks) else [True] * 7
        j = so.cell(rr, 10).value
        lt[k] = int(j) if is_num(j) else lt_default
    cat_kind = {}
    for r in range(SORT_CAT[0], SORT_CAT[1] + 1):
        c = so.cell(r, 2).value
        if c not in (None, ''):
            cat_kind[c] = so.cell(r, 3).value or None

    def override(code):
        r = master_row.get(code)
        if r is None:
            return None
        v = so.cell(SORT_OV_FIRST + r - 6, 6).value
        return v or None

    order = [c for c in master if c in seen_any]
    for key in ('cur', 'mid', 'prv'):
        for x in mine[key]:
            if x['code'] not in master and x['code'] not in order:
                order.append(x['code'])

    def csv_text(code, field):
        for key in ('cur', 'mid', 'prv'):
            for x in sheets[key]:
                if x['code'] == code:
                    return x[field]
        return ''

    def csv_num(code, field):
        for key in ('cur', 'mid', 'prv'):
            hits = [x[field] for x in mine[key] if x['code'] == code]
            if hits:
                return sum(hits) / len(hits)
        return 0

    pl = wb['動員を入れる']
    att = {}
    for r in range(16, 416):
        d = as_date(pl.cell(r, 2).value)
        if isinstance(d, datetime.date):
            att[d] = num(pl.cell(r, 7).value)
    days = [base + datetime.timedelta(days=d) for d in range(NDT)]
    day_att = [att.get(d, std) for d in days[:DAYS]]
    att_sum = sum(day_att)

    def flag(kind, d):
        return deliv[kind][days[d].weekday()]

    problems = []
    checked = 0
    sheet_rows = []
    amount = {k: 0 for k in KINDS}
    seq = {k: 0 for k in KINDS}
    for i, code in enumerate(order):
        if i >= 400:
            break
        m = master.get(code, {})
        in_cur = any(x['code'] == code for x in mine['cur'])

        def mval(field, csv_field, numeric):
            v = m.get(field)
            if v not in (None, ''):
                return v
            return csv_num(code, csv_field) if numeric else csv_text(code, csv_field)
        name = mval('name', 'name', False)
        cat = mval('cat', 'cat', False)
        vendor = mval('vendor', 'vendor', False)
        pack = max(1, num(mval('pack', 'pack', True)))
        price = num(mval('price', 'price', True))
        kind = override(code) or cat_kind.get(cat)
        if kind not in KINDS:
            kind = None
        pi = m.get('pi')
        pi = pi if is_num(pi) and pi > 0 else None
        par_case = m.get('par_case')
        if is_num(par_case) and par_case > 0:
            par = par_case * pack
        elif csv_num(code, 'kitei') > 0:
            par = csv_num(code, 'kitei')
        else:
            par = None
        stock = sum(x['total'] for x in mine['cur'] if x['code'] == code) if in_cur else 0
        if kind is None:
            checked += 1
            continue                                   # 未仕分けはどのシートにも出ない
        seq[kind] += 1
        o = wb[f'②{kind}']
        r = ORD_FIRST + seq[kind] - 1
        cell = lambda n: o.cell(r, COL[n]).value       # noqa: E731
        counted, counted_date = cell('counted'), as_date(cell('counted_date'))
        if is_num(counted) and isinstance(counted_date, datetime.date) and counted_date >= base:
            stock = counted
        ins = [max(0, num(o.cell(r, IN_FIRST + d).value)) for d in range(DAYS)]
        start = max(0, stock)
        use = [max(0, pi) * max(0, a) / 100 if pi is not None else 0 for a in day_att]
        b, a, ov, valid, gd = [], [], [], [], []
        pb = pa = start
        for d in range(DAYS):
            units = ins[d] * pack
            over = 0
            if par is not None and ins[d] > 0:
                over = math.ceil(max(0, pa + units - par) / pack - 1e-9)
            if par is None:
                cap = 9999
            elif par - pb <= 0:
                cap = 0
            else:
                cap = math.ceil((par - pb) / pack - 1e-9)
            gd.append(0 if not flag(kind, d) else cap)
            pb = max(0, pb - use[d])
            pa = max(0, pa + units - use[d])
            b.append(pb)
            a.append(pa)
            ov.append(over)
            valid.append(1 if (ins[d] > 0 and flag(kind, d) and over == 0) else 0)
        n_in = sum(1 for x in ins if x > 0)
        in_days = [days[d] for d in range(DAYS) if ins[d] > 0]
        first_in = min(in_days) if in_days else None
        last_in = max(in_days) if in_days else None
        n_before = sum(1 for x in b if x >= 1) if pi is not None else None
        cut_before = base + datetime.timedelta(days=n_before) if n_before is not None and n_before < DAYS else None
        post_idx = [d for d in range(DAYS) if a[d] < 1 and (last_in is None or days[d] >= last_in)]
        n_post = min(post_idx) if pi is not None and post_idx else None
        cut_post = base + datetime.timedelta(days=n_post) if n_post is not None else None
        nd = [days[d] for d in range(NDT) if days[d] >= base + datetime.timedelta(days=lt[kind]) and flag(kind, d)]
        next_del = min(nd) if nd else None
        lo = [days[d] for d in range(DAYS) if cut_before is not None and days[d] <= cut_before and flag(kind, d)]
        last_ok = max(lo) if lo else None
        deadline = last_ok - datetime.timedelta(days=lt[kind]) if last_ok else None
        lo2 = [days[d] for d in range(DAYS) if cut_post is not None and days[d] <= cut_post and flag(kind, d)]
        last_ok2 = max(lo2) if lo2 else None
        deadline2 = last_ok2 - datetime.timedelta(days=lt[kind]) if last_ok2 else None
        guide = None
        if par is not None and next_del is not None and next_del in days[:DAYS]:
            guide = gd[days.index(next_del)]
        bad_day = min([days[d] for d in range(DAYS) if ins[d] > 0 and not flag(kind, d)], default=None)
        over_day = min([days[d] for d in range(DAYS) if ov[d] > 0], default=None)
        n_valid = sum(valid)
        amt = sum(valid[d] * ins[d] for d in range(DAYS)) * pack * price
        amount[kind] += amt

        if not in_cur:
            sym = '－'
        elif stock <= 0:
            sym = '●'
        elif pi is None or att_sum == 0:
            sym = '？'
        elif cut_before is None:
            sym = '○'
        elif deadline is None or deadline <= base:
            sym = '●'
        else:
            sym = '△'
        if n_in == 0:
            after = ''
        elif bad_day or over_day:
            after = '⚠'
        elif pi is None or att_sum == 0:
            after = '？'
        elif cut_before is not None and first_in > cut_before:
            earlier = next_del is not None and next_del < first_in and next_del <= cut_before
            after = '●' if earlier else '△'
        elif cut_post is None:
            after = '○'
        elif deadline2 is None or deadline2 <= base:
            after = '●'
        else:
            after = '△'

        got = {k: cell(k) for k in ('name', 'par', 'stock', 'guide', 'code')}
        got['cut'] = as_date(cell('cut'))
        got['next'] = as_date(cell('next'))
        got['status'] = cell('status') or ''
        got['after'] = cell('after') or ''
        got['grid'] = [o.cell(r, PJ_FIRST + d).value for d in range(DAYS)]

        def bad(what, exp, act):
            problems.append(f'②{kind} 行{r} {code} {name}: {what} 期待={exp!r} 実際={act!r}')

        if str(got['code']) != code:
            bad('商品コード', code, got['code'])
        if got['name'] != name:
            bad('商品名', name, got['name'])
        if abs(num(got['stock']) - stock) > 1e-6:
            bad('在庫', stock, got['stock'])
        if par is None:
            if got['par'] not in (None, ''):
                bad('定数', '', got['par'])
        elif abs(num(got['par']) - par) > 1e-6:
            bad('定数', par, got['par'])
        if got['cut'] != cut_before:
            bad('切れる日', cut_before, got['cut'])
        if got['next'] != next_del:
            bad('次の便', next_del, got['next'])
        if not got['status'].startswith(sym):
            bad('このままだと', sym, got['status'])
        if sym in '●△' and stock > 0 and cut_before is not None and cut_before.strftime('%-m/%-d') not in got['status']:
            bad('このままだと の日付', cut_before, got['status'])
        if after == '':
            if got['after'] != '':
                bad('入庫を入れると', '', got['after'])
        elif not got['after'].startswith(after):
            bad('入庫を入れると', after, got['after'])
        if guide is None:
            if got['guide'] not in (None, ''):
                bad('目安', '', got['guide'])
        elif num(got['guide'], -1) != guide:
            bad('目安', guide, got['guide'])
        if pi is None:
            if any(x not in (None, '') for x in got['grid']):
                bad('見込み(減り数未設定)', '空', got['grid'])
        else:
            for d in range(DAYS):
                if abs(num(got['grid'][d]) - a[d]) > 1e-6:
                    bad(f'見込み {d}日目', round(a[d], 3), got['grid'][d])
                    break
        for d in range(DAYS):
            if valid[d]:
                sheet_rows.append((vendor, kind, code, days[d], ins[d], ins[d] * pack * price))
        checked += 1

    # ---- ③発注書 ----
    sh = wb['③発注書']
    sel_v = sh['C4'].value or 'すべて'
    sel_k = sh['C5'].value or 'すべて'
    exp_rows = [x for x in sheet_rows if (sel_v in ('すべて', '') or x[0] == sel_v) and (sel_k in ('すべて', '') or x[1] == sel_k)]
    got_rows = []
    for r in range(11, 111):
        if sh.cell(r, 2).value in (None, ''):
            continue
        got_rows.append((sh.cell(r, 3).value, sh.cell(r, 4).value, str(sh.cell(r, 5).value), as_date(sh.cell(r, 7).value),
                         sh.cell(r, 9).value, sh.cell(r, 11).value))
    if len(exp_rows) != len(got_rows):
        problems.append(f'③発注書 行数 期待={len(exp_rows)} 実際={len(got_rows)}')
    else:
        for e, g in zip(exp_rows, got_rows):
            if e[:5] != g[:5] or abs(e[5] - num(g[5])) > 0.5:
                problems.append(f'③発注書 明細 期待={e} 実際={g}')
                break
    total_exp = sum(x[5] for x in exp_rows)
    if abs(total_exp - num(sh['K9'].value)) > 0.5:
        problems.append(f'③発注書 合計 期待={total_exp} 実際={sh["K9"].value}')
    for k in KINDS:
        o = wb[f'②{k}']
        if abs(amount[k] - num(o['L5'].value)) > 0.5:
            problems.append(f'②{k} 発注予定額 期待={amount[k]} 実際={o["L5"].value}')
        crit_exp = sum(1 for r in range(ORD_FIRST, ORD_FIRST + 400) if str(o.cell(r, COL['status']).value or '').startswith('●'))
        if num(o['C5'].value) != crit_exp:
            problems.append(f'②{k} まとめ ● 期待={crit_exp} 実際={o["C5"].value}')

    print(f'{path}: 商品 {checked} 件（{", ".join(f"{k} {seq[k]}" for k in KINDS)}）を突き合わせ　'
          f'動員合計 {att_sum:.0f}　発注書 {len(got_rows)} 行　問題 {len(problems)} 件')
    for p in problems[:30]:
        print('  ', p)
    return len(problems)


if __name__ == '__main__':
    sys.exit(1 if main(sys.argv[1] if len(sys.argv) > 1 else '売店発注ツール_サンプルデータ入り.xlsx') else 0)
