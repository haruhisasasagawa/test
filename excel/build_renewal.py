"""
シン・コンセ発注ツール（刷新版）の計算エンジンを組み立てる。

    python3 excel/build_renewal.py [出力パス]

現行版のリバースエンジニアリング設計書に沿って、業務ロジックを踏襲する。
    販売数見込 ＝ 調整後動員 × 算出購買率
    使用見込   ＝ ROUNDUP(販売数見込 ÷ イールド, 0)
    必要数     ＝ (ストック設定 ＋ 使用見込) − 理論在庫（または実在庫換算）
    発注目安   ＝ ROUNDUP(必要数 ÷ 発注ロット, 0)

現行との違いは「持ち方」だけにする。
    ・商品マスタの ##1〜##150 の横持ち（323列）は、減算マスタの縦持ちに転換した
    ・サイズ換算比率は2シートへのベタ書きをやめ、単一のマスタに集約した
    ・VLOOKUP の列番号を数式で管理するのをやめ、INDEX/MATCH の直接参照にした

見た目とダッシュボードは既存の build_workbook.py と同じ言語を使う。
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from build_workbook import (  # noqa: E402
    BORDER, C_CRIT, C_CRIT_SOFT, C_INK, C_OK, C_OK_SOFT, C_TEXT2, C_TEXT3,
    C_WARN, C_WARN_SOFT, FILL_ACCENT, FILL_AUTO, FILL_HEAD, FILL_INPUT,
    FMT_DATE, FMT_INT, FMT_YEN, FONT, F_BOLD, F_HEAD, F_INPUT, F_NOTE, F_TITLE,
    add_validation, put_headers, sheet_title, style_row,
)

FMT_RATE = '0.0000'                      # 購買率は小数4桁まで見せる
FMT_DEC1 = '#,##0.0;[Red]\\-#,##0.0'

# ---------------------------------------------------------------------------
# シートと範囲
# ---------------------------------------------------------------------------

S_INTRO = 'はじめに'
S_TOP = 'TOP'
S_STOCK = '①在庫CSV_発注日'
S_PA = '②実績CSV_期間A'
S_PB = '③実績CSV_期間B'
S_PC = '④実績CSV_期間C'
S_YIELD = '⑤減算マスタCSV'
S_LOOKUP = '⑥商品照会CSV'
S_PLAN = '⑦動員計画'
S_RATE = '購買率算出'
S_ITEM = '商品マスタ'
S_CALC = '⑧発注計算'
S_CHECK = '取込チェック'
S_SIZE = 'サイズ換算マスタ'
S_THEATER = '劇場マスタ'


def q(name):
    """数式で使うシート名。丸数字を含むので常にクォートする。"""
    return f"'{name}'"


Q_TOP, Q_STOCK, Q_PA, Q_PB, Q_PC = map(q, (S_TOP, S_STOCK, S_PA, S_PB, S_PC))
Q_YIELD, Q_LOOKUP, Q_PLAN, Q_RATE = map(q, (S_YIELD, S_LOOKUP, S_PLAN, S_RATE))
Q_ITEM, Q_CALC, Q_SIZE, Q_THEATER = map(q, (S_ITEM, S_CALC, S_SIZE, S_THEATER))

# HELLO『売上・在庫・原価出力』CSV のレイアウト（設計書 §3.1）
HELLO_COLS = [
    ('A', 'タイトル'), ('B', '劇場コード'), ('C', '劇場名'),
    ('D', '対象期間開始'), ('E', '対象期間終了'),
    ('F', '大分類コード'), ('G', '小分類コード'), ('H', '分類名'),
    ('I', '作品コード'), ('J', '作品名'), ('K', '支払先コード'), ('L', '支払先名'),
    ('M', '商品コード'), ('N', '商品名'), ('O', '買取区分'), ('P', '買取区分名'),
    ('Q', '前日残数'), ('R', '仕入数'), ('S', '移動入庫数'), ('T', '返品数'),
    ('U', '廃棄数'), ('V', '移動出庫数'), ('W', '差異数'), ('X', '資産廃棄数'),
    ('Y', '減算数'), ('Z', '残数（理論在庫）'), ('AA', '売上数（販売数）'),
    ('AB', '売上金額'), ('AC', '売上原価'), ('AD', '仕入先商品コード'),
]
HELLO_FIRST, HELLO_LAST = 2, 5001        # 1劇場あたり約800商品を想定

# 『在庫減算マスタ一覧』CSV（設計書 §3.2）
YIELD_COLS = [
    ('A', '販売商品コード'), ('B', '販売商品名'), ('C', '適用開始日'),
    ('D', '劇場コード'), ('E', '劇場名'), ('F', '減算商品コード'),
    ('G', '減算商品名'), ('H', '必要販売数量（イールド）'), ('I', '使用量'),
]
YIELD_FIRST, YIELD_LAST = 2, 8001

# 『商品一覧照会』CSV（設計書 §3.3）
LOOKUP_COLS = [
    ('A', 'SYOUHIN_CD'), ('B', '（予備）'), ('C', 'SYOUHIN_NM'),
    ('M', 'IRISU（発注ロット）'), ('Q', 'SIIRE_TANKA'), ('S', 'TORIATUKAI_KBN'),
    ('AK', 'IRISU_PACK'), ('AL', 'TANAOROSI_CASE'), ('AM', 'TANAOROSI_PACK'),
    ('AN', 'TANAOROSI_BARA'),
]
LOOKUP_FIRST, LOOKUP_LAST = 2, 3001

RATE_FIRST, RATE_LAST = 6, 3005          # 購買率算出（販売商品）
ITEM_FIRST, ITEM_LAST = 6, 3005          # 商品マスタ（在庫商品）
CALC_FIRST, CALC_LAST = 7, 3006          # 発注計算
PLAN_FIRST, PLAN_LAST = 8, 28            # 動員計画（21日）
THEATER_FIRST, THEATER_LAST = 6, 105

SIZE_RATES = [('Ｓ', 0.88), ('Ｍ', 1.0), ('Ｌ', 1.66)]
TEMP_KINDS = ['常温', '冷凍', '飲料']
STOCK_MODES = ['理論', '実数']


def rng(sheet_q, letter, first, last):
    return f'{sheet_q}!${letter}${first}:${letter}${last}'


def hello(sheet_q, letter):
    return rng(sheet_q, letter, HELLO_FIRST, HELLO_LAST)


def yld(letter):
    return rng(Q_YIELD, letter, YIELD_FIRST, YIELD_LAST)


def look(letter):
    return rng(Q_LOOKUP, letter, LOOKUP_FIRST, LOOKUP_LAST)


def rate(letter):
    return rng(Q_RATE, letter, RATE_FIRST, RATE_LAST)


def item(letter):
    return rng(Q_ITEM, letter, ITEM_FIRST, ITEM_LAST)


# ---------------------------------------------------------------------------
# 入力シート
# ---------------------------------------------------------------------------

def build_hello_sheet(wb, name, label, note):
    ws = wb.create_sheet(name)
    sheet_title(ws, label, note + '　データ行を A2 セルから貼り付けてください（見出し行は不要です）。')
    for letter, title in HELLO_COLS:
        c = ws[f'{letter}1']
        c.value = title
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for letter in ('B', 'M'):
        for r in range(HELLO_FIRST, HELLO_LAST + 1):
            ws[f'{letter}{r}'].number_format = '@'   # 先頭ゼロを落とさない
    for letter, width in (('A', 14), ('B', 11), ('C', 10), ('L', 22),
                          ('M', 16), ('N', 30), ('Z', 12), ('AA', 12)):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = 'A2'
    return ws


def build_yield_sheet(wb):
    ws = wb.create_sheet(S_YIELD)
    sheet_title(ws, '⑤減算マスタCSV（イールド）',
                'HELLO『在庫減算マスタ一覧』のデータ行を A2 セルから貼り付けてください。'
                '販売商品1つが、どの在庫商品をいくつ減らすかの定義です。')
    for letter, title in YIELD_COLS:
        c = ws[f'{letter}1']
        c.value = title
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for letter, title in (('K', '適用'), ('L', '販売商品の購買率')):
        c = ws[f'{letter}1']
        c.value = title
        c.font = F_HEAD
        c.fill = FILL_ACCENT
        c.border = BORDER
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for r in range(YIELD_FIRST, YIELD_LAST + 1):
        # 劇場コード0は全劇場向けの定義。自劇場ぶんとあわせて有効にする
        ws[f'K{r}'] = (f'=IF($A{r}="",0,IF(OR(N($D{r})=0,$D{r}&""={Q_TOP}!$C$5&""),1,0))')
        # 縦持ちの肝。販売商品の購買率をこの行に持たせ、減算商品ごとにSUMIFSで集計する
        ws[f'L{r}'] = (f'=IF($K{r}=0,0,IFERROR(INDEX({rate("P")},'
                       f'MATCH($A{r}&"",{rate("C")},0)),0))')
        ws[f'K{r}'].font = F_NOTE
        ws[f'L{r}'].font = F_NOTE
        ws[f'L{r}'].number_format = FMT_RATE
        for letter in ('A', 'D', 'F'):
            ws[f'{letter}{r}'].number_format = '@'
    for letter, width in (('A', 16), ('B', 30), ('C', 12), ('D', 11), ('E', 12),
                          ('F', 16), ('G', 30), ('H', 20), ('I', 10), ('K', 7), ('L', 16)):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = 'A2'
    return ws


def build_lookup_sheet(wb):
    ws = wb.create_sheet(S_LOOKUP)
    sheet_title(ws, '⑥商品照会CSV',
                'HELLO『商品一覧照会』のデータ行を A2 セルから貼り付けてください。'
                '発注ロット・仕入単価・パック入り数の出どころです。')
    for letter, title in LOOKUP_COLS:
        c = ws[f'{letter}1']
        c.value = title
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for r in range(LOOKUP_FIRST, LOOKUP_LAST + 1):
        ws[f'A{r}'].number_format = '@'
    for letter, width in (('A', 16), ('C', 30), ('M', 16), ('Q', 14), ('AK', 14)):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = 'A2'
    return ws


# ---------------------------------------------------------------------------
# マスタ
# ---------------------------------------------------------------------------

def build_size_master(wb):
    ws = wb.create_sheet(S_SIZE)
    sheet_title(ws, 'サイズ換算マスタ',
                '購買率をM換算するための比率です。現行版では2シートにベタ書きされていました。')
    put_headers(ws, 5, ['サイズ', '換算比率', '備考'], ['入力', '入力', '入力'])
    notes = ['Sサイズ', 'Mサイズ（基準）', 'Lサイズ']
    for i, (size, ratio) in enumerate(SIZE_RATES):
        r = 6 + i
        ws[f'B{r}'] = size
        ws[f'C{r}'] = ratio
        ws[f'D{r}'] = notes[i]
        style_row(ws, r, 'BCD', fill=FILL_INPUT)
        ws[f'C{r}'].number_format = '0.00'
    ws['B10'] = '※ 商品名の末尾1文字をサイズとみなします。該当しない商品は比率1として扱います。'
    ws['B10'].font = F_NOTE
    for c, w in zip('BCD', [10, 12, 30]):
        ws.column_dimensions[c].width = w
    return ws


def build_theater_master(wb):
    ws = wb.create_sheet(S_THEATER)
    sheet_title(ws, '劇場マスタ', '全劇場共通のブックとして配布し、TOPで自劇場を選びます。')
    headers = ['劇場コード', '劇場名', '劇場タイプ', '規模区分', 'ドリンク提供方式', '備考']
    put_headers(ws, 5, headers, ['入力', '入力', '選択', '選択', '選択', '入力'])
    for r in range(THEATER_FIRST, THEATER_LAST + 1):
        style_row(ws, r, 'BCDEFG', fill=FILL_INPUT)
        ws[f'B{r}'].number_format = '@'
    ws['B6'], ws['C6'], ws['D6'] = '0761', '新宿', '通常劇場'
    ws['E6'], ws['F6'] = '大規模', '1杯売り'
    add_validation(ws, '"通常劇場,前田農産,ライスト"', f'D{THEATER_FIRST}:D{THEATER_LAST}')
    add_validation(ws, '"大規模,中規模,小規模"', f'E{THEATER_FIRST}:E{THEATER_LAST}')
    add_validation(ws, '"1杯売り,ドリンクバー"', f'F{THEATER_FIRST}:F{THEATER_LAST}')
    for c, w in zip('BCDEFG', [12, 22, 14, 12, 18, 34]):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = 'B6'
    return ws


# ---------------------------------------------------------------------------
# TOP（設定）
# ---------------------------------------------------------------------------

def build_top(wb):
    ws = wb.create_sheet(S_TOP)
    sheet_title(ws, 'TOP（設定）', '黄色のセルを入力・選択してください。')
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 66

    plan_att = rng(Q_PLAN, 'D', PLAN_FIRST, PLAN_LAST)   # 生の日別見込み
    rows = [
        ('自劇場コード', None, '劇場マスタから選びます。CSVの劇場コードと一致させてください。',
         FILL_INPUT, '@'),
        ('劇場名', f'=IFERROR(INDEX({rng(Q_THEATER, "C", THEATER_FIRST, THEATER_LAST)},'
                  f'MATCH($C$5,{rng(Q_THEATER, "B", THEATER_FIRST, THEATER_LAST)},0)),"")',
         '劇場マスタから自動取得します。', FILL_AUTO, None),
        ('劇場タイプ', f'=IFERROR(INDEX({rng(Q_THEATER, "D", THEATER_FIRST, THEATER_LAST)},'
                     f'MATCH($C$5,{rng(Q_THEATER, "B", THEATER_FIRST, THEATER_LAST)},0)),"")',
         'ポップコーン豆の商品コードを切り替えます（通常劇場／前田農産／ライスト）。',
         FILL_AUTO, None),
        ('発注日', None, '発注する日です。動員計画の起点になります。', FILL_INPUT, FMT_DATE),
        ('分納日数', 11, '冷凍品・飲料品の分納スケジュールの日数です（現行は11日固定）。',
         FILL_INPUT, FMT_INT),
        ('参照購買率', None, '購買率をどの期間の実績から取るかを選びます。', FILL_INPUT, None),
        ('動員見込み調整値', 1.2, '日別動員見込みに掛ける係数です（現行の既定は120%）。',
         FILL_INPUT, '0%'),
        ('【期間A】実動員', None, '予算実績管理表の 有料＋無料 の合計です。', FILL_INPUT, FMT_INT),
        ('【期間B】実動員', None, '同上。', FILL_INPUT, FMT_INT),
        ('【期間C】実動員', None, '同上。', FILL_INPUT, FMT_INT),
        ('期間合計動員', f'=SUM({plan_att})',
         '動員計画の日別見込みの合計です（調整値を掛ける前）。', FILL_AUTO, FMT_INT),
        ('調整後動員', '=ROUND($C$15*$C$11,0)',
         'すべての発注計算の需要ドライバです（期間合計動員 × 調整値）。', FILL_AUTO, FMT_INT),
        ('発注担当者', None, '発注書に記載されます。', FILL_INPUT, None),
    ]
    for i, (label, value, note, fill, fmt) in enumerate(rows):
        r = 5 + i
        ws[f'B{r}'] = label
        ws[f'B{r}'].font = F_BOLD
        ws[f'B{r}'].fill = FILL_HEAD
        ws[f'B{r}'].border = BORDER
        c = ws[f'C{r}']
        if value is not None:
            c.value = value
        c.fill = fill
        c.border = BORDER
        c.font = F_INPUT if fill is FILL_INPUT else Font(name=FONT, size=11, color=C_INK)
        if fmt:
            c.number_format = fmt
        ws[f'D{r}'] = note
        ws[f'D{r}'].font = F_NOTE

    ws['C5'] = '0761'
    ws['C10'] = '【期間A】'
    add_validation(ws, f'={rng(Q_THEATER, "B", THEATER_FIRST, THEATER_LAST)}', 'C5')
    add_validation(ws, '"【期間A】,【期間B】,【期間C】"', 'C10')
    add_validation(ws, '"1,1.1,1.2,1.3,1.4,1.5"', 'C11')

    ws['B19'] = 'ポップコーン豆の商品コード（劇場タイプ別）'
    ws['B19'].font = F_BOLD
    beans = [('通常劇場', '1812000007106'), ('前田農産', '1812000011561'),
             ('ライスト', '1812000011387')]
    for i, (kind, code) in enumerate(beans):
        r = 20 + i
        ws[f'B{r}'] = kind
        ws[f'B{r}'].font = F_NOTE
        ws[f'C{r}'] = code
        ws[f'C{r}'].number_format = '@'
        ws[f'C{r}'].font = F_NOTE
        style_row(ws, r, 'BC', fill=FILL_INPUT)
    ws['C24'] = f'=IFERROR(INDEX($C$20:$C$22,MATCH($C$7,$B$20:$B$22,0)),"")'
    ws['C24'].number_format = '@'
    ws['B24'] = '採用する豆の商品コード'
    ws['B24'].font = F_BOLD
    ws['B24'].fill = FILL_HEAD
    ws['B24'].border = BORDER
    ws['C24'].fill = FILL_AUTO
    ws['C24'].border = BORDER
    return ws


# ---------------------------------------------------------------------------
# 動員計画
# ---------------------------------------------------------------------------

def build_plan(wb):
    ws = wb.create_sheet(S_PLAN)
    sheet_title(ws, '⑦動員計画',
                '発注日からの日別動員見込みです。ここが来場者予測の唯一の入力場所です。')
    put_headers(ws, PLAN_FIRST - 1,
                ['日付', '曜日', '日別動員見込み', '主な作品・備考', '調整後動員'],
                ['自動', '自動', '入力', '入力', '自動'])
    for r in range(PLAN_FIRST, PLAN_LAST + 1):
        offset = r - PLAN_FIRST
        ws[f'B{r}'] = (f'=IF({Q_TOP}!$C$8="","",IF({offset}<{Q_TOP}!$C$9,'
                       f'{Q_TOP}!$C$8+{offset},""))')
        ws[f'C{r}'] = (f'=IF($B{r}="","",'
                       f'CHOOSE(WEEKDAY($B{r}),"日","月","火","水","木","金","土"))')
        ws[f'F{r}'] = f'=IF($B{r}="","",ROUND(N($D{r})*N({Q_TOP}!$C$11),0))'
        style_row(ws, r, 'BCDEF', fill=FILL_AUTO)
        ws[f'D{r}'].fill = FILL_INPUT
        ws[f'D{r}'].font = F_INPUT
        ws[f'E{r}'].fill = FILL_INPUT
        ws[f'E{r}'].font = F_INPUT
        ws[f'B{r}'].number_format = FMT_DATE
        ws[f'D{r}'].number_format = FMT_INT
        ws[f'F{r}'].number_format = FMT_INT
        ws[f'C{r}'].alignment = Alignment(horizontal='center')
    ws.conditional_formatting.add(
        f'B{PLAN_FIRST}:F{PLAN_LAST}',
        FormulaRule(formula=[f'AND($B{PLAN_FIRST}<>"",'
                             f'OR(WEEKDAY($B{PLAN_FIRST})=1,WEEKDAY($B{PLAN_FIRST})=7))'],
                    fill=PatternFill('solid', bgColor='FFF4EFE2'), stopIfTrue=False))
    for c, w in zip('BCDEF', [14, 6, 16, 40, 14]):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = f'B{PLAN_FIRST}'
    return ws


# ---------------------------------------------------------------------------
# 購買率算出
# ---------------------------------------------------------------------------

def build_rate(wb):
    ws = wb.create_sheet(S_RATE)
    sheet_title(ws, '購買率算出',
                '販売商品ごとの購買率です。期間A/B/Cの実績CSVと実動員から自動で出ます。')
    headers = ['No', '販売商品コード', '販売商品名', 'サイズ', '換算比率',
               '期間A 販売数', '期間B 販売数', '期間C 販売数',
               'M換算A', 'M換算B', 'M換算C',
               '購買率A', '購買率B', '購買率C', '採用購買率']
    put_headers(ws, 5, headers, ['自動'] * len(headers))

    size_letters = rng(Q_SIZE, 'B', 6, 8)
    size_ratio = rng(Q_SIZE, 'C', 6, 8)
    letters = [get_column_letter(i) for i in range(2, 2 + len(headers))]

    for r in range(RATE_FIRST, RATE_LAST + 1):
        src = HELLO_FIRST + (r - RATE_FIRST)
        ws[f'B{r}'] = r - RATE_FIRST + 1
        ws[f'C{r}'] = f'=IF({Q_PA}!$M{src}="","",{Q_PA}!$M{src}&"")'
        ws[f'D{r}'] = f'=IF($C{r}="","",{Q_PA}!$N{src})'
        ws[f'E{r}'] = f'=IF($C{r}="","",RIGHT($D{r},1))'
        ws[f'F{r}'] = (f'=IF($C{r}="","",IFERROR(INDEX({size_ratio},'
                       f'MATCH($E{r},{size_letters},0)),1))')
        for out, sheet_q in (('G', Q_PA), ('H', Q_PB), ('I', Q_PC)):
            ws[f'{out}{r}'] = (f'=IF($C{r}="","",SUMIFS({hello(sheet_q, "AA")},'
                               f'{hello(sheet_q, "M")},$C{r}))')
        for out, src_col in (('J', 'G'), ('K', 'H'), ('L', 'I')):
            ws[f'{out}{r}'] = f'=IF($C{r}="","",N($F{r})*N(${src_col}{r}))'
        for out, src_col, att in (('M', 'J', '$C$12'), ('N', 'K', '$C$13'), ('O', 'L', '$C$14')):
            ws[f'{out}{r}'] = (f'=IF(OR($C{r}="",N({Q_TOP}!{att})=0),"",'
                               f'N(${src_col}{r})/{Q_TOP}!{att})')
        ws[f'P{r}'] = (f'=IF($C{r}="","",IFERROR(CHOOSE(MATCH({Q_TOP}!$C$10,'
                       f'{{"【期間A】","【期間B】","【期間C】"}},0),N($M{r}),N($N{r}),N($O{r})),0))')
        style_row(ws, r, letters, fill=FILL_AUTO)
        ws[f'C{r}'].number_format = '@'
        ws[f'F{r}'].number_format = '0.00'
        for c in 'GHIJKL':
            ws[f'{c}{r}'].number_format = FMT_DEC1
        for c in 'MNOP':
            ws[f'{c}{r}'].number_format = FMT_RATE
        ws[f'P{r}'].font = F_BOLD

    widths = [5, 16, 30, 7, 9, 13, 13, 13, 11, 11, 11, 11, 11, 11, 12]
    for c, w in zip(letters, widths):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = f'D{RATE_FIRST}'
    return ws


# ---------------------------------------------------------------------------
# 商品マスタ（在庫商品）
# ---------------------------------------------------------------------------

def build_item_master(wb):
    ws = wb.create_sheet(S_ITEM)
    sheet_title(ws, '商品マスタ（在庫商品）',
                '発注日在庫CSVの商品を並べ、購買率・イールド・発注属性をまとめます。'
                '黄色の2列だけが手入力です。')
    headers = ['No', '商品コード', '商品名', '分類名', '支払先名', '温度区分',
               '理論在庫', '算出購買率', 'イールド', '発注ロット', '仕入単価',
               '発注単価', 'パック入り数', '棚卸単位', '設定', 'ストック(ケース)',
               'ストック設定']
    kinds = ['自動', '自動', '自動', '自動', '自動', '入力', '自動', '自動', '自動',
             '自動', '自動', '自動', '自動', '自動', '入力', '入力', '自動']
    put_headers(ws, 5, headers, kinds)
    letters = [get_column_letter(i) for i in range(2, 2 + len(headers))]

    for r in range(ITEM_FIRST, ITEM_LAST + 1):
        src = HELLO_FIRST + (r - ITEM_FIRST)
        ws[f'B{r}'] = r - ITEM_FIRST + 1
        ws[f'C{r}'] = f'=IF({Q_STOCK}!$M{src}="","",{Q_STOCK}!$M{src}&"")'
        ws[f'D{r}'] = f'=IF($C{r}="","",{Q_STOCK}!$N{src})'
        ws[f'E{r}'] = f'=IF($C{r}="","",{Q_STOCK}!$H{src})'
        ws[f'F{r}'] = f'=IF($C{r}="","",{Q_STOCK}!$L{src})'
        ws[f'H{r}'] = f'=IF($C{r}="","",N({Q_STOCK}!$Z{src}))'
        # 縦持ちの減算マスタから、この在庫商品を減らす販売商品の購買率を合計する
        ws[f'I{r}'] = (f'=IF($C{r}="","",SUMIFS({yld("L")},{yld("F")},$C{r},{yld("K")},1))')
        ws[f'J{r}'] = (f'=IF($C{r}="","",IFERROR(INDEX({yld("H")},'
                       f'MATCH($C{r},{yld("F")},0)),1))')
        ws[f'K{r}'] = (f'=IF($C{r}="","",MAX(1,IFERROR(INDEX({look("M")},'
                       f'MATCH($C{r},{look("A")},0)),1)))')
        ws[f'L{r}'] = (f'=IF($C{r}="","",IFERROR(INDEX({look("Q")},'
                       f'MATCH($C{r},{look("A")},0)),0))')
        ws[f'M{r}'] = f'=IF($C{r}="","",N($L{r})*N($K{r}))'
        ws[f'N{r}'] = (f'=IF($C{r}="","",MAX(1,IFERROR(INDEX({look("AK")},'
                       f'MATCH($C{r},{look("A")},0)),1)))')
        ws[f'O{r}'] = (f'=IF($C{r}="","",IFERROR(IF(INDEX({look("AN")},'
                       f'MATCH($C{r},{look("A")},0))=1,"バラ",'
                       f'IF(INDEX({look("AM")},MATCH($C{r},{look("A")},0))=1,'
                       f'"中間包装","ケース")),"ケース"))')
        ws[f'R{r}'] = f'=IF($C{r}="","",N($Q{r})*N($K{r}))'
        style_row(ws, r, letters, fill=FILL_AUTO)
        for letter in ('G', 'P', 'Q'):
            ws[f'{letter}{r}'].fill = FILL_INPUT
            ws[f'{letter}{r}'].font = F_INPUT
        ws[f'C{r}'].number_format = '@'
        for c in 'HJKNQR':
            ws[f'{c}{r}'].number_format = FMT_INT
        ws[f'I{r}'].number_format = FMT_RATE
        for c in 'LM':
            ws[f'{c}{r}'].number_format = FMT_YEN

    add_validation(ws, '"' + ','.join(TEMP_KINDS) + '"', f'G{ITEM_FIRST}:G{ITEM_LAST}')
    add_validation(ws, '"' + ','.join(STOCK_MODES) + '"', f'P{ITEM_FIRST}:P{ITEM_LAST}')
    widths = [5, 16, 30, 16, 22, 10, 11, 12, 10, 11, 11, 12, 13, 11, 8, 14, 13]
    for c, w in zip(letters, widths):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = f'D{ITEM_FIRST}'
    return ws


# ---------------------------------------------------------------------------
# 発注計算
# ---------------------------------------------------------------------------

def build_calc(wb):
    ws = wb.create_sheet(S_CALC)
    sheet_title(ws, '⑧発注計算',
                '現行版の算出式をそのまま実装しています。'
                '黄色の「発注数」と「カウント数」だけが入力欄です。')
    ws['B4'] = ('販売数見込 ＝ 調整後動員 × 算出購買率　／　使用見込 ＝ ROUNDUP(販売数見込 ÷ イールド, 0)　／　'
                '必要数 ＝ (ストック設定 ＋ 使用見込) − 在庫　／　発注目安 ＝ ROUNDUP(必要数 ÷ 発注ロット, 0)')
    ws['B4'].font = F_NOTE

    headers = ['No', '商品コード', '商品名', '温度区分', '設定', '理論在庫',
               'カウント数', '実在庫換算', '理論差', '算出購買率', 'イールド',
               '販売数見込', '使用見込', 'ストック設定', '必要数', '発注ロット',
               '発注目安', '発注数', '発注単価', '金額', '備考']
    kinds = ['自動', '自動', '自動', '自動', '自動', '自動', '入力', '自動', '自動',
             '自動', '自動', '自動', '自動', '自動', '自動', '自動', '自動', '入力',
             '自動', '自動', '入力']
    put_headers(ws, 6, headers, kinds)
    letters = [get_column_letter(i) for i in range(2, 2 + len(headers))]

    for r in range(CALC_FIRST, CALC_LAST + 1):
        m = ITEM_FIRST + (r - CALC_FIRST)
        ws[f'B{r}'] = r - CALC_FIRST + 1
        ws[f'C{r}'] = f'=IF({Q_ITEM}!$C{m}="","",{Q_ITEM}!$C{m})'
        ws[f'D{r}'] = f'=IF($C{r}="","",{Q_ITEM}!$D{m})'
        ws[f'E{r}'] = f'=IF($C{r}="","",{Q_ITEM}!$G{m})'
        ws[f'F{r}'] = f'=IF($C{r}="","",IF({Q_ITEM}!$P{m}="","理論",{Q_ITEM}!$P{m}))'
        ws[f'G{r}'] = f'=IF($C{r}="","",N({Q_ITEM}!$H{m}))'
        ws[f'I{r}'] = f'=IF($C{r}="","",N($H{r})*N({Q_ITEM}!$N{m}))'
        ws[f'J{r}'] = f'=IF($C{r}="","",N($I{r})-N($G{r}))'
        ws[f'K{r}'] = f'=IF($C{r}="","",N({Q_ITEM}!$I{m}))'
        ws[f'L{r}'] = f'=IF($C{r}="","",MAX(1,N({Q_ITEM}!$J{m})))'
        ws[f'M{r}'] = f'=IF($C{r}="","",N({Q_TOP}!$C$16)*N($K{r}))'
        ws[f'N{r}'] = f'=IF($C{r}="","",ROUNDUP(N($M{r})/N($L{r}),0))'
        ws[f'O{r}'] = f'=IF($C{r}="","",N({Q_ITEM}!$R{m}))'
        ws[f'P{r}'] = (f'=IF($C{r}="","",IF($F{r}="理論",'
                       f'(N($O{r})+N($N{r}))-N($G{r}),(N($O{r})+N($N{r}))-N($I{r})))')
        ws[f'Q{r}'] = f'=IF($C{r}="","",MAX(1,N({Q_ITEM}!$K{m})))'
        ws[f'R{r}'] = (f'=IF($C{r}="","",IF(N($P{r})/N($Q{r})<=0,"",'
                       f'ROUNDUP(N($P{r})/N($Q{r}),0)))')
        ws[f'T{r}'] = f'=IF($C{r}="","",N({Q_ITEM}!$M{m}))'
        ws[f'U{r}'] = f'=IF($S{r}="","",N($S{r})*N($T{r}))'
        style_row(ws, r, letters, fill=FILL_AUTO)
        for letter in ('H', 'S', 'V'):
            ws[f'{letter}{r}'].fill = FILL_INPUT
            ws[f'{letter}{r}'].font = F_INPUT
        ws[f'C{r}'].number_format = '@'
        for c in 'GHIJNOPQRS':
            ws[f'{c}{r}'].number_format = FMT_INT
        ws[f'K{r}'].number_format = FMT_RATE
        ws[f'L{r}'].number_format = FMT_INT
        ws[f'M{r}'].number_format = FMT_DEC1
        for c in 'TU':
            ws[f'{c}{r}'].number_format = FMT_YEN
        ws[f'R{r}'].font = F_BOLD

    ws.conditional_formatting.add(
        f'R{CALC_FIRST}:R{CALC_LAST}',
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill('solid', bgColor=C_WARN_SOFT)))
    ws.conditional_formatting.add(
        f'K{CALC_FIRST}:K{CALC_LAST}',
        CellIsRule(operator='equal', formula=['0'],
                   fill=PatternFill('solid', bgColor=C_CRIT_SOFT)))

    widths = [5, 16, 28, 10, 8, 11, 11, 12, 10, 12, 10, 12, 11, 13, 11, 11,
              11, 11, 12, 13, 22]
    for c, w in zip(letters, widths):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = f'D{CALC_FIRST}'
    ws.print_area = f'B2:V{CALC_LAST}'
    ws.page_setup.orientation = 'landscape'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    return ws


# ---------------------------------------------------------------------------
# 取込チェック
# ---------------------------------------------------------------------------

def build_check(wb):
    ws = wb.create_sheet(S_CHECK)
    sheet_title(ws, '取込チェック',
                'CSVを貼り替えたら、まずここを見てください。想定と違う数字が出ていたら取込ミスです。')
    put_headers(ws, 5, ['チェック項目', '結果', '判定', '意味'],
                ['自動', '自動', '自動', '自動'])

    calc_rate = rng(Q_CALC, 'K', CALC_FIRST, CALC_LAST)
    calc_lot = rng(Q_CALC, 'Q', CALC_FIRST, CALC_LAST)
    calc_code = rng(Q_CALC, 'C', CALC_FIRST, CALC_LAST)
    checks = [
        ('①在庫CSV 行数', f'=COUNTIF({hello(Q_STOCK, "M")},"?*")',
         '0なら未貼付です。'),
        ('　うち自劇場ぶん', f'=COUNTIF({hello(Q_STOCK, "B")},{Q_TOP}!$C$5)',
         '0のままなら劇場コードの型がずれています（先頭ゼロの脱落）。'),
        ('②実績CSV_期間A 行数', f'=COUNTIF({hello(Q_PA, "M")},"?*")', '購買率の元になります。'),
        ('③実績CSV_期間B 行数', f'=COUNTIF({hello(Q_PB, "M")},"?*")', '期間B購買率を使う場合に必要です。'),
        ('④実績CSV_期間C 行数', f'=COUNTIF({hello(Q_PC, "M")},"?*")', '期間C購買率を使う場合に必要です。'),
        ('⑤減算マスタ 行数', f'=COUNTIF({yld("A")},"?*")', 'イールドの元です。0なら購買率が全て0になります。'),
        ('　うち自劇場に適用', f'=SUM({yld("K")})', '劇場コード0（全劇場）と自劇場ぶんの合計です。'),
        ('⑥商品照会CSV 行数', f'=COUNTIF({look("A")},"?*")', '発注ロット・単価・パック入り数の元です。'),
        ('動員計画の入力日数', f'=COUNT({rng(Q_PLAN, "D", PLAN_FIRST, PLAN_LAST)})',
         '分納日数ぶん入っているか確認してください。'),
        ('調整後動員', f'={Q_TOP}!$C$16', 'すべての発注計算の需要ドライバです。'),
        ('発注対象の商品数', f'=COUNTIF({calc_code},"?*")', '発注計算シートの行数です。'),
        ('購買率が0の商品', f'=COUNTIFS({calc_code},"?*",{calc_rate},0)',
         '減算マスタに紐付きが無い商品です。マスタ登録漏れの可能性があります。'),
        ('発注ロットが1の商品', f'=COUNTIFS({calc_code},"?*",{calc_lot},1)',
         '商品照会CSVに無い商品です。既定値1で計算しています。'),
        ('発注目安が出た商品', f'=COUNTIF({rng(Q_CALC, "R", CALC_FIRST, CALC_LAST)},">0")',
         '今回の発注対象です。'),
    ]
    for i, (label, formula, meaning) in enumerate(checks):
        r = 6 + i
        ws[f'B{r}'] = label
        ws[f'C{r}'] = formula
        ws[f'C{r}'].number_format = FMT_INT
        ws[f'D{r}'] = f'=IF(N($C{r})>0,"OK","要確認")'
        ws[f'E{r}'] = meaning
        ws[f'E{r}'].font = F_NOTE
        style_row(ws, r, 'BCDE', fill=FILL_AUTO)
        ws[f'D{r}'].alignment = Alignment(horizontal='center')
    last = 5 + len(checks)
    # 「購買率0」「ロット1」は多いほど悪いので、判定の向きを逆にする
    for r in (17, 18):
        ws[f'D{r}'] = f'=IF(N($C{r})=0,"OK","要確認")'
    ws.conditional_formatting.add(
        f'D6:D{last}',
        CellIsRule(operator='equal', formula=['"要確認"'],
                   fill=PatternFill('solid', bgColor=C_CRIT_SOFT),
                   font=Font(name=FONT, size=11, bold=True, color=C_CRIT)))
    ws.conditional_formatting.add(
        f'D6:D{last}',
        CellIsRule(operator='equal', formula=['"OK"'],
                   fill=PatternFill('solid', bgColor=C_OK_SOFT),
                   font=Font(name=FONT, size=11, bold=True, color=C_OK)))
    for c, w in zip('BCDE', [26, 14, 10, 62]):
        ws.column_dimensions[c].width = w
    return ws


# ---------------------------------------------------------------------------
# はじめに
# ---------------------------------------------------------------------------

def build_intro(wb):
    ws = wb.create_sheet(S_INTRO)
    sheet_title(ws, 'シン・コンセ発注ツール（刷新版・計算エンジン）')
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 82

    rows = [
        ('h', '', 'このブックの位置づけ'),
        ('t', '', '現行版の業務ロジックを踏襲したまま、データの持ち方を作り直したものです。'),
        ('t', '', '受入条件である 理論残／購買率／使用見込／必要数／発注目安 の5項目を算出します。'),
        ('t', '', '分納スケジュール（冷凍・飲料）と3区分の出力シートは次の段階で載せます。'),
        ('', '', ''),
        ('h', '', '週次の流れ'),
        ('s', 'TOP', '自劇場・発注日・参照購買率・調整値・実動員を入れる'),
        ('s', '①在庫CSV_発注日', '発注日時点の理論在庫CSVを貼る'),
        ('s', '②〜④実績CSV', '期間A／B／Cの実績CSVを貼る（購買率の元）'),
        ('s', '⑤減算マスタCSV', 'イールド（販売商品→在庫商品の減算定義）を貼る'),
        ('s', '⑥商品照会CSV', '発注ロット・仕入単価・パック入り数を貼る'),
        ('s', '⑦動員計画', '発注日からの日別動員見込みを入れる'),
        ('s', '取込チェック', '貼付の抜けや型ずれが無いか確認する ★まずここ'),
        ('s', '⑧発注計算', '発注目安を見て発注数を決める ★ここで発注'),
        ('', '', ''),
        ('h', '', '算出式（現行踏襲）'),
        ('t', '', '調整後動員　 ＝ 期間合計動員 × 動員見込み調整値'),
        ('t', '', '購買率　　　 ＝ M換算販売数 ÷ 実動員　（S=0.88 / M=1 / L=1.66）'),
        ('t', '', '算出購買率　 ＝ その在庫商品を減らす販売商品の購買率の合計'),
        ('t', '', '販売数見込　 ＝ 調整後動員 × 算出購買率'),
        ('t', '', '使用見込　　 ＝ ROUNDUP(販売数見込 ÷ イールド, 0)'),
        ('t', '', '実在庫換算　 ＝ カウント数 × パック入り数'),
        ('t', '', '必要数　　　 ＝ (ストック設定 ＋ 使用見込) − 理論在庫（設定が「実数」なら実在庫換算）'),
        ('t', '', '発注目安　　 ＝ ROUNDUP(必要数 ÷ 発注ロット, 0)　※0以下は空欄'),
        ('', '', ''),
        ('h', '', '現行版から変えたところ'),
        ('t', '', '・商品マスタの ##1〜##150（323列の横持ち）を、減算マスタの縦持ちに転換しました。'),
        ('t', '', '　商品が増えても列を足す必要がなくなり、紐付けの追加はCSVの行追加だけで済みます。'),
        ('t', '', '・サイズ換算比率（S=0.88 / L=1.66）を単一のマスタに集約しました。'),
        ('t', '', '・VLOOKUPの列番号を数式で管理するのをやめ、INDEX/MATCHの直接参照にしました。'),
        ('t', '', '・貼付の抜けや型ずれを見つける「取込チェック」を用意しました。'),
        ('t', '', '・数式はINDEX/MATCH・SUMIFSなど旧来関数のみです（XLOOKUP・LET・スピルは不使用）。'),
        ('', '', ''),
        ('h', '', '確認をお願いしたいこと'),
        ('t', '', '・イールドは「在庫商品1つにつき1種類」として引いています（現行のVLOOKUP踏襲）。'),
        ('t', '', '　同じ在庫商品に複数のイールドが定義されている場合、最初の1件が使われます。'),
        ('t', '', '・温度区分（常温／冷凍／飲料）は商品マスタの手入力です。CSVから判定できるなら教えてください。'),
        ('t', '', '・分納の11日は可変にしてありますが、配送サイクル固定かどうかの確認が要ります。'),
    ]
    r = 5
    for kind, sheet, text in rows:
        if sheet:
            c = ws[f'C{r}']
            c.value = sheet
            c.font = Font(name=FONT, size=10, bold=True, color=C_INK)
            c.fill = FILL_INPUT if '★' in text else FILL_HEAD
            c.border = BORDER
            c.alignment = Alignment(horizontal='center', vertical='center')
        cell = ws[f'D{r}']
        cell.value = text
        if kind == 'h':
            cell.font = Font(name=FONT, size=12, bold=True, color=C_INK)
            cell.fill = FILL_ACCENT
        elif kind == 's':
            cell.font = F_BOLD
        else:
            cell.font = Font(name=FONT, size=10, color=C_TEXT2)
        r += 1
    return ws


# ---------------------------------------------------------------------------

def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('シン・コンセ発注ツール.xlsx')
    wb = Workbook()
    wb.remove(wb.active)

    build_intro(wb)
    build_top(wb)
    build_hello_sheet(wb, S_STOCK, '①在庫CSV（発注日）',
                      '発注日時点の理論在庫です。Z列（残数）が理論在庫になります。')
    build_hello_sheet(wb, S_PA, '②実績CSV（期間A）', '購買率を出すための販売実績です。')
    build_hello_sheet(wb, S_PB, '③実績CSV（期間B）', '同上。')
    build_hello_sheet(wb, S_PC, '④実績CSV（期間C）', '同上。')
    build_yield_sheet(wb)
    build_lookup_sheet(wb)
    build_plan(wb)
    build_rate(wb)
    build_item_master(wb)
    build_calc(wb)
    build_check(wb)
    build_size_master(wb)
    build_theater_master(wb)

    order = [S_INTRO, S_TOP, S_STOCK, S_PA, S_PB, S_PC, S_YIELD, S_LOOKUP,
             S_PLAN, S_CHECK, S_CALC, S_RATE, S_ITEM, S_SIZE, S_THEATER]
    wb._sheets = [wb[n] for n in order]
    for ws in wb.worksheets:
        if ws.title == S_CALC:
            ws.sheet_properties.tabColor = C_CRIT[2:]
        elif ws.title in (S_TOP, S_CHECK):
            ws.sheet_properties.tabColor = 'E8B44A'
        elif ws.title in (S_RATE, S_ITEM):
            ws.sheet_properties.tabColor = C_INK[2:]

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f'saved: {out}  sheets={len(wb.worksheets)}')


if __name__ == '__main__':
    main()
