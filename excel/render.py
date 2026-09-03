"""指定シートを1枚ずつPNGにする（LibreOffice経由）。見た目の確認用。
    python3 excel/render.py ホーム ②発注数を決める ...
"""
import os, shutil, subprocess, sys
from pathlib import Path
from openpyxl import load_workbook
import warnings; warnings.filterwarnings('ignore')
HERE = Path(__file__).resolve().parent
SRC = Path(os.environ.get('RENDER_SRC', HERE / '売店発注ツール_サンプルデータ入り.xlsx'))
OUT = Path('/tmp/claude-0/-home-user-test/6ebcf1c1-1102-5af1-96f9-c94aebb0a201/scratchpad/render')
AREAS = {'ホーム': 'A1:E52', '①在庫を貼る': 'A1:W20', '②冷凍': 'A1:AR40', '②飲料': 'A1:AR40', '②常温': 'A1:AR40',
         '③発注書': 'A1:K40', '仕分け': 'A1:K60', '動員を入れる': 'A1:L46', 'つかいかた': 'A1:D64',
         '商品マスタ': 'A1:O30', '劇場マスタ': 'A1:G30', '設定': 'A1:F42', '月1回_1ヶ月前の在庫': 'A1:W20'}
OUT.mkdir(parents=True, exist_ok=True)
for target in sys.argv[1:]:
    tmp = OUT / 'r.xlsx'
    shutil.copy(SRC, tmp)
    w = load_workbook(tmp)
    for ws in w.worksheets:
        if ws.title == target:
            ws.print_area = AREAS.get(target, 'A1:Z40')
            ws.page_setup.orientation = 'landscape'
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
        else:
            ws.print_area = 'A1:A1'
    w.active = w.sheetnames.index(target)
    w.save(tmp)
    subprocess.run(['soffice', '--headless', '--convert-to', 'pdf', str(tmp), '--outdir', str(OUT)],
                   capture_output=True)
    safe = target.replace('（', '').replace('）', '').replace('①', '1').replace('②', '2').replace('③', '3')
    subprocess.run(['pdftoppm', '-r', '120', '-png', '-f', '1', '-l', '1', str(OUT / 'r.pdf'), str(OUT / safe)],
                   capture_output=True)
    pngs = sorted(OUT.glob(f'{safe}*.png'))
    print(target, '->', pngs[-1] if pngs else 'FAILED')
