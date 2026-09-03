#!/usr/bin/env python3
"""
売店発注ツール（シンプル版）を生成する。

やることは4つだけ。
  1. 在庫CSVを貼る         → 商品ごとの今の在庫が入る
  2. 動員を入れる           → 商品ごとの「減り数のスピード」が決まる
  3. このままだと足りない  → 「○/○ に切れる」とアラートが出る
  4. 届く日の列に入庫数を入れる → 在庫が復活して、切れる日が延びる

  減り数/日 ＝ 100人あたりの減り数（商品マスタ） × その日の動員 ÷ 100
  在庫の予測 ＝ MAX(0, 前日の在庫 ＋ その日の入庫 − 減り数)   ※ 0 で止める

商品は「仕分け」ページで 冷凍・飲料・常温 に分け、区分ごとに ②冷凍 ②飲料 ②常温 のシートで発注する。
区分ごとに「納品日（届く曜日、最大7日）」と「発注から納品までの日数」を持つ。
入庫は 14日の表の、納品日の列に数を入れる（1日1便、最大14日ぶん）。

行の並びは商品マスタの順（CSVの順ではない）。
  入庫数は行に書くので、CSVの並びが日々変わっても入力が別の商品に付け替わらないよう、
  「商品マスタにある商品を、マスタの順に」並べ、マスタに無い新商品はその後ろに足す。
  マスタの行の挿入・削除・並べ替え、商品の区分の変更は行をずらすので、入庫の無い時に社員がやる。

社歴の浅い社員・アルバイトが触る前提なので、
  ・入力するセルは黄色だけ（入庫数・数えた在庫・動員）
  ・それ以外はシート保護でロック（パスワードなし。校閲→保護の解除で外せる）
  ・納品日でない列には入庫数が入らない。目安（定数まで）を超える数も入らない
  ・過去のCSV・定数超え・減り数未設定・動員未入力・未仕分け・納品日未設定は、文字で警告する
"""
import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

# ---------------------------------------------------------------------------
# 見た目
# ---------------------------------------------------------------------------
FONT = 'BIZ UDPゴシック'

C_INK = 'FF1B1D24'
C_TEXT2 = 'FF4E5163'
C_TEXT3 = 'FF8A8D9C'
C_PAPER = 'FFF7F5F0'
C_HAIR = 'FFEFECE5'
C_BUTTER = 'FFE8B44A'
C_BUTTER_SOFT = 'FFFDF3D8'
C_BUTTER_STRONG = 'FFF6D58A'
C_OK = 'FF00795A'
C_OK_SOFT = 'FFE3F1EC'
C_WARN = 'FFC9761A'
C_WARN_SOFT = 'FFFBEEDF'
C_CRIT = 'FFA32A3C'
C_CRIT_SOFT = 'FFF8E3E6'
C_WHITE = 'FFFFFFFF'
C_GREY_SOFT = 'FFEDEBE6'

F_BASE = Font(name=FONT, size=11, color=C_INK)
F_BOLD = Font(name=FONT, size=11, bold=True, color=C_INK)
F_TITLE = Font(name=FONT, size=18, bold=True, color=C_INK)
F_NOTE = Font(name=FONT, size=9, color=C_TEXT3)
F_NOTE2 = Font(name=FONT, size=10, color=C_TEXT2)
F_HEAD = Font(name=FONT, size=10, bold=True, color=C_TEXT2)
F_INPUT = Font(name=FONT, size=11, bold=True, color='FF7A5200')
F_LINK = Font(name=FONT, size=11, bold=True, color=C_INK, underline='single')
F_BAR = Font(name=FONT, size=11, bold=True, color=C_WHITE)
F_BAR_LINK = Font(name=FONT, size=10, bold=True, color=C_BUTTER, underline='single')
F_SMALL = Font(name=FONT, size=8, color=C_TEXT3)

FILL_HEAD = PatternFill('solid', fgColor='FFF3F1EC')
FILL_INPUT = PatternFill('solid', fgColor=C_BUTTER_SOFT)
FILL_AUTO = PatternFill('solid', fgColor='FFFCFBF8')
FILL_ACCENT = PatternFill('solid', fgColor=C_PAPER)
FILL_BAR = PatternFill('solid', fgColor=C_INK)
FILL_CARD = PatternFill('solid', fgColor=C_WHITE)

HAIR = Side(style='thin', color=C_HAIR)
RULE = Side(style='medium', color=C_INK)
BORDER_ROW = Border(bottom=HAIR)
BORDER_HEAD = Border(bottom=RULE)
BORDER_BOX = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
UNLOCKED = Protection(locked=False)

FMT_INT = '#,##0;[Red]\\-#,##0'
FMT_DEC1 = '#,##0.0'
FMT_YEN = '[$¥-411]#,##0;[Red][$¥-411]\\-#,##0'
FMT_DATE = 'yyyy/m/d'
FMT_MD = 'm/d'
FMT_GRID = '#,##0;"×";"×"'      # 0以下は「×」＝切れている

# ---------------------------------------------------------------------------
# シート名と配置
# ---------------------------------------------------------------------------
KINDS = ['冷凍', '飲料', '常温']
KIND_COLOR = {'冷凍': 'FF2E6FB3', '飲料': 'FF00795A', '常温': 'FFC9761A'}

S_HOME = 'ホーム'
S_CUR = '①在庫を貼る'
S_K = {k: f'②{k}' for k in KINDS}
S_SHEET = '③発注書'
S_SORT = '仕分け'
S_PLAN = '動員を入れる'
S_INTRO = 'つかいかた'
S_ITEM = '商品マスタ'
S_TH = '劇場マスタ'
S_SET = '設定'
S_MID = '月1回_1ヶ月前の在庫'
S_PRV = '月1回_2ヶ月前の在庫'
S_CALC = '計算'

BAR_LABEL = {
    S_HOME: 'ホーム', S_CUR: 'STEP 1　在庫を貼る',
    S_SHEET: 'STEP 3　発注書を印刷する', S_SORT: '仕分け（社員）　冷凍・飲料・常温と納品日',
    S_PLAN: '随時　動員を入れる',
    S_INTRO: 'つかいかた', S_ITEM: '商品マスタ（社員）', S_TH: '劇場マスタ（社員）',
    S_SET: '設定（社員）', S_MID: '月1回　1ヶ月前の在庫を貼る', S_PRV: '月1回　2ヶ月前の在庫を貼る',
    S_CALC: '計算（さわらない）',
}
for _k in KINDS:
    BAR_LABEL[S_K[_k]] = f'STEP 2　{_k} を発注する'

CSV_HEADERS = ['対象日付', '劇場コード', '劇場名', '支払先コード', '支払先名',
               '大分類コード', '小分類コード', '商品分類名', '商品コード', '商品名',
               '入数', '在庫数ケース', '在庫数バラ', '総数', '規定数', '資産廃棄',
               '税抜単価', '仕入税区分', '仕入税率']
CSV_HEAD = 5
CSV_FIRST = 6
CSV_LAST = 10005
CSV_HELPER = 'U'        # マスタに無い商品に振る通し番号
CSV_TH = 'V'            # 劇場コード（文字に揃えたもの）
CSV_CODE = 'W'          # 商品コード（文字に揃えたもの）

MASTER_FIRST, MASTER_LAST = 6, 1005
MASTER_ACTIVE = 'L'     # 取扱中(自動)
MASTER_SEQ = 'M'        # 並び(自動)
TH_FIRST, TH_LAST = 6, 105
CALC_FIRST, CALC_LAST = 10, 409
N_ITEMS = CALC_LAST - CALC_FIRST + 1
ORD_FIRST = 10
ORD_LAST = ORD_FIRST + N_ITEMS - 1
DAYS = 14
NDT = 21                                # 日付・納品日フラグは21日分持つ（次の便の探索用）
IN_FIRST_COL = 13                       # 区分シート：入庫数の表 M..Z
PJ_FIRST_COL = IN_FIRST_COL + DAYS      # 区分シート：在庫の見込み AA..AN
CAP_FIRST_COL = PJ_FIRST_COL + DAYS + 4  # 区分シート：入力上限の補助列 AS..BF
PLAN_FIRST, PLAN_LAST = 16, 415         # 400日分
PLAN_DOW_ROW = 6
SHEET_FIRST, SHEET_LAST = 11, 110
VENDOR_FIRST, VENDOR_LAST = 4, 40       # 設定シートF列の仕入先一覧（F4＝すべて）
LT_MAX = 10

# 仕分けシート
SORT_DELIV_ROW = {'冷凍': 6, '飲料': 7, '常温': 8}     # 納品日の表（C..I＝月..日、J＝発注から納品までの日数）
SORT_CAT_FIRST, SORT_CAT_LAST = 12, 31                # 分類ごとの区分
SORT_OV_FIRST = 36                                     # 商品ごとの上書き（商品マスタの行と1対1）
SORT_OV_LAST = SORT_OV_FIRST + (MASTER_LAST - MASTER_FIRST)

SIZES = ['大規模', '中規模', '小規模']

Q_SET = f"'{S_SET}'"
Q_CALC = f"'{S_CALC}'"
Q_SHEET = f"'{S_SHEET}'"
Q_PLAN = f"'{S_PLAN}'"
Q_SORT = f"'{S_SORT}'"
Q_ITEM = f"'{S_ITEM}'"

# 区分シートの列
OC = {'no': 'B', 'name': 'C', 'par': 'D', 'stock': 'E', 'use': 'F', 'cut': 'G', 'status': 'H',
      'next': 'I', 'guide': 'J', 'unit': 'K', 'after': 'L',
      'counted': get_column_letter(PJ_FIRST_COL + DAYS), 'counted_date': get_column_letter(PJ_FIRST_COL + DAYS + 1),
      'code': get_column_letter(PJ_FIRST_COL + DAYS + 2), 'vendor': get_column_letter(PJ_FIRST_COL + DAYS + 3)}
IN_COLS = [get_column_letter(IN_FIRST_COL + d) for d in range(DAYS)]
PJ_COLS = [get_column_letter(PJ_FIRST_COL + d) for d in range(DAYS)]
CAP_COLS = [get_column_letter(CAP_FIRST_COL + d) for d in range(DAYS)]

# 在庫CSVの劇場コードは4桁（新宿＝0761）で、公式3桁＋"1" の形になっている。
THEATER_SEED = [
    ('089', 'すすきの', '北海道'), ('049', 'おいらせ下田', '青森'), ('050', '秋田', '秋田'),
    ('078', '仙台', '宮城'), ('024', 'ひたちなか', '茨城'), ('015', '宇都宮', '栃木'),
    ('075', 'ららぽーと富士見', '埼玉'), ('003', '市川コルトンプラザ', '千葉'),
    ('018', 'ららぽーと船橋', '千葉'), ('028', '八千代緑が丘', '千葉'),
    ('035', '流山おおたかの森', '千葉'), ('071', '市原', '千葉'), ('077', '柏', '千葉'),
    ('006', '南大沢', '東京'), ('009', '六本木ヒルズ', '東京'), ('012', '府中', '東京'),
    ('029', '錦糸町（楽天地・オリナス）', '東京'), ('040', '西新井', '東京'),
    ('044', 'お台場（シネマメディアージュ）', '東京'), ('073', '日本橋', '東京'),
    ('076', '新宿', '東京'), ('080', '上野', '東京'), ('081', '日比谷／シャンテ', '東京'),
    ('084', '池袋', '東京'), ('085', '立川立飛', '東京'), ('007', '海老名', '神奈川'),
    ('008', '小田原', '神奈川'), ('010', '川崎', '神奈川'), ('036', 'ららぽーと横浜', '神奈川'),
    ('067', '甲府', '山梨'), ('068', '上田', '長野'), ('053', 'ファボーレ富山', '富山'),
    ('054', '高岡', '富山'), ('020', '岐阜', '岐阜'), ('030', 'モレラ岐阜', '岐阜'),
    ('004', '浜松', '静岡'), ('016', '木曽川', '愛知'), ('021', '東浦', '愛知'),
    ('026', '津島', '愛知'), ('079', '赤池', '愛知'), ('023', '二条', '京都'),
    ('005', '泉北', '大阪'), ('032', 'なんば', '大阪'), ('037', '梅田', '大阪'),
    ('045', '鳳', '大阪'), ('072', 'くずはモール', '大阪'), ('088', 'ららぽーと門真', '大阪'),
    ('038', '伊丹', '兵庫'), ('064', '西宮OS', '兵庫'), ('013', '橿原', '奈良'),
    ('031', '岡南', '岡山'), ('019', '緑井', '広島'), ('017', '高知', '高知'),
    ('022', '直方', '福岡'), ('056', '天神（ソラリア館）', '福岡'), ('087', 'ららぽーと福岡', '福岡'),
    ('046', '長崎', '長崎'), ('014', '光の森', '熊本'), ('027', 'はません', '熊本'),
    ('057', '宇城', '熊本'), ('083', '熊本サクラマチ', '熊本'), ('055', '大分わさだ', '大分'),
    ('074', 'アミュプラザおおいた', '大分'), ('033', '与次郎', '鹿児島'),
]

# 分類名 → 区分 の初期値。仕分けページで直せる
DEFAULT_KIND = {
    'アルコール': '飲料', 'コールド': '飲料', 'コーヒー': '飲料', 'ホット': '飲料', 'その他ドリンク': '飲料',
    'ポップコーン': '常温', 'コンセ包材': '常温', 'フード調味料': '常温', 'ＳＥＴ作品コンボ': '常温',
    'ホットドッグ': '冷凍', '軽食系フード': '冷凍', '調理系スイーツ': '冷凍', 'その他フード': '冷凍',
}

# 計算シートの列。名前で引けるようにして、列ずれのバグを構造的に消す
_CALC_NAMES = (['no', 'code', 'name', 'cat', 'vendor', 'pack', 'price', 'csv_stock', 'kind',
                'seq_冷凍', 'seq_飲料', 'seq_常温', 'seq', 'lt',
                'counted', 'counted_date', 'stock', 'pi', 'par_case', 'par']
               + [f'dt{d}' for d in range(NDT)]
               + [f'in{d}' for d in range(DAYS)]
               + [f'use{d}' for d in range(DAYS)]
               + [f'b{d}' for d in range(DAYS)]
               + [f'a{d}' for d in range(DAYS)]
               + [f'k{d}' for d in range(DAYS)]
               + [f'ov{d}' for d in range(DAYS)]
               + [f'valid{d}' for d in range(DAYS)]
               + [f'cnt{d}' for d in range(DAYS)]
               + [f'gd{d}' for d in range(DAYS)]
               + [f'lo{d}' for d in range(DAYS)]
               + [f'lo2_{d}' for d in range(DAYS)]
               + [f'bad{d}' for d in range(DAYS)]
               + [f'od{d}' for d in range(DAYS)]
               + [f'nd{d}' for d in range(NDT)]
               + ['n_before', 'cut_before', 'n_post', 'cut_post', 'use_avg', 'n_in', 'first_in', 'last_in',
                  'next_del', 'last_ok', 'deadline', 'last_ok2', 'deadline2', 'guide', 'n_valid',
                  'bad_day', 'over_day', 'in_amt', 'post', 'status', 'after', 'unit', 'absent', 'pick',
                  'in_cur', 'in_mid', 'in_prv', 'change'])
CC = {name: get_column_letter(2 + i) for i, name in enumerate(_CALC_NAMES)}
DT = [CC[f'dt{d}'] for d in range(NDT)]
FLAG_ROW = {'冷凍': 5, '飲料': 6, '常温': 7}     # 計算シートの納品日フラグ行


# ---------------------------------------------------------------------------
# 小さな道具
# ---------------------------------------------------------------------------
def col(sheet, letter, first, last):
    return f"'{sheet}'!${letter}${first}:${letter}${last}"


def csv_col(sheet, letter):
    return col(sheet, letter, CSV_FIRST, CSV_LAST)


def item_col(letter):
    return col(S_ITEM, letter, MASTER_FIRST, MASTER_LAST)


def th_col(letter):
    return col(S_TH, letter, TH_FIRST, TH_LAST)


def calc_col(name):
    return col(S_CALC, CC[name], CALC_FIRST, CALC_LAST)


def kind_col(kind, letter):
    return col(S_K[kind], letter, ORD_FIRST, ORD_LAST)


def plan_col(letter):
    return col(S_PLAN, letter, PLAN_FIRST, PLAN_LAST)


def sort_cat_col(letter):
    return col(S_SORT, letter, SORT_CAT_FIRST, SORT_CAT_LAST)


def sort_ov_col(letter):
    return col(S_SORT, letter, SORT_OV_FIRST, SORT_OV_LAST)


def weekday_jp(date_expr):
    """曜日の文字。書式コード aaa は環境依存で壊れるので CHOOSE で組む。"""
    return f'CHOOSE(WEEKDAY({date_expr}),"日","月","火","水","木","金","土")'


def md(expr):
    """5/8(金) の形。"""
    return f'TEXT({expr},"m/d")&"("&{weekday_jp(expr)}&")"'


def style_row(ws, row, cols, font=F_BASE, fill=None, fmt=None, align=None, border=True):
    for c in cols:
        cell = ws[f'{c}{row}']
        cell.font = font
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        if align:
            cell.alignment = Alignment(horizontal=align, vertical='center')
        if border:
            cell.border = BORDER_ROW
    if ws.row_dimensions[row].height is None:
        ws.row_dimensions[row].height = 18


def put_headers(ws, row, headers, kinds=None, start_col=2):
    for i, name in enumerate(headers):
        letter = get_column_letter(start_col + i)
        if kinds and kinds[i]:
            k = ws[f'{letter}{row - 1}']
            k.value = kinds[i]
            k.font = F_NOTE
            k.alignment = Alignment(horizontal='center')
        cell = ws[f'{letter}{row}']
        cell.value = name
        cell.font = F_HEAD
        cell.border = BORDER_HEAD
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30


def app_bar(ws, width=16):
    label = BAR_LABEL.get(ws.title, ws.title)
    last = get_column_letter(1 + width)
    ws.merge_cells(f'B1:{last}1')
    c = ws['B1']
    c.value = '  ' + label
    c.font = F_BAR
    c.fill = FILL_BAR
    c.alignment = Alignment(horizontal='left', vertical='center')
    for i in range(2, 2 + width):
        ws.cell(1, i).fill = FILL_BAR
    home = ws.cell(1, 2 + width)
    if ws.title != S_HOME:
        home.value = '◀ ホーム  '
        home.hyperlink = Hyperlink(ref=home.coordinate, location=f"'{S_HOME}'!A1", display='◀ ホーム')
        home.font = F_BAR_LINK
    home.fill = FILL_BAR
    home.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[1].height = 24
    ws.sheet_view.showGridLines = False


def sheet_title(ws, text, note=None, width=16):
    app_bar(ws, width)
    ws['B2'] = text
    ws['B2'].font = F_TITLE
    ws.row_dimensions[2].height = 30
    if note:
        ws['B3'] = note
        ws['B3'].font = F_NOTE2


def link_cell(cell, sheet, text=None):
    cell.value = text if text is not None else sheet
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet}'!A1", display=cell.value)
    cell.font = F_LINK
    return cell


def list_validation(ws, formula, cells, title='選択', error='一覧から選んでください。'):
    dv = DataValidation(type='list', formula1=formula, allow_blank=True,
                        showErrorMessage=True, errorTitle=title, error=error)
    ws.add_data_validation(dv)
    dv.add(cells)
    return dv


def number_validation(ws, cells, kind, lo, hi, title, error, prompt_title=None, prompt=None):
    dv = DataValidation(type=kind, operator='between', formula1=str(lo), formula2=str(hi),
                        showErrorMessage=True, errorTitle=title, error=error,
                        showInputMessage=bool(prompt), promptTitle=prompt_title or title, prompt=prompt)
    ws.add_data_validation(dv)
    dv.add(cells)
    return dv


def protect(ws, allow_filter=True):
    """数式のセルを守る。入力セルは各所で locked=False にしてある。パスワードは無し。"""
    ws.protection.sheet = True
    ws.protection.autoFilter = not allow_filter      # True＝禁止
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


def status_rules(ws, rng, size=11):
    """● △ ⚠ ○ ？ － の先頭記号で色を付ける。色だけに頼らず記号と文章でも読める。"""
    first = rng.split(':')[0]
    letter = ''.join(ch for ch in first if ch.isalpha())
    row = ''.join(ch for ch in first if ch.isdigit())
    for mark, color, soft, bold in (('●', C_CRIT, C_CRIT_SOFT, True), ('△', C_WARN, C_WARN_SOFT, True),
                                    ('⚠', C_WARN, C_WARN_SOFT, True), ('○', C_OK, None, False),
                                    ('？', C_TEXT3, None, False), ('－', C_TEXT3, None, False)):
        kw = dict(font=Font(name=FONT, size=size, bold=bold, color=color))
        if soft:
            kw['fill'] = PatternFill('solid', bgColor=soft)
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'LEFT(${letter}{row},1)="{mark}"'], stopIfTrue=True, **kw))


# ---------------------------------------------------------------------------
# ホーム
# ---------------------------------------------------------------------------
def build_home(wb):
    ws = wb.create_sheet(S_HOME)
    app_bar(ws, width=3)
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 44
    ws.column_dimensions['E'].width = 3

    ws['B2'] = '売店発注ツール'
    ws['B2'].font = Font(name=FONT, size=22, bold=True, color=C_INK)
    ws.row_dimensions[2].height = 34
    ws['B3'] = (f'=IF({Q_SET}!$C$6="","まだ在庫CSVが貼られていません",'
                f'"きょう "&{md(f"{Q_SET}!$C$6")}&"　"&{Q_SET}!$C$5&"　"&{Q_SET}!$C$4)')
    ws['B3'].font = F_NOTE2

    ws['B5'] = '毎日やること　上から順に'
    ws['B5'].font = F_BOLD

    def card(r, spine, title, sheet, sub, number, fmt, caption, status):
        ws.row_dimensions[r].height = 8
        for rr in range(r, r + 4):
            ws[f'B{rr}'].fill = PatternFill('solid', fgColor=spine)
            for c in 'CD':
                ws[f'{c}{rr}'].fill = FILL_CARD
        link_cell(ws[f'C{r + 1}'], sheet, title + '  ›')
        ws[f'C{r + 1}'].font = Font(name=FONT, size=13, bold=True, color=C_INK, underline='single')
        ws.row_dimensions[r + 1].height = 24
        ws[f'C{r + 2}'] = sub
        ws[f'C{r + 2}'].font = F_NOTE2
        ws[f'D{r + 1}'] = number
        ws[f'D{r + 1}'].font = Font(name=FONT, size=20, bold=True, color=C_INK)
        ws[f'D{r + 1}'].number_format = fmt
        ws[f'D{r + 1}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'D{r + 2}'] = caption
        ws[f'D{r + 2}'].font = F_NOTE
        ws[f'D{r + 2}'].alignment = Alignment(horizontal='right')
        ws[f'C{r + 3}'] = status
        ws[f'C{r + 3}'].font = F_NOTE2
        ws.row_dimensions[r + 3].height = 20
        status_rules(ws, f'C{r + 3}', size=10)
        return r + 5

    hit = f'COUNTIF({csv_col(S_CUR, CSV_TH)},{Q_SET}!$C$4)'
    r = card(6, C_BUTTER, 'STEP 1　在庫を貼る', S_CUR,
             '在庫システムの在庫一覧CSVを貼ります。', f'={hit}', '#,##0"行"', '読み込めた行',
             f'=IF({hit}=0,"● まだ貼られていません",IF({Q_SET}!$C$6<TODAY()-1,'
             f'"● 在庫が "&TEXT({Q_SET}!$C$6,"m/d")&" のものです（"&(TODAY()-{Q_SET}!$C$6)&"日前）。今日のCSVを貼ってください",'
             f'"○ 貼れています（"&TEXT({Q_SET}!$C$6,"m/d")&" の在庫）"))')
    for kind in KINDS:
        st = kind_col(kind, OC['status'])
        af = kind_col(kind, OC['after'])
        crit = f'COUNTIF({st},"●*")'
        warn = f'COUNTIF({st},"△*")'
        chk = f'COUNTIF({af},"⚠*")'
        cnt = f'COUNTIF({calc_col("kind")},"{kind}")'
        r = card(r, KIND_COLOR[kind], f'STEP 2　{kind} を発注する', S_K[kind],
                 f'="{kind} の商品 "&{cnt}&" 品目。「切れる日」を見て、届く日の列に入庫数を入れます。"',
                 f'={crit}', '#,##0"品目"', '今日発注が必要',
                 f'=IF({Q_CALC}!$B$3=0,"● 動員が入っていません → 動員を入れる",'
                 f'IF({Q_CALC}!$B$2=1,"● これから14日が動員の表の外です → 動員を入れる の「表の開始日」を直す",'
                 f'IF({chk}>0,"● ⚠ が "&{chk}&"件。直すまで発注書に載りません",'
                 f'IF({crit}>0,"● 赤 "&{crit}&"品目は今日中に発注。橙 "&{warn}&"品目も近く切れます",'
                 f'IF({warn}>0,"△ 今日は急ぎなし。橙 "&{warn}&"品目が近く切れます","○ 14日以内に切れる商品はありません")))))')
    cnt_all = f'COUNTIF({calc_col("n_valid")},">0")'
    amt = f'SUM({calc_col("in_amt")})'
    r = card(r, C_INK, 'STEP 3　発注書を印刷する', S_SHEET,
             '入庫を入れた商品が発注書に集まります。区分と仕入先を選んで印刷。', f'={amt}', FMT_YEN, '発注予定額（税抜）',
             f'=IF({cnt_all}=0,"入庫を入れた商品はまだありません","○ "&{cnt_all}&"品目を発注します")')

    ws[f'B{r}'] = '随時・社員'
    ws[f'B{r}'].font = F_BOLD
    r += 1
    unsorted = f'COUNTIFS({calc_col("code")},"?*",{calc_col("kind")},"")'
    nodeliv = '+'.join(f'IF(COUNTIF({Q_SORT}!$C${SORT_DELIV_ROW[k]}:$I${SORT_DELIV_ROW[k]},"○")=0,1,0)' for k in KINDS)
    r = card(r, C_TEXT2, '仕分け', S_SORT,
             '商品を 冷凍・飲料・常温 に分け、区分ごとの納品日（届く曜日）を決めます。', f'={unsorted}', '#,##0"件"',
             'まだ仕分けされていない商品',
             f'=IF({unsorted}>0,"● 未仕分けの商品があります。どのシートにも出ないので、仕分けページで区分を選んでください",'
             f'IF({nodeliv}>0,"△ 納品日（○）が付いていない区分があります。毎日届く前提で計算しています",'
             f'"○ 仕分けと納品日は設定済みです"))')
    filled = (f'COUNTIFS({plan_col("B")},">="&N({Q_SET}!$C$6),{plan_col("B")},"<="&(N({Q_SET}!$C$6)+{DAYS - 1}),'
              f'{plan_col("D")},">0")')
    r = card(r, C_OK, '動員を入れる', S_PLAN,
             '予測動員を入れると減り数のスピードが変わります。実績は月1回まとめてでも。', f'={filled}',
             f'0"/{DAYS}日"', 'これから14日で予測が入っている日',
             f'=IF({Q_CALC}!$B$2=1,"● これから14日が表の外です。「表の開始日」を今年の日付にしてください",'
             f'IF({Q_CALC}!$B$3=0,"● 動員が0です。標準の動員（設定）か予測を入れてください",'
             f'IF({filled}<{DAYS},"△ 予測が無い日は曜日の実績平均（無ければ標準の動員）を使います","○ 14日分そろっています")))')

    ws[f'B{r}'] = '月1回・社員'
    ws[f'B{r}'].font = F_BOLD
    r += 1
    hit_mid = f'COUNTIF({csv_col(S_MID, CSV_TH)},{Q_SET}!$C$4)'
    hit_prv = f'COUNTIF({csv_col(S_PRV, CSV_TH)},{Q_SET}!$C$4)'
    rows = [
        (S_MID, '取扱商品が変わっていないかの確認用。1ヶ月前の在庫CSVを貼り替えます。',
         f'=IF({hit_mid}=0,"未使用（貼らなくても動きます）","○ 貼れています（"&TEXT({Q_SET}!$C$12,"m/d")&" の在庫）")'),
        (S_PRV, '同じく2ヶ月前の在庫CSV。半年ごとに入れ替わる商品もこれで漏れません。',
         f'=IF({hit_prv}=0,"未使用（貼らなくても動きます）","○ 貼れています（"&TEXT({Q_SET}!$C$13,"m/d")&" の在庫）")'),
        (S_ITEM, '定数（最大保管数）と、100人あたりの減り数を商品ごとに決めます。',
         f'=IF({Q_SET}!$C$11=0,"",IF({Q_SET}!$C$17+{Q_SET}!$C$18+{Q_SET}!$C$19>0,'
         f'"● マスタに無い商品が "&({Q_SET}!$C$17+{Q_SET}!$C$18+{Q_SET}!$C$19)&"件（区分シートの末尾に出ています）→ マスタに追加して減り数を入れる",'
         f'IF({Q_SET}!$C$14<{Q_SET}!$C$20,"● 減り数が未設定の商品が "&({Q_SET}!$C$20-{Q_SET}!$C$14)&"件あります",'
         f'"○ 減り数はすべて設定済み（定数 "&{Q_SET}!$C$15&"/"&{Q_SET}!$C$20&"）")))'),
        (S_TH, '劇場コードと名前。CSVの劇場コードと一致しているか照合できます。', ''),
        (S_SET, '対象劇場・標準の動員・担当者・仕入先一覧。', ''),
        (S_INTRO, '色の意味と、迷ったときの読み方。', ''),
    ]
    for sheet, sub, status in rows:
        link_cell(ws[f'C{r}'], sheet, sheet + '  ›')
        ws[f'C{r + 1}'] = sub
        ws[f'C{r + 1}'].font = F_NOTE
        if status:
            ws[f'D{r}'] = status
            ws[f'D{r}'].font = F_NOTE2
            status_rules(ws, f'D{r}', size=10)
        r += 2
    ws[f'B{r + 1}'] = '黄色のセルだけ入力します。それ以外は保護されていて書き換えられません。'
    ws[f'B{r + 1}'].font = F_NOTE
    protect(ws, allow_filter=False)
    return ws


# ---------------------------------------------------------------------------
# つかいかた
# ---------------------------------------------------------------------------
def build_intro(wb):
    ws = wb.create_sheet(S_INTRO)
    sheet_title(ws, 'つかいかた', width=4)
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 100

    r = 5

    def head(text):
        nonlocal r
        ws[f'C{r}'] = text
        ws[f'C{r}'].font = F_BOLD
        ws[f'C{r}'].fill = FILL_HEAD
        ws[f'D{r}'].fill = FILL_HEAD
        r += 1

    def step(sheet, title, lines):
        nonlocal r
        link_cell(ws[f'C{r}'], sheet, sheet + ' ›')
        ws[f'D{r}'] = title
        ws[f'D{r}'].font = F_BOLD
        r += 1
        for line in lines:
            ws[f'D{r}'] = '　' + line
            ws[f'D{r}'].font = F_NOTE2
            r += 1
        r += 1

    head('毎日やること　これだけです')
    step(S_CUR, '在庫システムの「在庫一覧CSV」を貼る', [
        f'CSVを開いて、見出し行を除いたデータ部分をコピーし、A{CSV_FIRST} セルに貼り付けます。',
        '貼り替えるときは、前のデータ（A6 から S 列まで）だけを消してから貼ります。右の方の列は消さないでください。',
        '貼れたら上に「✔ ○行 読み込めました」と出ます。✖ のままなら貼る場所か劇場コードが違います。',
    ])
    step(S_K['冷凍'], '②冷凍 ②飲料 ②常温 の順に、「切れる日」を見て、届く日の列に入庫数を入れる', [
        '左から順に読みます。定数 → 在庫 → 減り数/日 → 切れる日 → このままだと（アラート）→ 次の便 → 目安。',
        '● 赤＝今日中に発注しないと間に合わない。△ 橙＝近く切れる（○/○ までに発注）。○ 緑＝14日以上もつ。',
        '「入庫を入れる」の表で、届く日の列に数を入れます（ケース入りはケース数、それ以外は個数）。黄色の日が納品日です。',
        '納品日でない日と、目安（定数まで入る数）を超える数は入りません。もっと入れるなら商品マスタの定数を直します。',
        '入れると右の「在庫の見込み」の表が復活し、「入庫を入れると」に結果が出ます。× は切れている日です。',
        '⚠ が出た行は直すまで発注書に載りません（納品日でない日／定数超え）。',
        '届いた分は翌日のCSVの在庫に入るので、過ぎた日の列は自動で無視されます。',
    ])
    step(S_SHEET, '区分と仕入先を選んで印刷する', [
        '入庫を入れた商品が、届く日ごとに1行ずつ集まります。区分と仕入先で絞れます。',
    ])
    head('随時・月1回（社員）')
    step(S_SORT, '商品を 冷凍・飲料・常温 に仕分けする、納品日を決める', [
        '上の表で、区分ごとの納品日（届く曜日）に ○ を付けます（1日でも7日でも）。「発注から納品までの日数」も入れます。',
        '「分類ごとの区分」で、CSVの商品分類名ごとにまとめて区分を選びます。ほとんどはこれで決まります。',
        '例外の商品だけ、下の「商品ごとの上書き」で区分を選びます。空欄なら分類の区分になります。',
        '区分を変えると、その区分のシートの行が並び直ります。入庫を入れていない朝いちにやってください。',
    ])
    step(S_PLAN, '動員（来場者数）を入れる', [
        '「予測」は大型作品の公開週など読みが変わる日だけで十分です。空欄はその曜日の実績平均→標準の動員の順で埋まります。',
        '「実績」は毎日でも、月1回まとめてでも構いません。入れるほど曜日の平均が正確になります。',
        '減り数/日 ＝ 100人あたりの減り数 × その日の動員 ÷ 100 で決まります。',
        '表は「表の開始日」から400日分です。空欄なら今年の1月1日から。年が変わったら開始日を新しい年にします。',
    ])
    step(S_MID, '月に1回、1ヶ月前・2ヶ月前の在庫CSVを貼り替える（貼らなくても動きます）', [
        '取扱商品が変わっていないかを見るためのものです。新商品・終売の件数がシートの上に出ます。',
    ])
    step(S_ITEM, '定数と減り数を決める', [
        '定数（ケース）＝その商品を棚に置ける最大数。入数を掛けた「定数（個）」が発注画面に出ます。',
        '100人あたりの減り数＝来場者100人につき何個減るか。1ヶ月の減り ÷ 1ヶ月の動員 × 100 で見当を付けます。',
        '減り数が空欄か 0 の商品は「？ 減り数が未設定」と出て、切れる日が計算できません。',
        '区分シートの行はこのマスタの順に並びます。行の挿入・削除・並べ替えは、入庫を入れていない朝いちに。',
        'CSVにあってマスタに無い商品は区分シートの末尾に出ます。マスタに追加して減り数を入れてください。',
    ])
    head('色と記号の意味')
    for text in [
        '黄色のセル … 入力するところ。それ以外は保護されていて書き換えられません（校閲 → シート保護の解除、パスワード無し）。',
        '● 赤 … 今日発注しないと間に合わない／入庫が遅い　　△ 橙 … 近く切れる、欠品の期間がある　　○ 緑 … 14日以上もつ',
        '⚠ … 直すところ（直すまで発注書に載らない）　　？ … 未仕分け・減り数か動員が未設定で計算できない　　－ … 今日のCSVに無い商品',
        '入庫を入れる の表 … 黄色の列が納品日。灰色の列には入りません。',
        '在庫の見込み の表 … 数字はその日の終わりの在庫。× は切れている。入庫のある日は太字。',
        '在庫が緑 … 「数えた在庫」を使っています（数えた日が今日のとき）。数えた日が古いと橙になり、CSVの在庫に戻ります。',
    ]:
        ws[f'D{r}'] = '　' + text
        ws[f'D{r}'].font = F_NOTE2
        r += 1
    protect(ws, allow_filter=False)
    return ws


# ---------------------------------------------------------------------------
# 在庫CSVを貼るシート（当日・1ヶ月前・2ヶ月前）
# ---------------------------------------------------------------------------
def build_csv_sheet(wb, name, others, label, note, paste_hint):
    ws = wb.create_sheet(name)
    app_bar(ws, width=12)
    ws['B2'] = label
    ws['B2'].font = F_TITLE
    ws.row_dimensions[2].height = 30
    ws['B3'] = note
    ws['B3'].font = F_NOTE2
    ws['B4'] = paste_hint
    ws['B4'].font = Font(name=FONT, size=11, bold=True, color=C_INK)

    hit = f'COUNTIF($V${CSV_FIRST}:$V${CSV_LAST},{Q_SET}!$C$4)'
    ws['H2'] = f'=IF({hit}=0,"✖ まだ貼れていません","✔ "&{hit}&"行 読み込めました")'
    ws['H2'].font = Font(name=FONT, size=16, bold=True, color=C_INK)
    ws['H2'].alignment = Alignment(horizontal='left', vertical='center')
    if not others:
        ws['H3'] = (f'=IF({hit}=0,"CSVの見出し行を除いたデータを A{CSV_FIRST} セルに貼り付けてください。'
                    f'貼ったのに0のままなら、設定の劇場コードとCSVの劇場コードが違います。",'
                    f'IF({Q_SET}!$C$11>{N_ITEMS},"⚠ 取扱商品が "&{Q_SET}!$C$11&" 件あります。'
                    f'計算できるのは {N_ITEMS} 件までなので、超えた分は発注画面に出ません。",'
                    f'IF(COUNTIF($A${CSV_FIRST}:$A${CSV_LAST},"<>"&INDEX($A${CSV_FIRST}:$A${CSV_LAST},'
                    f'MATCH({Q_SET}!$C$4,$V${CSV_FIRST}:$V${CSV_LAST},0)))-COUNTBLANK($A${CSV_FIRST}:$A${CSV_LAST})>0,'
                    f'"⚠ 日付の違う行が混ざっています。前のデータを消してから同じ日のCSVだけを貼ってください。",'
                    f'IF({Q_SET}!$C$21>0,"⚠ 同じ商品コードの行が "&{Q_SET}!$C$21&" 件重なっています（在庫は合算して1行で出します）。'
                    f'貼り替えのときに前のデータを消し忘れていませんか。",'
                    f'"対象劇場："&{Q_SET}!$C$4&"　"&{Q_SET}!$C$5&"　　CSVの日付："&TEXT({Q_SET}!$C$6,"yyyy/m/d")))))')
    else:
        ws['H3'] = (f'=IF({hit}=0,"貼らなくても動きます。貼ると、取扱商品の入れ替わりが分かります。",'
                    f'"対象劇場："&{Q_SET}!$C$4&"　"&{Q_SET}!$C$5)')
    ws['H3'].font = F_NOTE2
    ws.conditional_formatting.add('H2', FormulaRule(
        formula=[f'{hit}=0'], fill=PatternFill('solid', bgColor=C_CRIT_SOFT),
        font=Font(name=FONT, size=16, bold=True, color=C_CRIT), stopIfTrue=True))
    ws.conditional_formatting.add('H2', FormulaRule(
        formula=['TRUE'], fill=PatternFill('solid', bgColor=C_OK_SOFT),
        font=Font(name=FONT, size=16, bold=True, color=C_OK), stopIfTrue=True))

    if others:
        change = calc_col('change')
        checks = [('N', '新商品（今月から）'), ('P', '新商品（先月から）'),
                  ('R', '終売（今月から）'), ('T', '終売（先月から）'), ('V', '復活（先月は無し）')]
        ws['N4'] = '取扱の変化（当日・1ヶ月前・2ヶ月前を突き合わせた結果）'
        ws['N4'].font = Font(name=FONT, size=10, bold=True, color=C_INK)
        for letter, cname in checks:
            nxt = get_column_letter(ws[f'{letter}1'].column + 1)
            c1 = ws[f'{letter}2']
            c1.value = f'=COUNTIF({change},"{cname}")'
            c1.font = Font(name=FONT, size=18, bold=True, color=C_INK)
            c1.alignment = Alignment(horizontal='center')
            c2 = ws[f'{letter}3']
            c2.value = cname
            c2.font = F_NOTE
            c2.alignment = Alignment(horizontal='center')
            for cc in (c1, c2):
                cc.fill = FILL_HEAD
            ws.column_dimensions[letter].width = 6
            ws.column_dimensions[nxt].width = 12

    for i, h in enumerate(CSV_HEADERS):
        cell = ws.cell(CSV_HEAD, 1 + i, h)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.border = BORDER_BOX
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for ci, lab in ((21, 'マスタに無い商品の連番(自動)'), (22, '劇場コード(自動)'), (23, '商品コード(自動)')):
        helper = ws.cell(CSV_HEAD, ci, lab)
        helper.font = F_HEAD
        helper.fill = FILL_ACCENT
        helper.border = BORDER_BOX
        helper.alignment = Alignment(horizontal='center', wrap_text=True)
    ws['U4'] = '← この右の3列は自動です。消さないでください（消えたら社員に）。'
    ws['U4'].font = F_NOTE

    for r in range(CSV_FIRST, CSV_LAST + 1):
        # 貼ると「0761」が 761 に、商品コードが数値になることがある。文字に揃えてから比較する
        ws[f'V{r}'] = f'=IF($B{r}="","",TEXT($B{r},"0000"))'
        ws[f'W{r}'] = f'=IF($I{r}="","",IF(ISNUMBER($I{r}),TEXT($I{r},"0"),$I{r}))'
        ws[f'V{r}'].font = F_NOTE
        ws[f'W{r}'].font = F_NOTE
        nim = f'COUNTIF({item_col("B")},$W{r})=0'
        prev = f'N({CSV_HELPER}{r - 1})'
        if not others:
            # 当日CSV: 対象劇場の行のうち、商品マスタに無い商品だけに番号を振る（マスタの商品はマスタ順で並ぶ）
            ws[f'{CSV_HELPER}{r}'] = f'=IF(AND($V{r}={Q_SET}!$C$4,{nim}),{prev}+1,{prev})'
        else:
            # 過去のCSV: さらに、より新しいCSVにも無い商品だけ → 3枚を足すと重複なしになる
            dup = '+'.join(f'COUNTIFS({csv_col(o, CSV_TH)},$V{r},{csv_col(o, CSV_CODE)},$W{r})'
                           for o in others)
            ws[f'{CSV_HELPER}{r}'] = (
                f'=IF($V{r}<>{Q_SET}!$C$4,{prev},'
                f'IF(OR(NOT({nim}),{dup}>0),{prev},{prev}+1))')
        ws[f'{CSV_HELPER}{r}'].font = F_NOTE
        ws[f'B{r}'].number_format = '@'
        ws[f'I{r}'].number_format = '@'
    for letter, w in (('A', 12), ('B', 11), ('C', 10), ('E', 26), ('H', 16), ('I', 16),
                      ('J', 30), ('U', 14), ('V', 12), ('W', 14)):
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = f'A{CSV_FIRST}'
    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# 設定
# ---------------------------------------------------------------------------
def build_settings(wb, vendors):
    ws = wb.create_sheet(S_SET)
    sheet_title(ws, '設定', '黄色のセルだけ入力します。最初に1回、劇場コードを選んでください。', width=6)
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 64
    ws.column_dimensions['E'].width = 3
    ws.column_dimensions['F'].width = 30

    def csv_date(sheet):
        idx = f'MATCH($C$4,{csv_col(sheet, CSV_TH)},0)'
        raw = f'TEXT(INDEX({csv_col(sheet, "A")},{idx}),"00000000")'
        return (f'IFERROR(DATE(LEFT({raw},4),MID({raw},5,2),RIGHT({raw},2)),"")')

    rows = [
        # 4
        ('対象劇場コード', None, '劇場マスタから選びます。在庫CSVのこのコードの行だけを使います。', FILL_INPUT, '@'),
        # 5
        ('劇場名', f'=IFERROR(INDEX({th_col("C")},MATCH($C$4,{th_col("B")},0)),"")', '劇場マスタから自動。', FILL_AUTO, None),
        # 6
        ('当日基準日', f'={csv_date(S_CUR)}', '①に貼った在庫CSVの日付から自動。発注画面の「今日」になります。', FILL_AUTO, FMT_DATE),
        # 7
        ('発注から納品までの日数（既定）', None, f'仕分けページで区分ごとに空欄のときに使う日数（0〜{LT_MAX}）。', FILL_INPUT, '0"日"'),
        # 8
        ('標準の動員/日', None, '動員シートに予測も実績も無い日に使う来場者数。', FILL_INPUT, FMT_INT),
        # 9
        ('担当者', None, '発注書に印字します。', FILL_INPUT, None),
        # 10
        ('当日CSVの該当行数', f'=COUNTIF({csv_col(S_CUR, CSV_TH)},$C$4)', '0なら劇場コードが一致していません。', FILL_AUTO, FMT_INT),
        # 11
        ('取扱商品数', '=$C$20+$C$17+$C$18+$C$19', f'マスタの取扱商品＋マスタに無い商品。計算できるのは {N_ITEMS} 件まで。', FILL_AUTO, FMT_INT),
        # 12
        ('1ヶ月前の基準日', f'={csv_date(S_MID)}', '月1回_1ヶ月前の在庫 の日付から自動。', FILL_AUTO, FMT_DATE),
        # 13
        ('2ヶ月前の基準日', f'={csv_date(S_PRV)}', '月1回_2ヶ月前の在庫 の日付から自動。', FILL_AUTO, FMT_DATE),
        # 14
        ('減り数が設定済みの商品数', f'=COUNTIF({calc_col("pi")},">0")', '商品マスタの「100人あたりの減り数」が入っている商品。', FILL_AUTO, FMT_INT),
        # 15
        ('定数が設定済みの商品数', f'=COUNT({calc_col("par")})', '商品マスタの「定数（ケース）」（無ければCSVの規定数）が入っている商品。', FILL_AUTO, FMT_INT),
        # 16
        ('これから14日の動員（平均）', f'=IF({Q_CALC}!$B$3=0,0,ROUND({Q_CALC}!$B$3/{DAYS},0))', '0なら減り数が計算できません。', FILL_AUTO, FMT_INT),
        # 17
        ('当日CSVにあってマスタに無い商品数', f'=MAX({csv_col(S_CUR, CSV_HELPER)})', '区分シートの末尾に出ます。マスタに追加してください。', FILL_AUTO, FMT_INT),
        # 18
        ('1ヶ月前だけにあってマスタに無い商品数', f'=MAX({csv_col(S_MID, CSV_HELPER)})', '', FILL_AUTO, FMT_INT),
        # 19
        ('2ヶ月前だけにあってマスタに無い商品数', f'=MAX({csv_col(S_PRV, CSV_HELPER)})', '', FILL_AUTO, FMT_INT),
        # 20
        ('マスタの取扱商品数', f'=MAX({item_col(MASTER_SEQ)})', '商品マスタのうち、3枚のCSVのどれかに載っている商品。', FILL_AUTO, FMT_INT),
        # 21
        ('当日CSVで同じ商品コードが重なる行数', f'=$C$10-SUM({item_col("N")})+SUMPRODUCT(--({item_col("N")}>1))-$C$17',
         '0以外なら、同じ商品が複数行あります（在庫は合算して1行で出します）。', FILL_AUTO, FMT_INT),
    ]
    for i, (label, formula, note, fill, fmt) in enumerate(rows):
        r = 4 + i
        ws[f'B{r}'] = label
        ws[f'B{r}'].font = F_BOLD
        ws[f'B{r}'].fill = FILL_HEAD
        ws[f'B{r}'].border = BORDER_BOX
        cell = ws[f'C{r}']
        if formula:
            cell.value = formula
        cell.fill = fill
        cell.border = BORDER_BOX
        cell.font = F_INPUT if fill is FILL_INPUT else F_BASE
        if fmt:
            cell.number_format = fmt
        if fill is FILL_INPUT:
            cell.protection = UNLOCKED
        ws[f'D{r}'] = note
        ws[f'D{r}'].font = F_NOTE
    ws['C4'] = '0761'
    ws['C7'] = 3
    ws['C8'] = 1500
    list_validation(ws, f'={th_col("B")}', 'C4', '劇場コード', '劇場マスタにある4桁のコードを選んでください。')
    number_validation(ws, 'C7', 'whole', 0, LT_MAX, '日数', f'0〜{LT_MAX} の整数で入れてください。')
    number_validation(ws, 'C8', 'whole', 0, 100000, '動員', '1日の来場者数を整数で入れてください。')

    ws['F3'] = '仕入先の一覧（発注書の選択肢）'
    ws['F3'].font = F_BOLD
    ws[f'F{VENDOR_FIRST}'] = 'すべて'
    ws[f'F{VENDOR_FIRST}'].font = F_BASE
    for i in range(VENDOR_FIRST + 1, VENDOR_LAST + 1):
        c = ws[f'F{i}']
        c.fill = FILL_INPUT
        c.font = F_INPUT
        c.border = BORDER_ROW
        c.protection = UNLOCKED
    for i, v in enumerate(vendors[:VENDOR_LAST - VENDOR_FIRST]):
        ws[f'F{VENDOR_FIRST + 1 + i}'] = v
    ws[f'F{VENDOR_LAST + 1}'] = '仕入先が増えたらここに足してください（商品マスタの支払先名と同じ文字で）。'
    ws[f'F{VENDOR_LAST + 1}'].font = F_NOTE
    protect(ws, allow_filter=False)
    return ws


# ---------------------------------------------------------------------------
# 劇場マスタ・商品マスタ
# ---------------------------------------------------------------------------
def build_theater_master(wb):
    ws = wb.create_sheet(S_TH)
    sheet_title(ws, '劇場マスタ', '劇場コードは在庫CSVの4桁（例 0761）です。「CSV照合」が ✔ なら合っています。', width=6)
    put_headers(ws, 5, ['劇場コード', '劇場名', '都道府県', '規模', '備考', 'CSV照合'],
                ['入力', '入力', '入力', '選択', '入力', '自動'])
    for r in range(TH_FIRST, TH_LAST + 1):
        style_row(ws, r, 'BCDEF', fill=FILL_INPUT)
        style_row(ws, r, 'G', fill=FILL_AUTO)
        for letter in 'BCDEF':
            ws[f'{letter}{r}'].protection = UNLOCKED
        ws[f'B{r}'].number_format = '@'
        ws[f'G{r}'] = (
            f'=IF($B{r}="","",IF(COUNTIF({csv_col(S_CUR, CSV_TH)},$B{r})=0,"",'
            f'IFERROR(IF(INDEX({csv_col(S_CUR, "C")},MATCH($B{r},{csv_col(S_CUR, CSV_TH)},0))=$C{r},'
            f'"✔ 一致","△ CSVでは "&INDEX({csv_col(S_CUR, "C")},MATCH($B{r},{csv_col(S_CUR, CSV_TH)},0))),"")))')
        ws[f'G{r}'].font = Font(name=FONT, size=9, color=C_TEXT2)
    for i, (code3, name, pref) in enumerate(THEATER_SEED):
        r = TH_FIRST + i
        ws[f'B{r}'] = code3 + '1'
        ws[f'C{r}'] = name
        ws[f'D{r}'] = pref
    list_validation(ws, '"' + ','.join(SIZES) + '"', f'E{TH_FIRST}:E{TH_LAST}', '規模',
                    '大規模・中規模・小規模から選んでください。')
    for letter, w in (('B', 12), ('C', 30), ('D', 10), ('E', 10), ('F', 30), ('G', 18)):
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = 'B6'
    protect(ws, allow_filter=False)
    return ws


def build_item_master(wb, items):
    ws = wb.create_sheet(S_ITEM)
    sheet_title(ws, '商品マスタ',
                '「定数（ケース）」と「100人あたりの減り数」を入れると発注画面が動きます。'
                '区分シートはこの表の順に並びます。行の挿入・削除・並べ替えは入庫を入れていない朝いちに。', width=13)
    headers = ['商品コード', '商品名', '商品分類名', '支払先名', '入数', '税抜単価',
               '定数（ケース）', '定数（個）', '100人あたりの減り数', '備考', '取扱中', '並び', '当日CSVの行数', '区分']
    kinds = ['入力', '入力', '入力', '入力', '入力', '入力', '入力', '自動', '入力', '入力', '自動', '自動', '自動', '自動']
    put_headers(ws, 5, headers, kinds)
    th = f'{Q_SET}!$C$4'
    for r in range(MASTER_FIRST, MASTER_LAST + 1):
        style_row(ws, r, 'BCDEFGHJK', fill=FILL_INPUT)
        style_row(ws, r, 'ILMNO', fill=FILL_AUTO)
        for letter in 'BCDEFGHJK':
            ws[f'{letter}{r}'].protection = UNLOCKED
        ws[f'B{r}'].number_format = '@'
        ws[f'F{r}'].number_format = FMT_INT
        ws[f'G{r}'].number_format = FMT_YEN
        ws[f'H{r}'].number_format = FMT_INT
        ws[f'I{r}'] = f'=IF(OR($B{r}="",$H{r}=""),"",$H{r}*MAX(1,N($F{r})))'
        ws[f'I{r}'].number_format = FMT_INT
        ws[f'J{r}'].number_format = '0.00'
        ws[f'H{r}'].font = F_INPUT
        ws[f'J{r}'].font = F_INPUT
        # 3枚のCSVのどれかに載っていれば取扱中。区分シートはこの並び番号の順に商品を出す
        ws[f'{MASTER_ACTIVE}{r}'] = (
            f'=IF($B{r}="",0,IF(COUNTIFS({csv_col(S_CUR, CSV_TH)},{th},{csv_col(S_CUR, CSV_CODE)},$B{r})'
            f'+COUNTIFS({csv_col(S_MID, CSV_TH)},{th},{csv_col(S_MID, CSV_CODE)},$B{r})'
            f'+COUNTIFS({csv_col(S_PRV, CSV_TH)},{th},{csv_col(S_PRV, CSV_CODE)},$B{r})>0,1,0))')
        prev = f'N({MASTER_SEQ}{r - 1})' if r > MASTER_FIRST else '0'
        ws[f'{MASTER_SEQ}{r}'] = f'={prev}+${MASTER_ACTIVE}{r}'
        ws[f'N{r}'] = f'=IF($B{r}="",0,COUNTIFS({csv_col(S_CUR, CSV_TH)},{th},{csv_col(S_CUR, CSV_CODE)},$B{r}))'
        ws[f'O{r}'] = f'=IF($B{r}="","",{Q_SORT}!$G{SORT_OV_FIRST + r - MASTER_FIRST})'
        for letter in 'LMN':
            ws[f'{letter}{r}'].font = F_NOTE
    for i, it in enumerate(items[:MASTER_LAST - MASTER_FIRST + 1]):
        r = MASTER_FIRST + i
        ws[f'B{r}'] = it['商品コード']
        ws[f'C{r}'] = it['商品名']
        ws[f'D{r}'] = it['商品分類名']
        ws[f'E{r}'] = it['支払先名']
        ws[f'F{r}'] = int(it['入数'] or 1)
        ws[f'G{r}'] = float(it['税抜単価'] or 0)
    number_validation(ws, f'H{MASTER_FIRST}:H{MASTER_LAST}', 'decimal', 0, 9999, '定数',
                      'ケース数を 0 以上の数で入れてください。')
    number_validation(ws, f'J{MASTER_FIRST}:J{MASTER_LAST}', 'decimal', 0, 9999, '減り数',
                      '来場者100人あたりの減り数を 0 より大きい数で入れてください（例 6.5）。')
    ws.conditional_formatting.add(f'J{MASTER_FIRST}:J{MASTER_LAST}', FormulaRule(
        formula=[f'AND($J{MASTER_FIRST}<>"",$I{MASTER_FIRST}<>"",$J{MASTER_FIRST}*N({Q_SET}!$C$8)/100>$I{MASTER_FIRST})'],
        fill=PatternFill('solid', bgColor=C_WARN_SOFT), font=Font(name=FONT, size=11, bold=True, color=C_WARN)))
    ws['P5'] = '減り数が橙＝1日で定数以上減る計算です。「1日の減り数」を入れていませんか（100人あたりに直す）。区分は仕分けページで決めます。'
    ws['P5'].font = F_NOTE
    for letter, w in (('B', 16), ('C', 32), ('D', 16), ('E', 24), ('F', 7), ('G', 11),
                      ('H', 12), ('I', 10), ('J', 16), ('K', 26), ('L', 7), ('M', 6), ('N', 8), ('O', 8)):
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = 'D6'
    ws.auto_filter.ref = f'B5:O{MASTER_LAST}'
    protect(ws, allow_filter=True)
    return ws


# ---------------------------------------------------------------------------
# 仕分け（区分・納品日）
# ---------------------------------------------------------------------------
def build_sort(wb, categories):
    ws = wb.create_sheet(S_SORT)
    sheet_title(ws, '仕分け',
                '商品を 冷凍・飲料・常温 に分けます。区分ごとに納品日（届く曜日）を ○ で付けます（1日でも7日でも）。', width=10)

    # ---- 納品日 ----
    ws['B4'] = '納品日（届く曜日に ○）と、発注から納品までの日数'
    ws['B4'].font = F_BOLD
    put_headers(ws, 5, ['区分', '月', '火', '水', '木', '金', '土', '日', '発注から納品まで'],
                ['', '選択', '選択', '選択', '選択', '選択', '選択', '選択', '入力'])
    for kind in KINDS:
        r = SORT_DELIV_ROW[kind]
        ws[f'B{r}'] = kind
        ws[f'B{r}'].font = Font(name=FONT, size=11, bold=True, color=KIND_COLOR[kind])
        ws[f'B{r}'].border = BORDER_ROW
        for letter in 'CDEFGHI':
            c = ws[f'{letter}{r}']
            c.fill = FILL_INPUT
            c.font = F_INPUT
            c.border = BORDER_ROW
            c.alignment = Alignment(horizontal='center')
            c.protection = UNLOCKED
        j = ws[f'J{r}']
        j.fill = FILL_INPUT
        j.font = F_INPUT
        j.border = BORDER_ROW
        j.number_format = '0"日"'
        j.alignment = Alignment(horizontal='center')
        j.protection = UNLOCKED
        ws[f'K{r}'] = (f'=IF(COUNTIF($C{r}:$I{r},"○")=0,"⚠ ○ が無いので、毎日届く前提で計算しています",'
                       f'"○ 納品日 "&COUNTIF($C{r}:$I{r},"○")&"日／週　"&IF($J{r}="","日数は設定の既定（"&{Q_SET}!$C$7&"日）","発注から "&$J{r}&"日で届く"))')
        ws[f'K{r}'].font = F_NOTE2
        status_rules(ws, f'K{r}', size=10)
        ws.row_dimensions[r].height = 20
    list_validation(ws, '"○"', f'C{SORT_DELIV_ROW["冷凍"]}:I{SORT_DELIV_ROW["常温"]}', '納品日',
                    '届く曜日に ○ を入れます。届かない曜日は空欄にします。')
    number_validation(ws, f'J{SORT_DELIV_ROW["冷凍"]}:J{SORT_DELIV_ROW["常温"]}', 'whole', 0, LT_MAX, '日数',
                      f'発注してから届くまでの日数を 0〜{LT_MAX} で入れてください。')

    # ---- 分類ごとの区分 ----
    ws['B10'] = '分類ごとの区分（まとめて仕分け）　CSVの「商品分類名」ごとに区分を選びます'
    ws['B10'].font = F_BOLD
    put_headers(ws, SORT_CAT_FIRST - 1, ['商品分類名', '区分', '商品数'], ['入力', '選択', '自動'])
    for r in range(SORT_CAT_FIRST, SORT_CAT_LAST + 1):
        style_row(ws, r, 'BC', fill=FILL_INPUT)
        style_row(ws, r, 'D', fill=FILL_AUTO)
        ws[f'B{r}'].protection = UNLOCKED
        ws[f'C{r}'].protection = UNLOCKED
        ws[f'C{r}'].alignment = Alignment(horizontal='center')
        ws[f'D{r}'] = f'=IF($B{r}="","",COUNTIF({item_col("D")},$B{r}))'
        ws[f'D{r}'].number_format = FMT_INT
    for i, cat in enumerate(categories[:SORT_CAT_LAST - SORT_CAT_FIRST + 1]):
        ws[f'B{SORT_CAT_FIRST + i}'] = cat
        ws[f'C{SORT_CAT_FIRST + i}'] = DEFAULT_KIND.get(cat, '')
    list_validation(ws, '"' + ','.join(KINDS) + '"', f'C{SORT_CAT_FIRST}:C{SORT_CAT_LAST}', '区分',
                    '冷凍・飲料・常温 から選んでください。')

    # ---- まとめ ----
    ws['F10'] = '区分ごとの商品数'
    ws['F10'].font = F_BOLD
    for i, kind in enumerate(KINDS):
        r = SORT_CAT_FIRST + i
        link_cell(ws[f'F{r}'], S_K[kind], f'{kind} ›')
        ws[f'G{r}'] = f'=COUNTIF({calc_col("kind")},"{kind}")'
        ws[f'G{r}'].number_format = '#,##0"品目"'
        ws[f'G{r}'].font = F_BOLD
    r = SORT_CAT_FIRST + 3
    ws[f'F{r}'] = '未仕分け'
    ws[f'F{r}'].font = Font(name=FONT, size=11, bold=True, color=C_CRIT)
    ws[f'G{r}'] = f'=COUNTIFS({calc_col("code")},"?*",{calc_col("kind")},"")'
    ws[f'G{r}'].number_format = '#,##0"品目"'
    ws[f'G{r}'].font = Font(name=FONT, size=11, bold=True, color=C_CRIT)
    ws[f'F{r + 1}'] = '未仕分けの商品はどのシートにも出ません。下の表で「区分」が空の行を探して、分類の区分か上書きを入れてください。'
    ws[f'F{r + 1}'].font = F_NOTE
    ws[f'F{r + 3}'] = '区分を変えると、その区分のシートの行が並び直ります。'
    ws[f'F{r + 4}'] = '入庫を入れていない朝いちに、社員が行ってください。'
    for rr in (r + 3, r + 4):
        ws[f'F{rr}'].font = Font(name=FONT, size=10, bold=True, color=C_WARN)

    # ---- 商品ごとの上書き（商品マスタの行と1対1） ----
    ws[f'B{SORT_OV_FIRST - 3}'] = '商品ごとの上書き（例外だけ）　空欄なら分類の区分になります。行は商品マスタと同じ並びです'
    ws[f'B{SORT_OV_FIRST - 3}'].font = F_BOLD
    put_headers(ws, SORT_OV_FIRST - 1, ['商品コード', '商品名', '商品分類名', '分類の区分', '上書き', '区分（決定）'],
                ['自動', '自動', '自動', '自動', '選択', '自動'])
    for r in range(SORT_OV_FIRST, SORT_OV_LAST + 1):
        mr = MASTER_FIRST + (r - SORT_OV_FIRST)
        ws[f'B{r}'] = f'=IF({Q_ITEM}!$B{mr}="","",{Q_ITEM}!$B{mr})'
        ws[f'C{r}'] = f'=IF($B{r}="","",{Q_ITEM}!$C{mr})'
        ws[f'D{r}'] = f'=IF($B{r}="","",{Q_ITEM}!$D{mr})'
        ws[f'E{r}'] = (f'=IF($B{r}="","",IFERROR(IF(INDEX({sort_cat_col("C")},MATCH($D{r},{sort_cat_col("B")},0))="","",'
                       f'INDEX({sort_cat_col("C")},MATCH($D{r},{sort_cat_col("B")},0))),""))')
        ws[f'G{r}'] = f'=IF($B{r}="","",IF($F{r}<>"",$F{r},$E{r}))'
        style_row(ws, r, 'BCDEG', fill=FILL_AUTO)
        style_row(ws, r, 'F', fill=FILL_INPUT)
        ws[f'F{r}'].protection = UNLOCKED
        ws[f'F{r}'].alignment = Alignment(horizontal='center')
        ws[f'G{r}'].alignment = Alignment(horizontal='center')
        ws[f'B{r}'].font = F_NOTE
    list_validation(ws, '"' + ','.join(KINDS) + '"', f'F{SORT_OV_FIRST}:F{SORT_OV_LAST}', '区分',
                    '冷凍・飲料・常温 から選んでください。空欄なら分類の区分です。')
    ws.conditional_formatting.add(f'B{SORT_OV_FIRST}:G{SORT_OV_LAST}', FormulaRule(
        formula=[f'AND($B{SORT_OV_FIRST}<>"",$G{SORT_OV_FIRST}="")'],
        fill=PatternFill('solid', bgColor=C_CRIT_SOFT), font=Font(name=FONT, size=11, bold=True, color=C_CRIT)))
    for kind in KINDS:
        ws.conditional_formatting.add(f'G{SORT_OV_FIRST}:G{SORT_OV_LAST}', FormulaRule(
            formula=[f'$G{SORT_OV_FIRST}="{kind}"'], font=Font(name=FONT, size=11, bold=True, color=KIND_COLOR[kind])))
    for letter, w in (('B', 22), ('C', 30), ('D', 16), ('E', 12), ('F', 10), ('G', 12), ('H', 8), ('I', 8),
                      ('J', 16), ('K', 44)):
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = f'B{SORT_OV_FIRST}'
    protect(ws, allow_filter=False)
    return ws


# ---------------------------------------------------------------------------
# 動員を入れる
# ---------------------------------------------------------------------------
def build_plan(wb):
    ws = wb.create_sheet(S_PLAN)
    sheet_title(ws, '動員を入れる',
                '予測は読みが変わる日だけ。実績は月1回まとめてでも構いません。空欄の日は曜日の実績平均→標準の動員の順で埋まります。',
                width=14)
    ws['B5'] = '表の開始日'
    ws['B5'].font = F_BOLD
    ws['B5'].fill = FILL_HEAD
    ws['B5'].border = BORDER_BOX
    ws.merge_cells('C5:D5')
    ws['C5'].fill = FILL_INPUT
    ws['C5'].border = BORDER_BOX
    ws['C5'].font = F_INPUT
    ws['C5'].number_format = FMT_DATE
    ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C5'].protection = UNLOCKED
    ws['E5'] = ('← 空欄なら今年の1月1日から400日分。年が変わったら新しい年の1月1日を入れます。'
                '動かすと入力済みの行がずれるので、年に1回だけ。')
    ws['E5'].font = F_NOTE
    dv = DataValidation(type='date', operator='greaterThan', formula1='DATE(2020,1,1)',
                        showErrorMessage=True, errorTitle='開始日', error='日付を入れてください（例 2026/1/1）。')
    ws.add_data_validation(dv)
    dv.add('C5')

    date_c = plan_col('B')
    dow_c = plan_col('C')
    act_c = plan_col('E')
    recent = f'{date_c},">="&(N({Q_SET}!$C$6)-56),{date_c},"<"&N({Q_SET}!$C$6)'
    ws['J5'] = '曜日ごとの実績平均（直近8週）'
    ws['J5'].font = F_BOLD
    put_headers(ws, PLAN_DOW_ROW + 1, ['曜日', '実績の平均', '日数'], ['自動'] * 3, start_col=10)
    for i, day in enumerate('月火水木金土日'):
        r = PLAN_DOW_ROW + 2 + i
        ws[f'J{r}'] = day
        ws[f'K{r}'] = f'=IFERROR(AVERAGEIFS({act_c},{dow_c},$J{r},{recent}),"")'
        ws[f'L{r}'] = f'=COUNTIFS({dow_c},$J{r},{recent},{act_c},">0")'
        style_row(ws, r, 'JKL', fill=FILL_AUTO)
        ws[f'J{r}'].alignment = Alignment(horizontal='center')
        ws[f'K{r}'].number_format = FMT_INT
    dow_key = col(S_PLAN, 'J', PLAN_DOW_ROW + 2, PLAN_DOW_ROW + 8)
    dow_avg = col(S_PLAN, 'K', PLAN_DOW_ROW + 2, PLAN_DOW_ROW + 8)

    put_headers(ws, PLAN_FIRST - 1, ['日付', '曜日', '予測動員', '実績動員', '主な作品・備考', '採用する動員'],
                ['自動', '自動', '入力', '入力', '入力', '自動'])
    ws[f'B{PLAN_FIRST - 3}'] = '薄い黄色の帯＝これから14日（発注に使う範囲）。橙の実績＝入れ忘れ。'
    ws[f'B{PLAN_FIRST - 3}'].font = F_NOTE
    anchor = f'IF($C$5<>"",$C$5,IF({Q_SET}!$C$6="","",DATE(YEAR({Q_SET}!$C$6),1,1)))'
    for r in range(PLAN_FIRST, PLAN_LAST + 1):
        off = r - PLAN_FIRST
        ws[f'B{r}'] = f'=IF({anchor}="","",{anchor}+{off})'
        ws[f'C{r}'] = f'=IF($B{r}="","",{weekday_jp(f"$B{r}")})'
        fallback = (f'IF(N(IFERROR(INDEX({dow_avg},MATCH($C{r},{dow_key},0)),0))>0,'
                    f'IFERROR(INDEX({dow_avg},MATCH($C{r},{dow_key},0)),0),N({Q_SET}!$C$8))')
        ws[f'G{r}'] = (f'=IF($B{r}="","",IF(ISNUMBER($E{r}),$E{r},'
                       f'IF(ISNUMBER($D{r}),$D{r},ROUND({fallback},0))))')
        style_row(ws, r, 'BCDEFG', fill=FILL_AUTO)
        for letter in 'DEF':
            ws[f'{letter}{r}'].fill = FILL_INPUT
            ws[f'{letter}{r}'].font = F_INPUT
            ws[f'{letter}{r}'].protection = UNLOCKED
        ws[f'B{r}'].number_format = FMT_DATE
        for letter in 'DEG':
            ws[f'{letter}{r}'].number_format = FMT_INT
        ws[f'C{r}'].alignment = Alignment(horizontal='center')
    number_validation(ws, f'D{PLAN_FIRST}:E{PLAN_LAST}', 'whole', 0, 100000, '動員',
                      '来場者数を整数で入れてください。')

    rng = f'B{PLAN_FIRST}:G{PLAN_LAST}'
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND($B{PLAN_FIRST}<>"",$B{PLAN_FIRST}>={Q_SET}!$C$6,$B{PLAN_FIRST}<={Q_SET}!$C$6+{DAYS - 1})'],
        fill=PatternFill('solid', bgColor='FFFFF6DC'), stopIfTrue=False))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND($B{PLAN_FIRST}<>"",OR(WEEKDAY($B{PLAN_FIRST})=1,WEEKDAY($B{PLAN_FIRST})=7))'],
        font=Font(name=FONT, size=11, bold=True, color=C_TEXT2), stopIfTrue=False))
    ws.conditional_formatting.add(f'E{PLAN_FIRST}:E{PLAN_LAST}', FormulaRule(
        formula=[f'AND($B{PLAN_FIRST}<>"",$B{PLAN_FIRST}<{Q_SET}!$C$6,$E{PLAN_FIRST}="")'],
        fill=PatternFill('solid', bgColor=C_WARN_SOFT), stopIfTrue=True))
    for letter, w in (('B', 13), ('C', 6), ('D', 12), ('E', 12), ('F', 36), ('G', 13),
                      ('H', 3), ('I', 3), ('J', 8), ('K', 12), ('L', 8)):
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = f'B{PLAN_FIRST}'
    protect(ws, allow_filter=False)
    return ws


# ---------------------------------------------------------------------------
# 計算（非表示）
# ---------------------------------------------------------------------------
def build_calc(wb):
    ws = wb.create_sheet(S_CALC)
    ws['B1'] = 'このシートは計算専用です。書き換えないでください。'
    ws['B1'].font = F_NOTE
    ws['A2'] = '動員の表の範囲外なら1'
    ws['A3'] = '動員合計'
    ws['A4'] = '動員（日付は dt 列の3行目）'
    for kind in KINDS:
        ws[f'A{FLAG_ROW[kind]}'] = f'納品日フラグ {kind}'
    th = f'{Q_SET}!$C$4'
    base = f'{Q_SET}!$C$6'
    cur_v, cur_w = csv_col(S_CUR, CSV_TH), csv_col(S_CUR, CSV_CODE)
    mid_v, mid_w = csv_col(S_MID, CSV_TH), csv_col(S_MID, CSV_CODE)
    prv_v, prv_w = csv_col(S_PRV, CSV_TH), csv_col(S_PRV, CSV_CODE)
    cur_u, mid_u, prv_u = csv_col(S_CUR, CSV_HELPER), csv_col(S_MID, CSV_HELPER), csv_col(S_PRV, CSV_HELPER)

    for name, letter in CC.items():
        ws[f'{letter}{CALC_FIRST - 1}'] = name
        ws[f'{letter}{CALC_FIRST - 1}'].font = F_NOTE
    # 3行目＝日付、4行目＝その日の採用動員、5〜7行目＝区分ごとの納品日フラグ（21日分）
    for d, L in enumerate(DT):
        ws[f'{L}3'] = f'=IF({base}="","",{base}+{d})'
        ws[f'{L}3'].number_format = FMT_MD
        ws[f'{L}4'] = (f'=IF({L}3="",0,IFERROR(N(INDEX({plan_col("G")},MATCH({L}3,{plan_col("B")},0))),'
                       f'N({Q_SET}!$C$8)))')
        for kind in KINDS:
            wk = f'{Q_SORT}!$C${SORT_DELIV_ROW[kind]}:$I${SORT_DELIV_ROW[kind]}'
            ws[f'{L}{FLAG_ROW[kind]}'] = (f'=IF({L}3="",1,IF(COUNTIF({wk},"○")=0,1,'
                                          f'IF(INDEX({wk},WEEKDAY({L}3,2))="○",1,0)))')
    ws['B3'] = f'=SUM({DT[0]}4:{DT[DAYS - 1]}4)'
    ws['B2'] = (f'=IF({base}="",0,IF(OR({base}<N({Q_PLAN}!$B${PLAN_FIRST}),'
                f'{base}+{DAYS - 1}>N({Q_PLAN}!$B${PLAN_LAST})),1,0))')
    dates14 = f'${DT[0]}$3:${DT[DAYS - 1]}$3'

    def from_csv_text(letter):
        return (f'IFERROR(INDEX({csv_col(S_CUR, letter)},MATCH($C{{r}},{cur_w},0)),'
                f'IFERROR(INDEX({csv_col(S_MID, letter)},MATCH($C{{r}},{mid_w},0)),'
                f'IFERROR(INDEX({csv_col(S_PRV, letter)},MATCH($C{{r}},{prv_w},0)),"")))')

    def from_csv_num(letter):
        parts = []
        for sheet, v, w in ((S_CUR, cur_v, cur_w), (S_MID, mid_v, mid_w), (S_PRV, prv_v, prv_w)):
            cnt = f'COUNTIFS({v},{th},{w},$C{{r}})'
            parts.append((cnt, f'SUMIFS({csv_col(sheet, letter)},{v},{th},{w},$C{{r}})/{cnt}'))
        return (f'IF({parts[0][0]}>0,{parts[0][1]},IF({parts[1][0]}>0,{parts[1][1]},'
                f'IF({parts[2][0]}>0,{parts[2][1]},0)))')

    def from_master(letter, fallback):
        idx = f'INDEX({item_col(letter)},MATCH($C{{r}},{item_col("B")},0))'
        return f'IFERROR(IF({idx}="",{fallback},{idx}),{fallback})'

    def master_pos(letter):
        idx = f'INDEX({item_col(letter)},MATCH($C{{r}},{item_col("B")},0))'
        return f'IFERROR(IF(N({idx})>0,{idx},""),"")'

    for i in range(N_ITEMS):
        r = CALC_FIRST + i
        c = lambda n: f'${CC[n]}{r}'          # noqa: E731
        blank = f'$C{r}=""'
        R = lambda s: s.replace('{r}', str(r))   # noqa: E731
        f = {}
        f['no'] = f'=IF({i + 1}<={Q_SET}!$C$11,{i + 1},"")'
        f['code'] = (f'=IF($B{r}="","",IF($B{r}<={Q_SET}!$C$20,INDEX({item_col("B")},MATCH($B{r},{item_col(MASTER_SEQ)},0)),'
                     f'IF($B{r}<={Q_SET}!$C$20+{Q_SET}!$C$17,INDEX({cur_w},MATCH($B{r}-{Q_SET}!$C$20,{cur_u},0)),'
                     f'IF($B{r}<={Q_SET}!$C$20+{Q_SET}!$C$17+{Q_SET}!$C$18,'
                     f'INDEX({mid_w},MATCH($B{r}-{Q_SET}!$C$20-{Q_SET}!$C$17,{mid_u},0)),'
                     f'INDEX({prv_w},MATCH($B{r}-{Q_SET}!$C$20-{Q_SET}!$C$17-{Q_SET}!$C$18,{prv_u},0))))))')
        f['name'] = R(f'=IF({blank},"",{from_master("C", from_csv_text("J"))})')
        f['cat'] = R(f'=IF({blank},"",{from_master("D", from_csv_text("H"))})')
        f['vendor'] = R(f'=IF({blank},"",{from_master("E", from_csv_text("E"))})')
        f['pack'] = R(f'=IF({blank},"",MAX(1,N({from_master("F", from_csv_num("K"))})))')
        f['price'] = R(f'=IF({blank},"",N({from_master("G", from_csv_num("Q"))}))')
        f['csv_stock'] = (f'=IF({blank},"",IF(COUNTIFS({cur_v},{th},{cur_w},$C{r})=0,"",'
                          f'SUMIFS({csv_col(S_CUR, "N")},{cur_v},{th},{cur_w},$C{r})))')
        # 区分：マスタにある商品は仕分けページの「区分（決定）」、無い商品は分類の区分
        by_cat = (f'IFERROR(IF(INDEX({sort_cat_col("C")},MATCH({c("cat")},{sort_cat_col("B")},0))="","",'
                  f'INDEX({sort_cat_col("C")},MATCH({c("cat")},{sort_cat_col("B")},0))),"")')
        f['kind'] = (f'=IF({blank},"",IFERROR(IF(INDEX({sort_ov_col("G")},MATCH($C{r},{item_col("B")},0))="",'
                     f'{by_cat},INDEX({sort_ov_col("G")},MATCH($C{r},{item_col("B")},0))),{by_cat}))')
        for kind in KINDS:
            prev = f'N(${CC[f"seq_{kind}"]}{r - 1})' if i else '0'
            f[f'seq_{kind}'] = f'=IF({c("kind")}="{kind}",{prev}+1,{prev})'
        f['seq'] = (f'=IF({c("kind")}="冷凍",{c("seq_冷凍")},IF({c("kind")}="飲料",{c("seq_飲料")},'
                    f'IF({c("kind")}="常温",{c("seq_常温")},"")))')

        def lt_cell(k):
            cell = f'{Q_SORT}!$J${SORT_DELIV_ROW[k]}'
            return f'IF({cell}="",N({Q_SET}!$C$7),{cell})'
        f['lt'] = (f'=IF({c("kind")}="冷凍",{lt_cell("冷凍")},IF({c("kind")}="飲料",{lt_cell("飲料")},'
                   f'IF({c("kind")}="常温",{lt_cell("常温")},N({Q_SET}!$C$7))))')

        def from_sheet(letter):
            """区分シートの、この商品の行にある値。"""
            parts = ','.join(f'IF({c("kind")}="{k}",INDEX({kind_col(k, letter)},{c("seq")})' for k in KINDS)
            return f'IF({c("seq")}="","",{parts},""' + ')' * (len(KINDS) + 1)
        f['counted'] = f'=IF({blank},"",IF({from_sheet(OC["counted"])}="","",{from_sheet(OC["counted"])}))'
        f['counted_date'] = f'=IF({blank},"",IF({from_sheet(OC["counted_date"])}="","",{from_sheet(OC["counted_date"])}))'
        f['stock'] = (f'=IF({blank},"",IF(AND(ISNUMBER({c("counted")}),ISNUMBER({c("counted_date")}),'
                      f'{c("counted_date")}>={base}),{c("counted")},N({c("csv_stock")})))')
        f['pi'] = R(f'=IF({blank},"",{master_pos("J")})')
        f['par_case'] = R(f'=IF({blank},"",{master_pos("H")})')
        csv_par = R(f'N({from_csv_num("O")})')
        f['par'] = (f'=IF({blank},"",IF({c("par_case")}<>"",{c("par_case")}*{c("pack")},'
                    f'IF({csv_par}>0,{csv_par},"")))')
        lt = c('lt')

        def flag(d):
            return (f'IF({c("kind")}="冷凍",{DT[d]}$5,IF({c("kind")}="飲料",{DT[d]}$6,'
                    f'IF({c("kind")}="常温",{DT[d]}$7,1)))')
        start = f'MAX(0,N({c("stock")}))'
        for d in range(DAYS):
            L = DT[d]
            f[f'in{d}'] = f'=IF({blank},0,MAX(0,N({from_sheet(IN_COLS[d])})))'
            f[f'use{d}'] = f'=IF({blank},0,IF({c("pi")}="",0,MAX(0,{c("pi")})*MAX(0,{L}$4)/100))'
            prev_b = start if d == 0 else c(f'b{d - 1}')
            prev_a = start if d == 0 else c(f'a{d - 1}')
            units = f'{c(f"in{d}")}*{c("pack")}'
            f[f'b{d}'] = f'=IF({blank},0,MAX(0,{prev_b}-{c(f"use{d}")}))'
            f[f'a{d}'] = f'=IF({blank},0,MAX(0,{prev_a}+{units}-{c(f"use{d}")}))'
            f[f'k{d}'] = f'=IF({blank},99,IF(AND({c(f"a{d}")}<1,{L}$3>={c("last_in")}),{d},99))'
            f[f'ov{d}'] = (f'=IF(OR({blank},{c("par")}="",{c(f"in{d}")}=0),0,'
                           f'ROUNDUP(MAX(0,{prev_a}+{units}-{c("par")})/{c("pack")},0))')
            f[f'valid{d}'] = f'=IF(AND({c(f"in{d}")}>0,{flag(d)}=1,{c(f"ov{d}")}=0),1,0)'
            prev_cnt = '0' if d == 0 else c(f'cnt{d - 1}')
            f[f'cnt{d}'] = f'={prev_cnt}+{c(f"valid{d}")}'
            # その日に入れられる上限（ケース）。納品日でなければ 0
            f[f'gd{d}'] = (f'=IF({blank},0,IF({flag(d)}=0,0,IF({c("par")}="",9999,'
                           f'IF({c("par")}-{prev_b}<=0,0,ROUNDUP(({c("par")}-{prev_b})/{c("pack")},0)))))')
            f[f'lo{d}'] = f'=IF(AND({c("cut_before")}<>"",{L}$3<={c("cut_before")},{flag(d)}=1),{L}$3,0)'
            f[f'lo2_{d}'] = f'=IF(AND({c("cut_post")}<>"",{L}$3<={c("cut_post")},{flag(d)}=1),{L}$3,0)'
            f[f'bad{d}'] = f'=IF(AND({c(f"in{d}")}>0,{flag(d)}=0),{L}$3,99999)'
            f[f'od{d}'] = f'=IF({c(f"ov{d}")}>0,{L}$3,99999)'
        for d in range(NDT):
            L = DT[d]
            f[f'nd{d}'] = f'=IF(AND({L}$3<>"",{L}$3>=N({base})+{lt},{flag(d)}=1),{L}$3,99999)'

        def rng(prefix, n=DAYS):
            return f'${CC[f"{prefix}0"]}{r}:${CC[f"{prefix}{n - 1}"]}{r}'
        in_rng, b_rng, k_rng = rng('in'), rng('b'), rng('k')
        f['n_before'] = f'=IF(OR({blank},{c("pi")}=""),"",COUNTIF({b_rng},">=1"))'
        f['cut_before'] = f'=IF({c("n_before")}="","",IF({c("n_before")}>={DAYS},"",{base}+{c("n_before")}))'
        f['n_post'] = f'=IF(OR({blank},{c("pi")}=""),"",MIN({k_rng}))'
        f['cut_post'] = f'=IF({c("n_post")}="","",IF({c("n_post")}>=99,"",{base}+{c("n_post")}))'
        f['use_avg'] = f'=IF(OR({blank},{c("pi")}=""),"",AVERAGE({rng("use")}))'
        f['n_in'] = f'=IF({blank},0,COUNTIF({in_rng},">0"))'
        f['first_in'] = f'=IF({c("n_in")}=0,"",SUMPRODUCT(MIN(({in_rng}>0)*{dates14}+({in_rng}<=0)*99999)))'
        f['last_in'] = f'=IF({c("n_in")}=0,0,SUMPRODUCT(MAX(({in_rng}>0)*{dates14})))'
        f['next_del'] = f'=IF({blank},"",IF(MIN({rng("nd", NDT)})>=99999,"",MIN({rng("nd", NDT)})))'
        f['last_ok'] = f'=IF({c("cut_before")}="",0,MAX({rng("lo")}))'
        f['deadline'] = f'=IF({c("last_ok")}=0,"",{c("last_ok")}-{lt})'
        f['last_ok2'] = f'=IF({c("cut_post")}="",0,MAX({rng("lo2_")}))'
        f['deadline2'] = f'=IF({c("last_ok2")}=0,"",{c("last_ok2")}-{lt})'
        pos = f'MATCH({c("next_del")},{dates14},0)'
        f['guide'] = (f'=IF(OR({blank},{c("par")}="",{c("next_del")}=""),"",'
                      f'IFERROR(INDEX({rng("gd")},{pos}),""))')
        f['n_valid'] = f'=IF({blank},0,{c(f"cnt{DAYS - 1}")})'
        f['bad_day'] = f'=IF(MIN({rng("bad")})>=99999,"",MIN({rng("bad")}))'
        f['over_day'] = f'=IF(MIN({rng("od")})>=99999,"",MIN({rng("od")}))'
        f['in_amt'] = f'=IF({blank},0,SUMPRODUCT({rng("valid")},{in_rng})*{c("pack")}*N({c("price")}))'
        f['unit'] = f'=IF({blank},"",IF({c("pack")}>1,"ｹｰｽ("&{c("pack")}&"入)","個"))'
        f['absent'] = f'=IF({blank},0,IF(COUNTIFS({cur_v},{th},{cur_w},$C{r})>0,0,1))'
        cut_b = md(c('cut_before'))
        cut_p = md(c('cut_post'))
        f['post'] = (f'=IF({blank},"",IF({c("cut_post")}="","入庫後は 14日以上もつ",'
                     f'IF({c("deadline2")}="","入庫後も "&{cut_p}&" に切れる（それまでに便が無い）　もう1便を至急発注",'
                     f'IF({c("deadline2")}<={base},"入庫後も "&{cut_p}&" に切れる　もう1便を至急発注（"&{md(c("last_ok2"))}&" 便）",'
                     f'"入庫後も "&{cut_p}&" に切れる　"&{md(c("deadline2"))}&" までに次を発注（"&{md(c("last_ok2"))}&" 便）"))))')
        f['status'] = (
            f'=IF({blank},"",'
            f'IF({base}="","？ ①在庫を貼る に今日のCSVが貼られていない",'
            f'IF({c("absent")}=1,"－ 今日のCSVに無い（終売？）",'
            f'IF({c("kind")}="","？ 未仕分け（仕分けページで区分を選ぶ）",'
            f'IF({c("stock")}<0,"● 在庫がマイナス　数えて「数えた在庫」に入れる",'
            f'IF({c("stock")}=0,"● 在庫なし　至急発注",'
            f'IF({c("pi")}="","？ 減り数が未設定（商品マスタ）",'
            f'IF($B$3=0,"？ 動員が未入力（動員を入れる）",'
            f'IF({c("cut_before")}="","○ 14日以上もつ",'
            f'IF({c("deadline")}="","● "&{cut_b}&" に切れる　それまでに便が無い　至急発注",'
            f'IF({c("deadline")}<{base},"● "&{cut_b}&" に切れる　至急発注（"&{md(c("last_ok"))}&" 便）",'
            f'IF({c("deadline")}={base},"● "&{cut_b}&" に切れる　今日発注（"&{md(c("last_ok"))}&" 便）",'
            f'"△ "&{cut_b}&" に切れる　"&{md(c("deadline"))}&" までに発注（"&{md(c("last_ok"))}&" 便）"))))))))))))')
        gap = f'AND({c("cut_before")}<>"",{c("first_in")}>{c("cut_before")})'
        earlier = f'AND({c("next_del")}<>"",{c("next_del")}<{c("first_in")},{c("next_del")}<={c("cut_before")})'
        fi_prev = md(f'{c("first_in")}-1')
        f['after'] = (
            f'=IF({blank},"",'
            f'IF({c("n_in")}=0,"",'
            f'IF({c("bad_day")}<>"","⚠ "&{md(c("bad_day"))}&" は納品日ではありません → その列を消す",'
            f'IF({c("over_day")}<>"","⚠ "&{md(c("over_day"))}&" の入庫が定数を超えます → 減らす",'
            f'IF({c("pi")}="","？ 減り数が未設定（発注書には載ります）",'
            f'IF($B$3=0,"？ 動員が未入力（発注書には載ります）",'
            f'IF({gap},'
            f'IF({earlier},"● "&{cut_b}&" に切れる　"&{md(c("next_del"))}&" の便に入れてください",'
            f'"△ 欠品 "&{cut_b}&"〜"&{fi_prev}&"　"&{c("post")}),'
            f'IF({c("cut_post")}="","○ 入庫で 14日以上もつ",'
            f'IF({c("deadline2")}="","● 入庫しても "&{cut_p}&" に切れる（それまでに便が無い）　もう1便を至急発注",'
            f'IF({c("deadline2")}<={base},"● 入庫しても "&{cut_p}&" に切れる　もう1便を至急発注（"&{md(c("last_ok2"))}&" 便）",'
            f'"△ 入庫しても "&{cut_p}&" に切れる　"&{md(c("deadline2"))}&" までに次を発注（"&{md(c("last_ok2"))}&" 便）"))))))))))')
        sel_v = f'{Q_SHEET}!$C$4'
        sel_k = f'{Q_SHEET}!$C$5'
        prev_pick = f'N(${CC["pick"]}{r - 1})' if i else '0'
        f['pick'] = (f'=IF(AND({c("n_valid")}>0,OR({sel_v}="すべて",{sel_v}="",{c("vendor")}={sel_v}),'
                     f'OR({sel_k}="すべて",{sel_k}="",{c("kind")}={sel_k})),{prev_pick}+{c("n_valid")},{prev_pick})')
        f['in_cur'] = f'=IF({blank},0,IF(COUNTIFS({cur_v},{th},{cur_w},$C{r})>0,1,0))'
        f['in_mid'] = f'=IF({blank},0,IF(COUNTIFS({mid_v},{th},{mid_w},$C{r})>0,1,0))'
        f['in_prv'] = f'=IF({blank},0,IF(COUNTIFS({prv_v},{th},{prv_w},$C{r})>0,1,0))'
        cu, mi, pv = c('in_cur'), c('in_mid'), c('in_prv')
        f['change'] = (
            f'=IF(OR({blank},{Q_SET}!$C$12=""),"",'
            f'IF({Q_SET}!$C$13="",'
            f'IF(AND({cu}=1,{mi}=0),"新商品（今月から）",IF(AND({cu}=0,{mi}=1),"終売（今月から）","")),'
            f'IF(AND({cu}=1,{mi}=0,{pv}=0),"新商品（今月から）",'
            f'IF(AND({cu}=1,{mi}=1,{pv}=0),"新商品（先月から）",'
            f'IF(AND({cu}=0,{mi}=1,{pv}=1),"終売（今月から）",'
            f'IF(AND({cu}=0,{mi}=0,{pv}=1),"終売（先月から）",'
            f'IF(AND({cu}=1,{mi}=0,{pv}=1),"復活（先月は無し）","")))))))')
        for name, formula in f.items():
            ws[f'{CC[name]}{r}'] = formula
        for name in ('cut_before', 'cut_post', 'counted_date', 'first_in', 'last_in', 'next_del',
                     'last_ok', 'deadline', 'last_ok2', 'deadline2', 'bad_day', 'over_day'):
            ws[f'{CC[name]}{r}'].number_format = FMT_MD
    ws.sheet_state = 'hidden'
    protect(ws, allow_filter=False)
    return ws


# ---------------------------------------------------------------------------
# ②冷凍 ②飲料 ②常温
# ---------------------------------------------------------------------------
def build_category(wb, kind):
    ws = wb.create_sheet(S_K[kind])
    last_in, last_pj = IN_COLS[-1], PJ_COLS[-1]
    app_bar(ws, width=PJ_FIRST_COL + DAYS + 2)
    ws['B2'] = f'{kind} を発注する'
    ws['B2'].font = Font(name=FONT, size=18, bold=True, color=KIND_COLOR[kind])
    ws.row_dimensions[2].height = 30
    ws['E2'] = (f'=IF({Q_SET}!$C$6="","← まず ①在庫を貼る",'
                f'IF({Q_SET}!$C$6<TODAY()-1,"⚠ 在庫が "&TEXT({Q_SET}!$C$6,"m/d")&" のものです（"&(TODAY()-{Q_SET}!$C$6)&"日前）'
                f'→ 今日のCSVを ①在庫を貼る に貼ってください",'
                f'"きょう "&{md(f"{Q_SET}!$C$6")}&"　"&{Q_SET}!$C$5&"　　CSV "&{Q_SET}!$C$10&"行"))')
    ws['E2'].font = F_NOTE2
    ws.conditional_formatting.add('E2', FormulaRule(
        formula=['LEFT($E$2,1)="⚠"'], font=Font(name=FONT, size=11, bold=True, color=C_WARN)))
    drow = SORT_DELIV_ROW[kind]
    wk = f'{Q_SORT}!$C${drow}:$I${drow}'
    ws['L2'] = (f'=IF(COUNTIF({wk},"○")=0,"⚠ {kind} の納品日（曜日）が未設定。毎日届く前提で計算しています → 仕分けページで ○ を付ける",'
                f'"納品日："&IF({Q_SORT}!$C${drow}="○","月・","")&IF({Q_SORT}!$D${drow}="○","火・","")&IF({Q_SORT}!$E${drow}="○","水・","")'
                f'&IF({Q_SORT}!$F${drow}="○","木・","")&IF({Q_SORT}!$G${drow}="○","金・","")&IF({Q_SORT}!$H${drow}="○","土・","")'
                f'&IF({Q_SORT}!$I${drow}="○","日・","")&"　発注から "&IF({Q_SORT}!$J${drow}="",{Q_SET}!$C$7,{Q_SORT}!$J${drow})&"日で届く")')
    ws['L2'].font = F_NOTE2
    ws.conditional_formatting.add('L2', FormulaRule(
        formula=['LEFT($L$2,1)="⚠"'], font=Font(name=FONT, size=11, bold=True, color=C_WARN)))
    others = [k for k in KINDS if k != kind]
    for j, k in enumerate(others):
        link_cell(ws.cell(2, PJ_FIRST_COL + DAYS - 6 + j * 3), S_K[k], f'{k} ›')

    # ---- 上のまとめ ----
    st = f'${OC["status"]}${ORD_FIRST}:${OC["status"]}${ORD_LAST}'
    af = f'${OC["after"]}${ORD_FIRST}:${OC["after"]}${ORD_LAST}'
    crit = f'COUNTIF({st},"●*")'
    warn = f'COUNTIF({st},"△*")'
    chk = f'COUNTIF({af},"⚠*")'
    cnt = f'COUNTIFS({calc_col("kind")},"{kind}",{calc_col("n_valid")},">0")'
    amt = f'SUMIFS({calc_col("in_amt")},{calc_col("kind")},"{kind}")'
    cards = [
        ('C', 'C', '● 今日発注しないと間に合わない', f'={crit}', '#,##0"品目"', C_CRIT),
        ('D', 'G', '△ 近く切れる', f'={warn}', '#,##0"品目"', C_WARN),
        ('H', 'H', '⚠ 直すところ（直すまで発注書に載らない）', f'={chk}', '#,##0"件"', C_WARN),
        ('I', 'K', '入庫を入れた（発注書に載る）', f'={cnt}', '#,##0"品目"', C_INK),
        ('L', 'L', f'{kind} の発注予定額（税抜）', f'={amt}', FMT_YEN, C_INK),
    ]
    for c1, c2, label, formula, fmt, color in cards:
        if c1 != c2:
            ws.merge_cells(f'{c1}4:{c2}4')
            ws.merge_cells(f'{c1}5:{c2}5')
        ws[f'{c1}4'] = label
        ws[f'{c1}4'].font = Font(name=FONT, size=9, bold=True, color=color)
        ws[f'{c1}4'].alignment = Alignment(horizontal='left', vertical='bottom')
        ws[f'{c1}5'] = formula
        ws[f'{c1}5'].font = Font(name=FONT, size=20, bold=True, color=color)
        ws[f'{c1}5'].number_format = fmt
        ws[f'{c1}5'].alignment = Alignment(horizontal='left', vertical='center')
        for rr in (4, 5):
            for ci in range(ws[f'{c1}1'].column, ws[f'{c2}1'].column + 1):
                ws.cell(rr, ci).fill = FILL_ACCENT
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 30
    ws.merge_cells(f'{IN_COLS[0]}4:{last_pj}5')
    ws[f'{IN_COLS[0]}4'] = (
        f'=IF({Q_CALC}!$B$2=1,"⚠ これから14日が動員の表の外です。「動員を入れる」の表の開始日を今年の日付にしてください。",'
        f'IF({Q_CALC}!$B$3=0,"⚠ 動員が入っていません。減り数が計算できないので「動員を入れる」か「設定」の標準の動員を入れてください。",'
        f'"これから14日の動員：平均 "&TEXT({Q_SET}!$C$16,"#,##0")&" 人/日　　"'
        f'&"左の表：届く日の列に入庫数を入れる（黄色＝納品日）。右の表：その日の終わりの在庫。× は切れている。"))')
    ws[f'{IN_COLS[0]}4'].font = F_NOTE2
    ws[f'{IN_COLS[0]}4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.conditional_formatting.add(f'{IN_COLS[0]}4', FormulaRule(
        formula=[f'LEFT({IN_COLS[0]}$4,1)="⚠"'], fill=PatternFill('solid', bgColor=C_WARN_SOFT),
        font=Font(name=FONT, size=10, bold=True, color=C_WARN)))
    ws['B7'] = ('読み方：左から順に。「このままだと」が ● か △ なら、「入庫を入れる」の届く日の列（黄色）に数を入れる。'
                '迷ったら「次の便」の日に「目安」の数。⚠ は直す。')
    ws['B7'].font = F_NOTE2
    ws[f'{IN_COLS[0]}6'] = '入庫を入れる（届く日の列に数を入れる）'
    ws[f'{IN_COLS[0]}6'].font = F_BOLD
    ws[f'{PJ_COLS[0]}6'] = '在庫の見込み（その日の終わり）'
    ws[f'{PJ_COLS[0]}6'].font = F_BOLD

    # ---- 見出し ----
    headers = ['No', '商品名', '定数(個)', '在庫(個)', '減り数/日', '切れる日', 'このままだと',
               '次の便', '目安', '単位', '入庫を入れると']
    kinds_row = ['', '', '自動', '自動', '自動', '自動', '自動', '自動', '自動', '自動', '自動']
    put_headers(ws, ORD_FIRST - 1, headers, kinds_row)
    for cols in (IN_COLS, PJ_COLS):
        for d, L in enumerate(cols):
            h = ws[f'{L}{ORD_FIRST - 1}']
            h.value = f'=IF({Q_SET}!$C$6="","",{Q_SET}!$C$6+{d})'
            h.number_format = FMT_MD
            h.font = F_HEAD
            h.border = BORDER_HEAD
            h.alignment = Alignment(horizontal='center', vertical='center')
            w = ws[f'{L}{ORD_FIRST - 2}']
            w.value = f'=IF({L}{ORD_FIRST - 1}="","",{weekday_jp(f"{L}{ORD_FIRST - 1}")})'
            w.font = F_NOTE
            w.alignment = Alignment(horizontal='center')
            a = ws[f'{L}{ORD_FIRST - 3}']
            a.value = f'=IF({L}{ORD_FIRST - 1}="","",{Q_CALC}!{DT[d]}$4)'
            a.number_format = '#,##0'
            a.font = F_SMALL
            a.alignment = Alignment(horizontal='center')
    ws[f'{OC["after"]}{ORD_FIRST - 3}'] = '動員 →'
    ws[f'{OC["after"]}{ORD_FIRST - 3}'].font = F_SMALL
    ws[f'{OC["after"]}{ORD_FIRST - 3}'].alignment = Alignment(horizontal='right')
    put_headers(ws, ORD_FIRST - 1, ['数えた在庫', '数えた日', '商品コード', '仕入先'],
                ['任意', '任意', '自動', '自動'], start_col=PJ_FIRST_COL + DAYS)
    for d, L in enumerate(CAP_COLS):
        h = ws[f'{L}{ORD_FIRST - 1}']
        h.value = f'上限{d}'
        h.font = F_NOTE
    ws[f'{CAP_COLS[0]}{ORD_FIRST - 2}'] = '← 入庫数の上限（自動）。さわらない'
    ws[f'{CAP_COLS[0]}{ORD_FIRST - 2}'].font = F_NOTE
    flag_row = f'{Q_CALC}!{DT[0]}${FLAG_ROW[kind]}'
    ws.conditional_formatting.add(f'{IN_COLS[0]}{ORD_FIRST - 1}:{last_in}{ORD_FIRST - 1}', FormulaRule(
        formula=[f'AND({IN_COLS[0]}${ORD_FIRST - 1}<>"",{flag_row}=1)'],
        fill=PatternFill('solid', bgColor=C_BUTTER), font=Font(name=FONT, size=10, bold=True, color=C_INK)))
    ws.conditional_formatting.add(f'{IN_COLS[0]}{ORD_FIRST - 1}:{last_in}{ORD_FIRST - 1}', FormulaRule(
        formula=[f'AND({IN_COLS[0]}${ORD_FIRST - 1}<>"",{flag_row}=0)'],
        fill=PatternFill('solid', bgColor=C_GREY_SOFT), font=Font(name=FONT, size=10, color=C_TEXT3)))
    ws.conditional_formatting.add(f'{PJ_COLS[0]}{ORD_FIRST - 1}:{last_pj}{ORD_FIRST - 1}', FormulaRule(
        formula=[f'AND({PJ_COLS[0]}${ORD_FIRST - 1}<>"",OR(WEEKDAY({PJ_COLS[0]}${ORD_FIRST - 1})=1,WEEKDAY({PJ_COLS[0]}${ORD_FIRST - 1})=7))'],
        fill=PatternFill('solid', bgColor='FFF3F1EC')))

    # ---- 明細 ----
    seq_col = calc_col(f'seq_{kind}')
    for i in range(N_ITEMS):
        r = ORD_FIRST + i
        j = i + 1
        rc = f'MATCH({j},{seq_col},0)'
        exists = f'{j}<=MAX({seq_col})'

        def take(name):
            return f'=IF({exists},INDEX({calc_col(name)},{rc}),"")'
        ws[f'{OC["no"]}{r}'] = f'=IF({exists},{j},"")'
        for name, cn in (('name', 'name'), ('par', 'par'), ('stock', 'stock'), ('use', 'use_avg'),
                         ('cut', 'cut_before'), ('status', 'status'), ('next', 'next_del'), ('guide', 'guide'),
                         ('unit', 'unit'), ('after', 'after'), ('code', 'code'), ('vendor', 'vendor')):
            ws[f'{OC[name]}{r}'] = take(cn)
        for d, L in enumerate(PJ_COLS):
            ws[f'{L}{r}'] = f'=IF({exists},IF(INDEX({calc_col("pi")},{rc})="","",INDEX({calc_col(f"a{d}")},{rc})),"")'
        for d, L in enumerate(CAP_COLS):
            ws[f'{L}{r}'] = f'=IF({exists},INDEX({calc_col(f"gd{d}")},{rc}),0)'
            ws[f'{L}{r}'].font = F_NOTE

        auto_cols = [OC[n] for n in ('no', 'name', 'par', 'stock', 'use', 'cut', 'status', 'next', 'guide',
                                     'unit', 'after', 'code', 'vendor')] + PJ_COLS
        style_row(ws, r, auto_cols, fill=FILL_AUTO)
        for L in IN_COLS + [OC['counted'], OC['counted_date']]:
            cell = ws[f'{L}{r}']
            cell.fill = FILL_INPUT
            cell.font = F_INPUT
            cell.border = BORDER_ROW
            cell.protection = UNLOCKED
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '0'
        ws[f'{OC["no"]}{r}'].alignment = Alignment(horizontal='center')
        ws[f'{OC["no"]}{r}'].font = F_NOTE2
        ws[f'{OC["par"]}{r}'].number_format = FMT_INT
        ws[f'{OC["stock"]}{r}'].number_format = FMT_INT
        ws[f'{OC["stock"]}{r}'].font = F_BOLD
        ws[f'{OC["use"]}{r}'].number_format = FMT_DEC1
        for n in ('cut', 'next'):
            ws[f'{OC[n]}{r}'].number_format = FMT_MD
            ws[f'{OC[n]}{r}'].alignment = Alignment(horizontal='center')
        ws[f'{OC["cut"]}{r}'].font = F_BOLD
        ws[f'{OC["guide"]}{r}'].number_format = '0'
        ws[f'{OC["guide"]}{r}'].alignment = Alignment(horizontal='center')
        ws[f'{OC["guide"]}{r}'].font = F_BOLD
        ws[f'{OC["unit"]}{r}'].font = F_NOTE2
        ws[f'{OC["counted"]}{r}'].number_format = FMT_INT
        ws[f'{OC["counted_date"]}{r}'].number_format = FMT_MD
        ws[f'{OC["code"]}{r}'].font = F_NOTE
        ws[f'{OC["vendor"]}{r}'].font = F_NOTE
        for L in PJ_COLS:
            ws[f'{L}{r}'].number_format = FMT_GRID
            ws[f'{L}{r}'].alignment = Alignment(horizontal='center')
            ws[f'{L}{r}'].font = Font(name=FONT, size=10, color=C_TEXT2)

    # ---- 色 ----
    R1, RL = ORD_FIRST, ORD_LAST
    status_rules(ws, f'{OC["status"]}{R1}:{OC["status"]}{RL}')
    status_rules(ws, f'{OC["after"]}{R1}:{OC["after"]}{RL}')
    ws.conditional_formatting.add(f'{OC["no"]}{R1}:{OC["cut"]}{RL}', FormulaRule(
        formula=[f'LEFT(${OC["status"]}{R1},1)="－"'], font=Font(name=FONT, size=11, color=C_TEXT3)))
    for mark, color in (('●', C_CRIT), ('△', C_WARN)):
        ws.conditional_formatting.add(f'{OC["cut"]}{R1}:{OC["cut"]}{RL}', FormulaRule(
            formula=[f'LEFT(${OC["status"]}{R1},1)="{mark}"'], font=Font(name=FONT, size=11, bold=True, color=color)))
    ws.conditional_formatting.add(f'{OC["stock"]}{R1}:{OC["stock"]}{RL}', FormulaRule(
        formula=[f'AND(${OC["counted"]}{R1}<>"",ISNUMBER(${OC["counted_date"]}{R1}),${OC["counted_date"]}{R1}>={Q_SET}!$C$6)'],
        fill=PatternFill('solid', bgColor=C_OK_SOFT), font=Font(name=FONT, size=11, bold=True, color=C_OK)))
    ws.conditional_formatting.add(f'{OC["counted"]}{R1}:{OC["counted_date"]}{RL}', FormulaRule(
        formula=[f'AND(${OC["counted"]}{R1}<>"",NOT(AND(ISNUMBER(${OC["counted_date"]}{R1}),${OC["counted_date"]}{R1}>={Q_SET}!$C$6)))'],
        fill=PatternFill('solid', bgColor=C_WARN_SOFT), font=Font(name=FONT, size=11, bold=True, color=C_WARN)))
    ws.conditional_formatting.add(f'{OC["name"]}{R1}:{OC["name"]}{RL}', FormulaRule(
        formula=[f'COUNTIF(${IN_COLS[0]}{R1}:${last_in}{R1},">0")>0'],
        fill=PatternFill('solid', bgColor=C_BUTTER_SOFT), font=Font(name=FONT, size=11, bold=True, color=C_INK)))
    in_rng = f'{IN_COLS[0]}{R1}:{last_in}{RL}'
    ws.conditional_formatting.add(in_rng, FormulaRule(
        formula=[f'{Q_CALC}!{DT[0]}${FLAG_ROW[kind]}=0'],
        fill=PatternFill('solid', bgColor=C_GREY_SOFT), font=Font(name=FONT, size=10, color=C_TEXT3), stopIfTrue=True))
    ws.conditional_formatting.add(in_rng, FormulaRule(
        formula=[f'{IN_COLS[0]}{R1}>0'],
        fill=PatternFill('solid', bgColor=C_BUTTER), font=Font(name=FONT, size=11, bold=True, color=C_INK), stopIfTrue=True))
    ws.conditional_formatting.add(in_rng, FormulaRule(
        formula=[f'AND(${OC["next"]}{R1}<>"",{IN_COLS[0]}${ORD_FIRST - 1}=${OC["next"]}{R1},'
                 f'OR(LEFT(${OC["status"]}{R1},1)="●",LEFT(${OC["status"]}{R1},1)="△"))'],
        fill=PatternFill('solid', bgColor=C_BUTTER_STRONG), stopIfTrue=True))
    pj_rng = f'{PJ_COLS[0]}{R1}:{last_pj}{RL}'
    ws.conditional_formatting.add(pj_rng, FormulaRule(
        formula=[f'AND({PJ_COLS[0]}{R1}<>"",{PJ_COLS[0]}{R1}<1)'],
        fill=PatternFill('solid', bgColor=C_CRIT), font=Font(name=FONT, size=10, bold=True, color=C_WHITE),
        stopIfTrue=True))
    ws.conditional_formatting.add(pj_rng, FormulaRule(
        formula=[f'{IN_COLS[0]}{R1}>0'],
        fill=PatternFill('solid', bgColor=C_BUTTER_SOFT), font=Font(name=FONT, size=10, bold=True, color=C_INK),
        stopIfTrue=True))

    # ---- 入力規則：納品日でない列には入らない、上限は目安 ----
    for d, L in enumerate(IN_COLS):
        dv = DataValidation(type='whole', operator='between', formula1='0', formula2=f'${CAP_COLS[d]}{R1}',
                            showErrorMessage=True, errorTitle='入庫数',
                            error='この日は納品日でないか、定数まで在庫があります（上限 0）。納品日の列に、「目安」までの整数で入れてください。'
                                  '定数を超えて入れるなら商品マスタの定数を増やしてください。',
                            showInputMessage=True, promptTitle='入庫数', prompt='届く日の列に「単位」の数で。迷ったら「目安」の数。')
        ws.add_data_validation(dv)
        dv.add(f'{L}{R1}:{L}{RL}')
    number_validation(ws, f'{OC["counted"]}{R1}:{OC["counted"]}{RL}', 'whole', 0, 999999,
                      '数えた在庫', '個数を整数で入れてください。')
    dv = DataValidation(type='date', operator='between', formula1=f'N({Q_SET}!$C$6)-30', formula2=f'N({Q_SET}!$C$6)',
                        showErrorMessage=True, errorTitle='数えた日',
                        error='数えた日は今日か、それより前の日付です（30日前まで）。')
    ws.add_data_validation(dv)
    dv.add(f'{OC["counted_date"]}{R1}:{OC["counted_date"]}{RL}')

    widths = {'A': 2, 'B': 5, 'C': 30, 'D': 9, 'E': 9, 'F': 9, 'G': 9, 'H': 46, 'I': 9, 'J': 6, 'K': 10, 'L': 60,
              OC['counted']: 10, OC['counted_date']: 10, OC['code']: 15, OC['vendor']: 20}
    for L in IN_COLS + PJ_COLS:
        widths[L] = 6.4
    for L in CAP_COLS:
        widths[L] = 4
    for L, w in widths.items():
        ws.column_dimensions[L].width = w
    ws.freeze_panes = f'D{ORD_FIRST}'
    ws.auto_filter.ref = f'B{ORD_FIRST - 1}:{OC["vendor"]}{ORD_LAST}'
    ws.print_title_rows = f'{ORD_FIRST - 1}:{ORD_FIRST - 1}'
    ws.print_area = f'B2:{last_pj}{ORD_LAST}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.tabColor = KIND_COLOR[kind][2:]
    protect(ws, allow_filter=True)
    return ws


# ---------------------------------------------------------------------------
# ③発注書
# ---------------------------------------------------------------------------
def build_order_sheet(wb):
    ws = wb.create_sheet(S_SHEET)
    app_bar(ws, width=10)
    ws['B2'] = '発 注 書'
    ws['B2'].font = Font(name=FONT, size=22, bold=True, color=C_INK)
    ws.row_dimensions[2].height = 34
    ws['B3'] = 'STEP 2 で入庫を入れた商品が、届く日ごとに1行ずつ集まります（⚠ の行は載りません）。区分と仕入先を選んで印刷してください。'
    ws['B3'].font = F_NOTE2

    labels = [('仕入先', None, FILL_INPUT), ('区分', None, FILL_INPUT), ('発注日', f'={Q_SET}!$C$6', FILL_AUTO),
              ('発注元', f'={Q_SET}!$C$4&"　"&{Q_SET}!$C$5', FILL_AUTO), ('担当者', f'={Q_SET}!$C$9', FILL_AUTO)]
    for i, (label, formula, fill) in enumerate(labels):
        r = 4 + i
        ws[f'B{r}'] = label
        ws[f'B{r}'].font = F_BOLD
        ws[f'B{r}'].fill = FILL_HEAD
        ws[f'B{r}'].border = BORDER_BOX
        ws.merge_cells(f'C{r}:F{r}')
        c = ws[f'C{r}']
        if formula:
            c.value = formula
        c.fill = fill
        c.border = BORDER_BOX
        c.font = F_INPUT if fill is FILL_INPUT else F_BASE
        c.alignment = Alignment(horizontal='left', vertical='center')
    ws['C4'] = 'すべて'
    ws['C5'] = 'すべて'
    ws['C4'].protection = UNLOCKED
    ws['C5'].protection = UNLOCKED
    ws['C6'].number_format = FMT_DATE
    ws['G4'] = '← ここを選ぶ'
    ws['G4'].font = F_NOTE
    ws['G5'] = '← ここを選ぶ'
    ws['G5'].font = F_NOTE
    list_validation(ws, f"='{S_SET}'!$F${VENDOR_FIRST}:$F${VENDOR_LAST}", 'C4', '仕入先',
                    '一覧から選んでください（一覧は 設定 シートの F 列）。')
    list_validation(ws, '"すべて,' + ','.join(KINDS) + '"', 'C5', '区分', 'すべて・冷凍・飲料・常温 から選んでください。')
    ws['I9'] = '合計（税抜）'
    ws['I9'].font = F_BOLD
    ws['I9'].alignment = Alignment(horizontal='right')
    ws['K9'] = f'=SUM($K${SHEET_FIRST}:$K${SHEET_LAST})'
    ws['K9'].font = Font(name=FONT, size=14, bold=True, color=C_INK)
    ws['K9'].number_format = FMT_YEN

    headers = ['No', '仕入先', '区分', '商品コード', '商品名', '入庫日(届く日)', '単位', '発注数', '単価', '金額']
    put_headers(ws, SHEET_FIRST - 1, headers)
    pick = calc_col('pick')
    cnt_block = f"'{S_CALC}'!${CC['cnt0']}${CALC_FIRST}:${CC[f'cnt{DAYS - 1}']}${CALC_LAST}"
    in_block = f"'{S_CALC}'!${CC['in0']}${CALC_FIRST}:${CC[f'in{DAYS - 1}']}${CALC_LAST}"
    dates14 = f"'{S_CALC}'!${DT[0]}$3:${DT[DAYS - 1]}$3"
    for k in range(1, SHEET_LAST - SHEET_FIRST + 2):
        r = SHEET_FIRST + k - 1
        # pick は「その行までの有効な入庫の累計」。k 番目の入庫が属する商品は、累計 < k の行数 + 1
        prow = f'(COUNTIF({pick},"<"&{k})+1)'
        prev = f'IF({prow}=1,0,INDEX({pick},{prow}-1))'
        m = f'({k}-{prev})'
        dpos = f'MATCH({m},INDEX({cnt_block},{prow},0),0)'
        ws[f'B{r}'] = f'=IF({k}<=MAX({pick}),{k},"")'

        def take(name):
            return f'=IF($B{r}="","",IFERROR(INDEX({calc_col(name)},{prow}),""))'
        ws[f'C{r}'] = take('vendor')
        ws[f'D{r}'] = take('kind')
        ws[f'E{r}'] = f'=IF($B{r}="","",IFERROR(INDEX({calc_col("code")},{prow})&"",""))'
        ws[f'F{r}'] = take('name')
        ws[f'G{r}'] = f'=IF($B{r}="","",IFERROR(INDEX({dates14},{dpos}),""))'
        ws[f'H{r}'] = take('unit')
        ws[f'I{r}'] = f'=IF($B{r}="","",IFERROR(INDEX(INDEX({in_block},{prow},0),{dpos}),""))'
        ws[f'J{r}'] = take('price')
        ws[f'K{r}'] = f'=IF($B{r}="","",N($I{r})*IFERROR(INDEX({calc_col("pack")},{prow}),1)*N($J{r}))'
        style_row(ws, r, 'BCDEFGHIJK')
        ws[f'B{r}'].alignment = Alignment(horizontal='center')
        ws[f'D{r}'].alignment = Alignment(horizontal='center')
        ws[f'G{r}'].number_format = FMT_MD
        ws[f'G{r}'].alignment = Alignment(horizontal='center')
        ws[f'I{r}'].number_format = FMT_INT
        ws[f'I{r}'].font = F_BOLD
        ws[f'J{r}'].number_format = FMT_YEN
        ws[f'K{r}'].number_format = FMT_YEN
    for letter, w in (('B', 5), ('C', 22), ('D', 7), ('E', 16), ('F', 30), ('G', 12), ('H', 11),
                      ('I', 9), ('J', 11), ('K', 13)):
        ws.column_dimensions[letter].width = w
    ws.print_area = f'B2:K{SHEET_LAST}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    protect(ws, allow_filter=False)
    return ws


# ---------------------------------------------------------------------------
def load_items(csv_path):
    with open(csv_path, encoding='cp932') as fh:
        rows = list(csv.DictReader(fh))
    seen, items = set(), []
    for r in rows:
        code = r['商品コード'].strip()
        if code and code not in seen:
            seen.add(code)
            items.append(r)
    vendor_count = {}
    categories = []
    for r in rows:
        vendor_count[r['支払先名']] = vendor_count.get(r['支払先名'], 0) + 1
        if r['商品分類名'] not in categories:
            categories.append(r['商品分類名'])
    vendors = sorted(vendor_count, key=lambda v: -vendor_count[v])
    return items, vendors, categories


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('売店発注ツール.xlsx')
    src = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    items, vendors, categories = load_items(src) if src else ([], ['仕入先A'], list(DEFAULT_KIND))

    wb = Workbook()
    wb.remove(wb.active)
    build_home(wb)
    build_csv_sheet(wb, S_CUR, [], '在庫を貼る',
                    '在庫システムから出した「今日の在庫一覧CSV」を貼ります。これが今の在庫になります。',
                    f'CSVの見出し行を除いたデータを、A{CSV_FIRST} セルに貼り付けてください。'
                    f'貼り替えは A{CSV_FIRST} から S 列までを消してから。')
    for kind in KINDS:
        build_category(wb, kind)
    build_order_sheet(wb)
    build_sort(wb, categories)
    build_plan(wb)
    build_intro(wb)
    build_csv_sheet(wb, S_MID, [S_CUR], '1ヶ月前の在庫を貼る（月1回・任意）',
                    '取扱商品が変わっていないかを見るためのものです。1ヶ月前の1日時点の在庫一覧を貼ってください。',
                    f'月に1回、A{CSV_FIRST} セルに貼り替えてください。貼らなくても動きます。')
    build_csv_sheet(wb, S_PRV, [S_CUR, S_MID], '2ヶ月前の在庫を貼る（月1回・任意）',
                    '当月を含まない2ヶ月前の1日時点の在庫一覧を貼ってください。半年ごとに入れ替わる商品もこれで漏れません。',
                    f'月に1回、A{CSV_FIRST} セルに貼り替えてください。貼らなくても動きます。')
    build_item_master(wb, items)
    build_theater_master(wb)
    build_settings(wb, vendors)
    build_calc(wb)

    order = ([S_HOME, S_CUR] + [S_K[k] for k in KINDS]
             + [S_SHEET, S_SORT, S_PLAN, S_INTRO, S_MID, S_PRV, S_ITEM, S_TH, S_SET, S_CALC])
    wb._sheets = [wb[n] for n in order]
    for ws in wb.worksheets:
        if ws.title == S_HOME:
            ws.sheet_properties.tabColor = C_INK[2:]
        elif ws.title in (S_CUR, S_SHEET):
            ws.sheet_properties.tabColor = C_BUTTER[2:]
        elif ws.title in (S_PLAN, S_INTRO, S_SORT):
            ws.sheet_properties.tabColor = C_OK[2:]
        elif ws.title in [S_K[k] for k in KINDS]:
            pass
        else:
            ws.sheet_properties.tabColor = C_TEXT3[2:]
        for dim in ws.column_dimensions.values():
            if dim.width:
                dim.width = round(dim.width * 1.08, 2)
    wb.active = 0
    # 値は保存していない（LibreOffice で再計算して保存するとフォントとリンクが書き換わるため）。
    # Excel が開くときに全部計算する。
    wb.calculation.fullCalcOnLoad = True
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f'saved: {out}  sheets={len(wb.worksheets)}  items={len(items)}')


if __name__ == '__main__':
    main()
