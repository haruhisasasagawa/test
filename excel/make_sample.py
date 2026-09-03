#!/usr/bin/env python3
"""
動作確認用のサンプルデータを入れる。

  python3 make_sample.py [元のxlsx] [出力xlsx]

3劇場ぶんの在庫CSV（当日・1ヶ月前・2ヶ月前）、商品マスタの定数と減り数、
動員の予測と実績、そして「入庫を入れたらどう見えるか」の例を入れる。
入庫の例は、一度再計算して「このままだと」を読んでから、● の行に入れる。
実運用と同じ順番（CSVを貼る → アラートを見る → 入庫を入れる）で作っている。
"""
import csv
import datetime
import random
import subprocess
import sys
from pathlib import Path

import openpyxl

random.seed(7)
HERE = Path(__file__).resolve().parent
S = '/tmp/claude-0/-home-user-test/6ebcf1c1-1102-5af1-96f9-c94aebb0a201/scratchpad'
SRC = sys.argv[1] if len(sys.argv) > 1 else str(HERE / '売店発注ツール.xlsx')
OUT = sys.argv[2] if len(sys.argv) > 2 else str(HERE / '売店発注ツール_サンプルデータ入り.xlsx')
RECALC = '/root/.claude/skills/xlsx/scripts/recalc.py'

H = ['対象日付', '劇場コード', '劇場名', '支払先コード', '支払先名', '大分類コード',
     '小分類コード', '商品分類名', '商品コード', '商品名', '入数', '在庫数ケース',
     '在庫数バラ', '総数', '規定数', '資産廃棄', '税抜単価', '仕入税区分', '仕入税率']

rows = list(csv.DictReader(open(S + '/stock.csv', encoding='cp932')))
BASE = datetime.date(2026, 5, 1)          # 当日基準日
MID = datetime.date(2026, 4, 1)           # 1ヶ月前の1日
PREV = datetime.date(2026, 3, 1)          # 2ヶ月前の1日

THEATERS = [
    # (コード, 名前, 取扱商品の割合, 在庫の潤沢さ)
    ('0761', '新宿', 1.00, 1.0),
    ('0841', '池袋', 0.72, 1.4),
    ('0681', '上田', 0.55, 2.2),
]

ORD_FIRST = 10
CALC_FIRST = 6


def put(ws, i, r):
    for j, h in enumerate(H):
        v = r[h]
        if h in ('入数', '在庫数ケース', '在庫数バラ', '総数', '資産廃棄', '規定数'):
            v = int(v) if str(v).strip() else None
        elif h in ('税抜単価', '仕入税率'):
            v = float(v) if str(v).strip() else None
        elif h == '対象日付':
            v = int(v)
        ws.cell(i, 1 + j, v)


def row_for(date, code, name, src, total):
    p = dict(src)
    ins = max(1, int(src['入数']))
    p['対象日付'] = date.strftime('%Y%m%d')
    p['劇場コード'] = code
    p['劇場名'] = name
    p['在庫数ケース'] = total // ins
    p['在庫数バラ'] = total - (total // ins) * ins
    p['総数'] = total
    return p


def snapshots(code, name, items, scale):
    cur, mid, prv = [], [], []
    for src in items:
        base = max(1, int(src['総数']))
        today = int(base * random.choice([0.05, 0.2, 0.4, 0.7, 1.0, 1.5, 2.2]) * scale)
        if random.random() < 0.04:
            today = -random.randint(1, 30)            # 理論値がマイナスの商品もたまにある
        used1 = int(base * random.uniform(0.4, 1.3))
        used2 = int(base * random.uniform(0.4, 1.3))
        m = max(0, today) + used1
        cur.append(row_for(BASE, code, name, src, today))
        mid.append(row_for(MID, code, name, src, m))
        prv.append(row_for(PREV, code, name, src, m + used2))
    return cur, mid, prv


def variant(items, idx, suffix, label, qty):
    src = dict(items[idx])
    src.update({'商品コード': suffix, '商品名': label, '入数': '1'})
    return src, qty


wb = openpyxl.load_workbook(SRC)
ws_cur = wb['①在庫を貼る']
ws_mid = wb['月1回_1ヶ月前の在庫']
ws_prv = wb['月1回_2ヶ月前の在庫']

cur_rows, mid_rows, prv_rows = [], [], []
for code, name, share, scale in THEATERS:
    items = rows[:int(len(rows) * share)]
    c, m, p = snapshots(code, name, items, scale)
    churn = [
        ('（新商品）夏季限定フローズン', True, False, False),
        ('（新商品）コラボカップ', True, True, False),
        ('（終売）季節限定ドリンク', False, True, True),
        ('（終売）旧パッケージカップ', False, False, True),
        ('（復活）冬季限定ホットスナック', True, False, True),
    ]
    for n, (label, a, b, d) in enumerate(churn):
        src, qty = variant(items, 3 + n, f'9{n}{code}000001', label, 40 + n * 25)
        if a:
            c.append(row_for(BASE, code, name, src, qty))
        if b:
            m.append(row_for(MID, code, name, src, qty + 30))
        if d:
            p.append(row_for(PREV, code, name, src, qty + 60))
    cur_rows += c
    mid_rows += m
    prv_rows += p
for ws_, rows_ in ((ws_cur, cur_rows), (ws_mid, mid_rows), (ws_prv, prv_rows)):
    for i, r in enumerate(rows_):
        put(ws_, 6 + i, r)

# ---- 商品マスタ：定数（ケース）と100人あたりの減り数 ----
# 減り数は「新宿の今日の在庫が何日でなくなるか」から逆算する。
# こうすると ●（数日で切れる）△（1〜2週間）○（余裕）が混ざって、画面の見え方を確認できる。
im = wb['商品マスタ']
today_stock = {r['商品コード']: int(r['総数']) for r in cur_rows if r['劇場コード'] == '0761'}
AVG_ATT = 2400          # 動員の平均/日（下の weekday_base の平均に近い値）
for i in range(len(rows)):
    r = 6 + i
    code = str(im.cell(r, 2).value)
    pack = int(im.cell(r, 6).value or 1)
    stock = max(1, today_stock.get(code, 1))
    days = random.choice([1, 2, 3, 4, 5, 6, 8, 10, 12, 16, 20, 28, 40])
    if random.random() < 0.9:
        im.cell(r, 10, max(0.01, round(stock / days / (AVG_ATT / 100), 2)))    # J 減り数
    if random.random() < 0.85:
        cases = stock / pack
        im.cell(r, 8, max(1, int(cases * random.uniform(1.3, 2.5) + 0.999)))   # H 定数（ケース）

# ---- 動員：3/1 から。過去は実績、これからは予測 ----
plan = wb['動員を入れる']
weekday_base = {0: 2600, 1: 1050, 2: 980, 3: 1150, 4: 1250, 5: 1750, 6: 3000}   # 月=0
titles = {
    BASE: 'GW初日', BASE + datetime.timedelta(days=1): 'GW', BASE + datetime.timedelta(days=2): 'GW',
    BASE + datetime.timedelta(days=3): 'GW', BASE + datetime.timedelta(days=4): 'GWこどもの日／ファミリー作品',
    datetime.date(2026, 5, 15): '大型アニメ作品 公開初日', datetime.date(2026, 5, 16): '大型アニメ作品 公開2日目',
    datetime.date(2026, 5, 17): '大型アニメ作品 公開3日目', datetime.date(2026, 5, 22): '洋画大作 公開',
    datetime.date(2026, 6, 3): '映画の日', datetime.date(2026, 3, 20): '春休み作品 公開',
    datetime.date(2026, 4, 3): '話題作 公開',
}
plan['C5'] = PREV
for i in range(220):
    d = PREV + datetime.timedelta(days=i)
    r = 16 + i
    base = weekday_base[d.weekday()]
    if d in titles:
        base = int(base * random.uniform(1.5, 2.3))
    if d < BASE:
        actual = int(base * random.uniform(0.88, 1.16))
        plan.cell(r, 5, actual)                                  # E 実績
        if i % 2 == 0 or d in titles:
            plan.cell(r, 4, int(actual * random.uniform(0.85, 1.15)))   # D 予測
    elif i < 61 + 21 or d in titles:
        plan.cell(r, 4, int(base * random.uniform(0.9, 1.12)))
    if d in titles:
        plan.cell(r, 6, titles[d])

wb['設定']['C9'] = '笹川'
print(f'sample built: 当日{len(cur_rows)}行 / 1ヶ月前{len(mid_rows)}行 / 2ヶ月前{len(prv_rows)}行')

# ---- 仕分け：納品日と発注から納品までの日数 ----
so = wb['仕分け']
DELIV = {'冷凍': ('月', '水', '金'), '飲料': ('月', '火', '水', '木', '金', '土'), '常温': ('木',)}
LT = {'冷凍': 2, '飲料': 1, '常温': 3}
WD = '月火水木金土日'
for kind, row in (('冷凍', 6), ('飲料', 7), ('常温', 8)):
    for i, w in enumerate(WD):
        so.cell(row, 3 + i, '○' if w in DELIV[kind] else None)
    so.cell(row, 10, LT[kind])
# 商品ごとの上書きの例：1件だけ常温の商品を冷凍に
so.cell(36 + 3, 6, '冷凍')

wb.save(OUT)


# ---- 入庫の例：一度計算して「このままだと」を読んでから入れる ----
def recalc(path, tries=3):
    """LibreOffice はまれに起動でつまずいて時間切れになるので、何度か試す。"""
    import time
    for _ in range(tries):
        lock = Path(path).parent / f'.~lock.{Path(path).name}#'
        if lock.exists():
            lock.unlink()
        out = subprocess.run([sys.executable, RECALC, path, '240'], capture_output=True, text=True).stdout
        if '"total_formulas"' in out:
            return True
        time.sleep(3)
    raise SystemExit('再計算に失敗しました: ' + out[-300:])


# 再計算は控えのコピーで行う。納品するファイル（OUT）は openpyxl が書いたままにしておく。
import shutil
TMP = str(Path(S) / 'sample_calc.xlsx')
shutil.copy(OUT, TMP)
recalc(TMP)
vals = openpyxl.load_workbook(TMP, data_only=True)
wb2 = openpyxl.load_workbook(OUT)
demo = []
IN_FIRST = 13
for kind in ('冷凍', '飲料', '常温'):
    ov = vals[f'②{kind}']
    o = wb2[f'②{kind}']
    hdr_dates = [ov.cell(9, IN_FIRST + d).value for d in range(14)]
    hdr_dates = [h.date() if hasattr(h, 'date') else h for h in hdr_dates]
    rows = []
    for r in range(ORD_FIRST, ORD_FIRST + 400):
        s_ = ov.cell(r, 8).value or ''
        nxt = ov.cell(r, 9).value
        g = ov.cell(r, 10).value
        if not s_:
            break
        rows.append((r, s_, nxt.date() if hasattr(nxt, 'date') else nxt, g))
    crit = [x for x in rows if x[1].startswith('●') and isinstance(x[3], (int, float)) and x[3] > 0 and x[2] in hdr_dates]
    warn = [x for x in rows if x[1].startswith('△') and isinstance(x[3], (int, float)) and x[3] > 0 and x[2] in hdr_dates]
    # ● の行：次の便に目安どおり
    for r, _, nxt, g in crit[:4]:
        o.cell(r, IN_FIRST + hdr_dates.index(nxt), int(g))
        demo.append((f'{kind} 目安どおり', r))
    # ● の行：次の便に少なすぎる数
    if len(crit) > 4:
        r, _, nxt, g = crit[4]
        o.cell(r, IN_FIRST + hdr_dates.index(nxt), 1)
        demo.append((f'{kind} 少ない', r))
    # △ の行：納品日でない日に入れてしまう → ⚠
    if warn:
        r, _, nxt, g = warn[0]
        idx = hdr_dates.index(nxt)
        for d in range(idx + 1, 14):
            if ov.cell(r, 45 + d).value == 0:     # 上限0＝納品日でない
                o.cell(r, IN_FIRST + d, int(g))
                demo.append((f'{kind} 納品日でない', r))
                break
    # △ の行：目安より多い → ⚠
    if len(warn) > 1:
        r, _, nxt, g = warn[1]
        o.cell(r, IN_FIRST + hdr_dates.index(nxt), int(g) + 3)
        demo.append((f'{kind} 定数超え', r))
    # △ の行：2便に分けて入れる
    if len(warn) > 2:
        r, _, nxt, g = warn[2]
        idx = hdr_dates.index(nxt)
        o.cell(r, IN_FIRST + idx, max(1, int(g) // 2))
        for d in range(idx + 1, 14):
            if ov.cell(r, 45 + d).value not in (0, None):
                o.cell(r, IN_FIRST + d, max(1, int(g) // 2))
                break
        demo.append((f'{kind} 2便', r))
    # 数えた在庫
    ok_rows = [x for x in rows if x[1].startswith('○')]
    for r, *_ in ok_rows[:1]:
        o.cell(r, 41, 5)
        o.cell(r, 42, BASE)
        demo.append((f'{kind} 数えた在庫', r))
    if len(ok_rows) > 1:
        r = ok_rows[1][0]
        o.cell(r, 41, 40)
        o.cell(r, 42, BASE - datetime.timedelta(days=9))
        demo.append((f'{kind} 数えた日が古い', r))
wb2.calculation.fullCalcOnLoad = True
wb2.save(OUT)
print('入庫の例:', ', '.join(f'{k}=行{r}' for k, r in demo))
