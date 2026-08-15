"""月次予算割振ブックを生成する。

構成
  1. 設定          … 年・月を変えると週の区切り（金曜〜木曜）が入れ替わる
  2. 動員予測      … 日別（または週別）の動員予測 → 週ごとの構成比
  3. 予算割振      … 月予算を週別動員の構成比で自動配分（メイン）
  4. 明細（実施予定）… 週ごとの回数から積み上げたい項目用
  5. 接客部門      … 日別の予測動員から部門別の人件費を自動計算（動員帯マスター方式）
  6. 有休          … 有休計算表
  7. 人件費マスター … 動員帯別の部門別人件費テーブル
"""

import json
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Meiryo UI"
BLUE = "0000FF"       # 入力値
BLACK = "000000"      # 数式
YELLOW = "FFF2CC"     # 入力セル
GRAY = "F2F2F2"       # 見出し
HEAD = "D9E2F3"       # 表ヘッダ
TOTALFILL = "E2EFDA"  # 合計行
RATIOFILL = "FCE4D6"  # 構成比の行

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

YEN = '#,##0;[Red]-#,##0;"-"'
PCT = '0.0%;[Red]-0.0%;"-"'
PCT2 = '0.00%;[Red]-0.00%;"-"'
NUM = '#,##0.##;[Red]-#,##0.##;"-"'

wb = openpyxl.Workbook()


def style(ws, ref, *, bold=False, size=11, color=BLACK, fill=None, fmt=None,
          align=None, border=False, wrap=False):
    rng = ws[ref]
    rows = rng if isinstance(rng, tuple) else ((rng,),)
    if rows and not isinstance(rows[0], tuple):
        rows = (rows,)
    for row in rows:
        for c in row:
            c.font = Font(name=FONT, size=size, bold=bold, color=color)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
            if fmt:
                c.number_format = fmt
            if align or wrap:
                c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
            if border:
                c.border = BORDER


def put(ws, ref, value, **kw):
    ws[ref] = value
    style(ws, ref, **kw)


def title(ws, text, sub=None):
    put(ws, "A1", text, bold=True, size=14)
    if sub:
        put(ws, "A2", sub, size=9, color="808080")


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# =====================================================================
# 1. 設定
# =====================================================================
s = wb.active
s.title = "設定"
title(s, "月次予算割振",
      "対象月を変えるだけで、日付・週の区切り（金曜〜木曜）・WEEK番号がすべて入れ替わります。")

put(s, "A4", "■ 対象月", bold=True, fill=GRAY)
for row, (label, val) in enumerate([("年", 2026), ("月", 4)], start=5):
    put(s, f"A{row}", label, bold=True)
    put(s, f"B{row}", val, color=BLUE, fill=YELLOW, border=True, align="center")
put(s, "A7", "開始WEEK", bold=True)
put(s, "B7", "=B16", color=BLUE, fill=YELLOW, border=True, align="center")
put(s, "C7", "← 自動計算値が入っています。別の番号にしたいときは直接上書きしてください。",
    size=9, color="808080")

put(s, "A9", "■ 自動計算（変更しないでください）", bold=True, fill=GRAY)
auto = [
    ("月初", "=DATE(B5,B6,1)", "yyyy/m/d"),
    ("月末", "=EOMONTH(B11,0)", "yyyy/m/d"),
    ("第1週の木曜", "=B11+MOD(4-WEEKDAY(B11,2)+7,7)", "yyyy/m/d"),
    ("3/1(当年)の週の木曜", "=DATE(B5,3,1)+MOD(4-WEEKDAY(DATE(B5,3,1),2)+7,7)", "yyyy/m/d"),
    ("3/1(前年)の週の木曜", "=DATE(B5-1,3,1)+MOD(4-WEEKDAY(DATE(B5-1,3,1),2)+7,7)", "yyyy/m/d"),
    ("開始WEEK（自動）", "=IF(B13>=B14,(B13-B14)/7+1,(B13-B15)/7+1)", "0"),
    ("週数", "=IF(B12<=B13,1,1+CEILING((B12-B13)/7,1))", "0"),
]
for i, (label, f, fmt) in enumerate(auto, start=11):
    put(s, f"A{i}", label)
    put(s, f"B{i}", f, fmt=fmt, align="center", border=True)

put(s, "A19", "■ 週の区切り（金曜〜木曜。月初・月末は端数の週になります）", bold=True, fill=GRAY)
for col, label in zip("ABCDE", ["週", "WEEK", "開始", "終了", "日数"]):
    put(s, f"{col}20", label, bold=True, fill=HEAD, align="center", border=True)
for i in range(1, 7):
    r = 20 + i
    put(s, f"A{r}", i, align="center", border=True)
    put(s, f"B{r}", f'=IF($A{r}<=$B$17,"W"&($B$7+$A{r}-1),"")', align="center", border=True)
    put(s, f"C{r}", f'=IF($A{r}<=$B$17,IF($A{r}=1,$B$11,$B$13+7*($A{r}-2)+1),"")',
        fmt="m/d", align="center", border=True)
    put(s, f"D{r}", f'=IF($A{r}<=$B$17,MIN($B$12,$B$13+7*($A{r}-1)),"")',
        fmt="m/d", align="center", border=True)
    put(s, f"E{r}", f'=IF($A{r}<=$B$17,$D{r}-$C{r}+1,"")', align="center", border=True)

put(s, "A28", "■ 部門別の人件費について", bold=True, fill=GRAY)
put(s, "A29", "部門別（ボックス／コンセ／フロア／ストア／オフィス）の人件費は、",
    size=9, color="808080")
put(s, "A30", "「人件費マスター」シートの動員帯別テーブルから、日別の予測動員に応じて自動で引きます。",
    size=9, color="808080")
put(s, "A31", "計算結果は「接客部門」シートに出ます。", size=9, color="808080")

put(s, "A35", "■ 凡例", bold=True, fill=GRAY)
put(s, "A36", "入力するセル", fill=YELLOW, color=BLUE, border=True)
put(s, "B36", "黄色 ＝ 手で入力するところ（青字）", size=9, color="808080")
put(s, "A37", "自動計算", border=True)
put(s, "B37", "白 ＝ 数式が入っています（触らないでください）", size=9, color="808080")
put(s, "A39", "※ 初期値は 2026年4月シートから写しています。実態に合わせて書き換えてください。",
    size=9, color="808080")
widths(s, {"A": 24, "B": 14, "C": 54, "D": 12, "E": 8})

# =====================================================================
# 2. 動員予測
# =====================================================================
a = wb.create_sheet("動員予測")
title(a, "動員予測",
      "日別の予測を入れると、金曜〜木曜の週ごとに合計され、その月の何％にあたるかが出ます。"
      "週単位の予測しかない場合は右の表のK列に直接入れてください。")

for col, label in zip("ABCDE", ["日付", "曜日", "WEEK", "動員予測", "日別構成比"]):
    put(a, f"{col}3", label, bold=True, fill=HEAD, align="center", border=True)
for i in range(31):
    r = 4 + i
    put(a, f"A{r}", f'=IF({i + 1}<=DAY(設定!$B$12),設定!$B$11+{i},"")',
        fmt="m/d", align="center", border=True)
    put(a, f"B{r}", f'=IF($A{r}="","",CHOOSE(WEEKDAY($A{r},2),"月","火","水","木","金","土","日"))',
        align="center", border=True)
    put(a, f"C{r}", f'=IF($A{r}="","",IF($A{r}<=設定!$B$13,1,1+CEILING(($A{r}-設定!$B$13)/7,1)))',
        align="center", border=True)
    put(a, f"D{r}", None, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
    put(a, f"E{r}", f'=IF($A{r}="","",IF($D$35=0,0,$D{r}/$D$35))', fmt=PCT2, border=True)
put(a, "A35", "合計", bold=True, fill=TOTALFILL, border=True)
for col in "BC":
    put(a, f"{col}35", None, fill=TOTALFILL, border=True)
put(a, "D35", "=SUM(D4:D34)", bold=True, fill=TOTALFILL, fmt=YEN, border=True)
put(a, "E35", '=IF($D$35=0,"",SUM(E4:E34))', bold=True, fill=TOTALFILL, fmt=PCT2, border=True)

put(a, "G2", "■ 週別（金曜〜木曜）", bold=True, fill=GRAY)
for col, label in zip("GHIJKLM",
                      ["週", "WEEK", "期間", "日別からの合計", "週で直接入力", "動員予測", "構成比"]):
    put(a, f"{col}3", label, bold=True, fill=HEAD, align="center", border=True, wrap=True)
for i in range(1, 7):
    r = 3 + i
    sr = 20 + i  # 設定シートの週テーブル行
    put(a, f"G{r}", i, align="center", border=True)
    put(a, f"H{r}", f"=設定!$B${sr}", align="center", border=True)
    put(a, f"I{r}", f'=IF(設定!$B${sr}="","",TEXT(設定!$C${sr},"m/d")&"〜"&TEXT(設定!$D${sr},"m/d"))',
        align="center", border=True)
    put(a, f"J{r}", f'=IF(設定!$B${sr}="","",SUMIF($C$4:$C$34,$G{r},$D$4:$D$34))',
        fmt=YEN, border=True)
    put(a, f"K{r}", None, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
    put(a, f"L{r}", f'=IF(設定!$B${sr}="","",IF($K{r}="",$J{r},$K{r}))',
        bold=True, fmt=YEN, border=True)
    put(a, f"M{r}", f'=IF(設定!$B${sr}="","",IF($L$10=0,0,$L{r}/$L$10))',
        bold=True, fmt=PCT2, border=True)
put(a, "I10", "合計", bold=True, fill=TOTALFILL, align="center", border=True)
put(a, "J10", "=SUM(J4:J9)", bold=True, fill=TOTALFILL, fmt=YEN, border=True)
put(a, "K10", "=SUM(K4:K9)", bold=True, fill=TOTALFILL, fmt=YEN, border=True)
put(a, "L10", "=SUM(L4:L9)", bold=True, fill=TOTALFILL, fmt=YEN, border=True)
put(a, "M10", '=IF($L$10=0,"",SUM(M4:M9))', bold=True, fill=TOTALFILL, fmt=PCT2, border=True)
put(a, "G12", "※ K列（週で直接入力）に数字を入れると、その週は日別の合計ではなくK列の値を使います。",
    size=9, color="808080")
put(a, "G13", "※ M列の構成比が「予算割振」シートの割振に使われます。",
    size=9, color="808080")
widths(a, {"A": 9, "B": 7, "C": 7, "D": 12, "E": 11, "F": 3,
           "G": 6, "H": 8, "I": 15, "J": 13, "K": 12, "L": 13, "M": 10})
a.freeze_panes = "A4"

# =====================================================================
# 3. 予算割振（メイン）
# =====================================================================
b = wb.create_sheet("予算割振")
title(b, "予算割振（動員予測の構成比で自動配分）",
      "月予算（B列）を入れると、その月の週ごとの動員構成比に応じて金額が割り振られます。"
      "端数は最終週で調整するので、配分計は必ず月予算と一致します。")

put(b, "A4", "月動員予測", bold=True)
put(b, "B4", "=動員予測!$L$10", fmt=YEN, border=True)
put(b, "C4", "週数", bold=True, align="center")
put(b, "D4", "=設定!$B$17", align="center", border=True)

HEADR = 6
RATIOR = 7
ITEM_FIRST = 8
ITEM_LAST = 27
TOTR = 29
W1 = 4          # 週の先頭列（D列）
WCOLS = [get_column_letter(W1 + i) for i in range(6)]   # D..I
SUMC, DETC, DIFC, NOTEC = "J", "K", "L", "M"

put(b, f"A{HEADR}", "項目", bold=True, fill=HEAD, align="center", border=True)
put(b, f"B{HEADR}", "月予算", bold=True, fill=HEAD, align="center", border=True)
put(b, f"C{HEADR}", "割振方法", bold=True, fill=HEAD, align="center", border=True)
for i, col in enumerate(WCOLS):
    put(b, f"{col}{HEADR}", f"=設定!$B${21 + i}", bold=True, fill=HEAD, align="center", border=True)
for col, label in zip([SUMC, DETC, DIFC, NOTEC], ["配分計", "明細合計", "差（予算−明細）", "備考"]):
    put(b, f"{col}{HEADR}", label, bold=True, fill=HEAD, align="center", border=True, wrap=True)

put(b, f"A{RATIOR}", "動員構成比", bold=True, fill=RATIOFILL, border=True)
for col in ("B", "C"):
    put(b, f"{col}{RATIOR}", None, fill=RATIOFILL, border=True)
for i, col in enumerate(WCOLS):
    put(b, f"{col}{RATIOR}", f'=IF({col}${HEADR}="","",動員予測!$M${4 + i})',
        bold=True, fill=RATIOFILL, fmt=PCT2, border=True)
put(b, f"{SUMC}{RATIOR}", f'=IF($B$4=0,"",SUM({WCOLS[0]}{RATIOR}:{WCOLS[5]}{RATIOR}))',
    bold=True, fill=RATIOFILL, fmt=PCT2, border=True)
for col in (DETC, DIFC, NOTEC):
    put(b, f"{col}{RATIOR}", None, fill=RATIOFILL, border=True)

# 項目名と既定の割振方法
ITEMS = [
    ("接客部門", "動員比率"),
    ("障がい者雇用", "動員比率"),
    ("新規採用研修（トレーニー）", "動員比率"),
    ("新規採用研修（トレーナー）", "動員比率"),
    ("既存スタッフ研修", "動員比率"),
    ("防災訓練", "動員比率"),
    ("棚卸（予算分）", "動員比率"),
    ("ストア準備・返品（予算分）", "動員比率"),
    ("ストア準備・返品（追加分）", "動員比率"),
    ("その他", "動員比率"),
    ("リーダー手当", "最終週"),
]
LEAVE_ROW = ITEM_FIRST + len(ITEMS)  # 有給休暇の行

for idx in range(ITEM_LAST - ITEM_FIRST + 1):
    r = ITEM_FIRST + idx
    is_leave = (r == LEAVE_ROW)
    name = ITEMS[idx][0] if idx < len(ITEMS) else ("有給休暇" if is_leave else None)
    method = ITEMS[idx][1] if idx < len(ITEMS) else None

    put(b, f"A{r}", name, color=BLACK if is_leave else BLUE,
        fill=None if is_leave else YELLOW, border=True)
    if is_leave:
        put(b, f"B{r}", "=有休!$B$5", fmt=YEN, border=True)
        put(b, f"C{r}", "有休シート", align="center", border=True)
    else:
        put(b, f"B{r}", None, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
        put(b, f"C{r}", method if name else "動員比率",
            color=BLUE, fill=YELLOW, align="center", border=True)

    for i, col in enumerate(WCOLS):
        nxt = WCOLS[i + 1] if i < 5 else None
        if is_leave:
            f = f'=IF({col}${HEADR}="","",有休!$E${8 + i})'
        elif i == 0:
            # 最初の週。ここが最終週になるのは「週数=1」のときだけ
            f = (f'=IF(OR($A{r}="",{col}${HEADR}=""),"",'
                 f'IF({nxt}${HEADR}="",$B{r},'
                 f'IF($C{r}="最終週",0,'
                 f'IF($C{r}="均等",ROUND($B{r}/$D$4,0),ROUND($B{r}*{col}${RATIOR},0)))))')
        elif i < 5:
            prev_range = f'${WCOLS[0]}{r}:{WCOLS[i - 1]}{r}'
            f = (f'=IF(OR($A{r}="",{col}${HEADR}=""),"",'
                 f'IF({nxt}${HEADR}="",$B{r}-SUM({prev_range}),'
                 f'IF($C{r}="最終週",0,'
                 f'IF($C{r}="均等",ROUND($B{r}/$D$4,0),ROUND($B{r}*{col}${RATIOR},0)))))')
        else:
            prev_range = f'${WCOLS[0]}{r}:{WCOLS[4]}{r}'
            f = (f'=IF(OR($A{r}="",{col}${HEADR}=""),"",$B{r}-SUM({prev_range}))')
        put(b, f"{col}{r}", f, fmt=YEN, border=True)

    put(b, f"{SUMC}{r}", f'=IF($A{r}="","",SUM({WCOLS[0]}{r}:{WCOLS[5]}{r}))',
        bold=True, fmt=YEN, border=True)
    if is_leave:
        put(b, f"{DETC}{r}", None, border=True)
        put(b, f"{DIFC}{r}", None, border=True)
        put(b, f"{NOTEC}{r}", "有休シートの金額を使用", size=9, color="808080", border=True)
    else:
        put(b, f"{DETC}{r}", f'=IF($A{r}="","",SUMIF(明細!$A$5:$A$104,$A{r},明細!$M$5:$M$104))',
            fmt=YEN, border=True)
        put(b, f"{DIFC}{r}", f'=IF(OR($A{r}="",{DETC}{r}=0),"",$B{r}-{DETC}{r})',
            fmt=YEN, border=True)
        note = "月末の週にまとめて計上" if method == "最終週" else None
        put(b, f"{NOTEC}{r}", note, color=BLUE, fill=YELLOW, size=9, border=True)

# 割振方法の入力規則（プルダウン）
dv = DataValidation(type="list", formula1='"動員比率,最終週,均等"', allow_blank=True,
                    showDropDown=False)
dv.error = "「動員比率」「最終週」「均等」から選んでください"
dv.errorTitle = "割振方法"
b.add_data_validation(dv)
dv.add(f"C{ITEM_FIRST}:C{LEAVE_ROW - 1}")
dv.add(f"C{LEAVE_ROW + 1}:C{ITEM_LAST}")

put(b, f"A{TOTR}", "合計", bold=True, fill=TOTALFILL, border=True)
put(b, f"B{TOTR}", f"=SUM($B${ITEM_FIRST}:$B${ITEM_LAST})", bold=True, fill=TOTALFILL,
    fmt=YEN, border=True)
put(b, f"C{TOTR}", None, fill=TOTALFILL, border=True)
for col in WCOLS:
    put(b, f"{col}{TOTR}", f'=IF({col}${HEADR}="","",SUM({col}${ITEM_FIRST}:{col}${ITEM_LAST}))',
        bold=True, fill=TOTALFILL, fmt=YEN, border=True)
put(b, f"{SUMC}{TOTR}", f"=SUM(${SUMC}${ITEM_FIRST}:${SUMC}${ITEM_LAST})", bold=True,
    fill=TOTALFILL, fmt=YEN, border=True)
put(b, f"{DETC}{TOTR}", f"=SUM(${DETC}${ITEM_FIRST}:${DETC}${ITEM_LAST})", bold=True,
    fill=TOTALFILL, fmt=YEN, border=True)
put(b, f"{DIFC}{TOTR}", f"=$B{TOTR}-${DETC}{TOTR}", bold=True, fill=TOTALFILL,
    fmt=YEN, border=True)
put(b, f"{NOTEC}{TOTR}", None, fill=TOTALFILL, border=True)

put(b, f"A{TOTR + 2}", "■ チェック", bold=True, fill=GRAY)
put(b, f"A{TOTR + 3}", "配分計 − 月予算")
put(b, f"B{TOTR + 3}", f"=${SUMC}{TOTR}-$B{TOTR}", fmt=YEN, border=True)
put(b, f"C{TOTR + 3}", "← 0 なら、割り振った金額の合計が月予算とぴったり一致しています",
    size=9, color="808080")
put(b, f"A{TOTR + 4}", "明細シートの金額計")
put(b, f"B{TOTR + 4}", "=明細!$M$106", fmt=YEN, border=True)
put(b, f"A{TOTR + 5}", "うち上の表に載った額")
put(b, f"B{TOTR + 5}", f"=${DETC}{TOTR}", fmt=YEN, border=True)
put(b, f"A{TOTR + 6}", "差（項目名が一致していない額）", bold=True)
put(b, f"B{TOTR + 6}", f"=$B{TOTR + 4}-$B{TOTR + 5}", bold=True, fmt=YEN, border=True)
put(b, f"C{TOTR + 6}", "← 0 以外なら、明細シートの項目名がこの表の項目名と一致していません",
    size=9, color="808080")

put(b, f"A{TOTR + 8}", "■ 割振方法（C列）", bold=True, fill=GRAY)
put(b, f"A{TOTR + 9}", "動員比率")
put(b, f"B{TOTR + 9}", "その週の動員構成比に応じて配分（既定）", size=9, color="808080")
put(b, f"A{TOTR + 10}", "最終週")
put(b, f"B{TOTR + 10}", "その月の一番最後の週にまとめて計上（リーダー手当など）",
    size=9, color="808080")
put(b, f"A{TOTR + 11}", "均等")
put(b, f"B{TOTR + 11}", "週数で等分（端数は最終週で調整）", size=9, color="808080")
put(b, f"A{TOTR + 12}",
    "※ どの方法でも端数は最終週で調整するので、配分計は必ず月予算と一致します。",
    size=9, color="808080")

widths(b, {"A": 28, "B": 14, "C": 11, **{col: 13 for col in WCOLS},
           SUMC: 14, DETC: 13, DIFC: 15, NOTEC: 22})
b.freeze_panes = f"{WCOLS[0]}{ITEM_FIRST}"

# =====================================================================
# 4. 明細（実施予定）
# =====================================================================
p = wb.create_sheet("明細")
title(p, "明細（実施予定）",
      "週ごとの回数から金額を積み上げたい項目に使います。金額 ＝ 単価 × 時間 × 人数 × 回数。"
      "使わない場合は空のままで構いません。")

p.merge_cells("F3:K3")
put(p, "F3", "週ごとの回数（日数）を入力", bold=True, fill=HEAD, align="center", border=True)
p.merge_cells("N3:S3")
put(p, "N3", "週別金額（自動計算）", bold=True, fill=GRAY, align="center", border=True)

for col, label in {"A": "項目", "B": "内容", "C": "単価(時給)", "D": "時間", "E": "人数",
                   "L": "回数計", "M": "金額計"}.items():
    put(p, f"{col}4", label, bold=True, fill=HEAD, align="center", border=True, wrap=True)
for i in range(6):
    wk = get_column_letter(6 + i)     # F..K
    amt = get_column_letter(14 + i)   # N..S
    ref = f"設定!$B${21 + i}"
    put(p, f"{wk}4", f"={ref}", bold=True, fill=HEAD, align="center", border=True)
    put(p, f"{amt}4", f"={ref}", bold=True, fill=GRAY, align="center", border=True)

DETAIL = [
    ("障がい者雇用", "駒澤さん", 1236, 5, 1, [3, 3, 3, 3, 3, None]),
    ("新規採用研修（トレーニー）", "週2〜3名", 1226, 5.5, 2, [None] * 6),
    ("新規採用研修（トレーニー）", "カンポリ参加", 1226, 2.15, 3, [None] * 6),
    ("新規採用研修（トレーナー）", "週2〜3名", 1251, 5.5, 1, [None] * 6),
    ("新規採用研修（トレーナー）", "通常", 1251, 5, 4, [None] * 6),
    ("既存スタッフ研修", "週1", 1251, 1, 1, [1, None, None, None, None, None]),
    ("既存スタッフ研修", "週2", 1251, 1.5, 3, [None, 1, None, None, None, None]),
    ("既存スタッフ研修", "週3", 1251, 1.75, 6, [None, None, 1, None, None, None]),
    ("既存スタッフ研修", "週4", 1251, 1.5, 6, [None, None, None, 1, None, None]),
    ("防災訓練", "週1", 1251, 0.5, 10, [1, None, None, None, None, None]),
    ("防災訓練", "週2", 1251, 0.5, 15, [None, 1, None, None, None, None]),
    ("防災訓練", "週3", 1251, 0.75, 8, [None, None, 1, None, None, None]),
    ("防災訓練", "週4", 1251, 0.75, 2, [None, None, None, 1, None, None]),
    ("防災訓練", "総合訓練", 1251, 1, 8, [None] * 6),
    ("棚卸（予算分）", "棚卸", 1251, 5, 5, [None] * 6),
    ("ストア準備・返品（予算分）", "定例（単価以外は変更不可）", 1251, 2, 1, [1, 1, 1, 1, 1, None]),
    ("ストア準備・返品（追加分）", "準備", 1251, 2, 2, [None] * 6),
    ("ストア準備・返品（追加分）", "返品", 1251, 6, 2, [None] * 6),
    ("ストア準備・返品（追加分）", "陳列", 1251, 6, 4, [None] * 6),
    ("その他", "", 1251, None, None, [None] * 6),
    ("リーダー手当", "手当（金額を単価欄に入れ、時間・人数・回数を1にする）", 93000, 1, 1, [None] * 6),
]

FIRST, LAST = 5, 104
for r in range(FIRST, LAST + 1):
    idx = r - FIRST
    item = DETAIL[idx] if idx < len(DETAIL) else None
    put(p, f"A{r}", item[0] if item else None, color=BLUE, fill=YELLOW, border=True)
    put(p, f"B{r}", item[1] if item else None, color=BLUE, fill=YELLOW, border=True)
    put(p, f"C{r}", item[2] if item else None, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
    put(p, f"D{r}", item[3] if item else None, color=BLUE, fill=YELLOW, fmt=NUM, border=True)
    put(p, f"E{r}", item[4] if item else None, color=BLUE, fill=YELLOW, fmt=NUM, border=True)
    for i in range(6):
        wk = get_column_letter(6 + i)
        amt = get_column_letter(14 + i)
        put(p, f"{wk}{r}", item[5][i] if item else None,
            color=BLUE, fill=YELLOW, fmt=NUM, align="center", border=True)
        put(p, f"{amt}{r}", f'=IF(OR($A{r}="",{amt}$4=""),"",$C{r}*$D{r}*$E{r}*{wk}{r})',
            fmt=YEN, border=True)
    put(p, f"L{r}", f'=IF($A{r}="","",SUMPRODUCT(($F$4:$K$4<>"")*$F{r}:$K{r}))',
        fmt=NUM, align="center", border=True)
    put(p, f"M{r}", f'=IF($A{r}="","",SUM($N{r}:$S{r}))', fmt=YEN, border=True)

put(p, f"A{LAST + 2}", "合計", bold=True, fill=TOTALFILL, border=True)
for col in "BCDE":
    put(p, f"{col}{LAST + 2}", None, fill=TOTALFILL, border=True)
for i in range(6):
    wk = get_column_letter(6 + i)
    amt = get_column_letter(14 + i)
    put(p, f"{wk}{LAST + 2}", f"=SUM({wk}{FIRST}:{wk}{LAST})", bold=True, fill=TOTALFILL,
        fmt=NUM, align="center", border=True)
    put(p, f"{amt}{LAST + 2}", f"=SUM({amt}{FIRST}:{amt}{LAST})", bold=True, fill=TOTALFILL,
        fmt=YEN, border=True)
put(p, f"L{LAST + 2}", f"=SUM(L{FIRST}:L{LAST})", bold=True, fill=TOTALFILL, fmt=NUM,
    align="center", border=True)
put(p, f"M{LAST + 2}", f"=SUM(M{FIRST}:M{LAST})", bold=True, fill=TOTALFILL, fmt=YEN, border=True)
put(p, f"A{LAST + 4}",
    "※ 項目名は「予算割振」シートの項目名と同じ文字にしてください（違うと集計に載りません）。",
    size=9, color="808080")

widths(p, {"A": 26, "B": 30, "C": 11, "D": 8, "E": 8,
           **{get_column_letter(6 + i): 7 for i in range(6)},
           "L": 8, "M": 12, "N": 12,
           **{get_column_letter(15 + i): 12 for i in range(5)}})
p.freeze_panes = "F5"

# =====================================================================
# 5. 人件費マスター（動員帯別）
# =====================================================================
mst = wb.create_sheet("人件費マスター")
title(mst, "人件費マスター（動員帯別）",
      "日別の予測動員がどの帯に入るかで、部門別の人件費を決めます。"
      "現行の接客部門割振表と同じ数値を初期値にしています。")

put(mst, "A3", "時給", bold=True)
put(mst, "B3", 1251, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
put(mst, "D3", "■ 劇場マスター額（早朝・ナイト）", bold=True, fill=GRAY)
NIGHT = [("早朝", 28512), ("ナイトボックス", 8168), ("ナイトコンセ", 59400),
         ("ナイトフロア", 53460), ("ナイトストア", 9653)]
for i, (label, val) in enumerate(NIGHT):
    put(mst, f"D{4 + i}", label)
    put(mst, f"E{4 + i}", val, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
put(mst, "G4", "← 接客部門シートの早朝・ナイト欄に入れる金額です", size=9, color="808080")

MST_HEAD = 11
MST_FIRST = 12
SECTIONS = ["ボックス", "コンセ", "フロア", "ストア", "オフィス"]

mst.merge_cells(f"D{MST_HEAD - 1}:H{MST_HEAD - 1}")
put(mst, f"D{MST_HEAD - 1}", "構成比", bold=True, fill=GRAY, align="center", border=True)
mst.merge_cells(f"I{MST_HEAD - 1}:M{MST_HEAD - 1}")
put(mst, f"I{MST_HEAD - 1}", "金額（標準総額 × 構成比）", bold=True, fill=HEAD,
    align="center", border=True)
for col, label in zip("ABC", ["動員 下限", "動員 上限", "標準総額"]):
    put(mst, f"{col}{MST_HEAD}", label, bold=True, fill=HEAD, align="center", border=True)
for i, sec in enumerate(SECTIONS):
    put(mst, f"{get_column_letter(4 + i)}{MST_HEAD}", sec, bold=True, fill=GRAY,
        align="center", border=True)
    put(mst, f"{get_column_letter(9 + i)}{MST_HEAD}", sec, bold=True, fill=HEAD,
        align="center", border=True)
put(mst, f"N{MST_HEAD}", "部門計", bold=True, fill=HEAD, align="center", border=True)

BANDS = json.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "master.json")))
for i, band in enumerate(BANDS):
    r = MST_FIRST + i
    lo, hi, total = band[0], band[1], band[2]
    put(mst, f"A{r}", lo, color=BLUE, fill=YELLOW, fmt=YEN, align="center", border=True)
    put(mst, f"B{r}", hi, color=BLUE, fill=YELLOW, fmt=YEN, align="center", border=True)
    put(mst, f"C{r}", total, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
    for k in range(5):
        put(mst, f"{get_column_letter(4 + k)}{r}", band[3 + k], color=BLUE, fill=YELLOW,
            fmt=PCT2, border=True)
        put(mst, f"{get_column_letter(9 + k)}{r}",
            f"=$C{r}*{get_column_letter(4 + k)}{r}", fmt=YEN, border=True)
    put(mst, f"N{r}", f"=SUM(I{r}:M{r})", bold=True, fmt=YEN, border=True)
MST_LAST = MST_FIRST + len(BANDS) - 1

put(mst, f"A{MST_LAST + 2}",
    "※ 動員が帯の下限以上・次の帯の下限未満なら、その帯の金額を使います（VLOOKUPの近似一致）。",
    size=9, color="808080")
put(mst, f"A{MST_LAST + 3}",
    "※ 帯は動員の小さい順に並べてください。行の追加・単価や構成比の変更はそのまま反映されます。",
    size=9, color="808080")
widths(mst, {"A": 11, "B": 11, "C": 13, **{get_column_letter(4 + i): 11 for i in range(5)},
             **{get_column_letter(9 + i): 13 for i in range(5)}, "N": 13})
mst.freeze_panes = f"A{MST_FIRST}"

# =====================================================================
# 6. 接客部門（日別の見込人件費）
# =====================================================================
h = wb.create_sheet("接客部門")
title(h, "接客部門 人件費（動員帯マスターから自動計算）",
      "日別の予測動員（動員予測シート）から、部門別の人件費を自動で引きます。"
      "早朝・ナイトは金額を直接入力します。")

put(h, "A3", "予算動員", bold=True)
put(h, "B3", None, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
put(h, "C3", "予測動員", bold=True, align="center")
put(h, "D3", "=$D$40", fmt=YEN, border=True)
put(h, "E3", "動員 予算比", bold=True, align="center")
put(h, "F3", '=IF($B$3=0,"",$D$3/$B$3)', fmt=PCT2, border=True)
put(h, "H3", "早朝・ナイトの金額 →", size=9, color="808080")
for i, (label, _) in enumerate(NIGHT):
    put(h, f"{get_column_letter(10 + i)}4", f"=人件費マスター!$E${4 + i}",
        fmt=YEN, align="center", size=9, color="808080")

DAY_HEAD = 8
DAY_FIRST = 9
DAY_LAST = DAY_FIRST + 30   # 31日分
DAY_TOT = DAY_LAST + 1

heads = ["日付", "曜日", "週", "予測動員"] + SECTIONS + \
        ["早朝", "ナイトボックス", "ナイトコンセ", "ナイトフロア", "ナイトストア", "日計"]
for i, label in enumerate(heads):
    col = get_column_letter(1 + i)
    fill = HEAD if i < 9 else (YELLOW if i < 14 else TOTALFILL)
    put(h, f"{col}{DAY_HEAD}", label, bold=True, fill=fill, align="center",
        border=True, wrap=True)

for i in range(31):
    r = DAY_FIRST + i
    src = 4 + i  # 動員予測シートの行
    put(h, f"A{r}", f'=IF({i + 1}<=DAY(設定!$B$12),設定!$B$11+{i},"")',
        fmt="m/d", align="center", border=True)
    put(h, f"B{r}", f'=IF($A{r}="","",CHOOSE(WEEKDAY($A{r},2),"月","火","水","木","金","土","日"))',
        align="center", border=True)
    put(h, f"C{r}", f'=IF($A{r}="","",IF($A{r}<=設定!$B$13,1,1+CEILING(($A{r}-設定!$B$13)/7,1)))',
        align="center", border=True)
    put(h, f"D{r}", f'=IF($A{r}="","",動員予測!$D${src})', fmt=YEN, border=True)
    for k in range(5):
        col = get_column_letter(5 + k)
        put(h, f"{col}{r}",
            f'=IF(OR($A{r}="",$D{r}=0),"",'
            f'VLOOKUP($D{r},人件費マスター!$A${MST_FIRST}:$N${MST_LAST},{9 + k},TRUE))',
            fmt=YEN, border=True)
    for k in range(5):
        col = get_column_letter(10 + k)
        put(h, f"{col}{r}", None, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
    put(h, f"O{r}", f'=IF($A{r}="","",SUM($E{r}:$N{r}))', bold=True, fmt=YEN, border=True)

put(h, f"A{DAY_TOT}", "合計", bold=True, fill=TOTALFILL, border=True)
for col in "BC":
    put(h, f"{col}{DAY_TOT}", None, fill=TOTALFILL, border=True)
for i in range(4, 16):
    col = get_column_letter(i)
    put(h, f"{col}{DAY_TOT}", f"=SUM({col}{DAY_FIRST}:{col}{DAY_LAST})",
        bold=True, fill=TOTALFILL, fmt=YEN, border=True)

# ---- 週別集計 ----
WK_HEAD = DAY_TOT + 3
put(h, f"A{WK_HEAD - 1}", "■ 週別集計（金曜〜木曜。早朝はフロアに含めます）", bold=True, fill=GRAY)
put(h, f"A{WK_HEAD}", "項目", bold=True, fill=HEAD, align="center", border=True)
for i in range(6):
    col = get_column_letter(2 + i)
    put(h, f"{col}{WK_HEAD}", f"=設定!$B${21 + i}", bold=True, fill=HEAD,
        align="center", border=True)
put(h, f"H{WK_HEAD}", "月計", bold=True, fill=HEAD, align="center", border=True)

WK_ATT = WK_HEAD + 1
put(h, f"A{WK_ATT}", "予測動員", bold=True, border=True)
for i in range(6):
    col = get_column_letter(2 + i)
    put(h, f"{col}{WK_ATT}",
        f'=IF({col}${WK_HEAD}="","",SUMIF($C${DAY_FIRST}:$C${DAY_LAST},{i + 1},$D${DAY_FIRST}:$D${DAY_LAST}))',
        fmt=YEN, border=True)
put(h, f"H{WK_ATT}", f"=SUM(B{WK_ATT}:G{WK_ATT})", bold=True, fmt=YEN, border=True)

# 部門別（早朝はフロア、ナイトは各部門に加算）
EXTRA = {"ボックス": ["K"], "コンセ": ["L"], "フロア": ["M", "J"], "ストア": ["N"], "オフィス": []}
WK_SEC = WK_ATT + 1
for k, sec in enumerate(SECTIONS):
    r = WK_SEC + k
    base = get_column_letter(5 + k)
    put(h, f"A{r}", sec, border=True)
    for i in range(6):
        col = get_column_letter(2 + i)
        terms = [f'SUMIF($C${DAY_FIRST}:$C${DAY_LAST},{i + 1},${c}${DAY_FIRST}:${c}${DAY_LAST})'
                 for c in [base] + EXTRA[sec]]
        put(h, f"{col}{r}", f'=IF({col}${WK_HEAD}="","",{"+".join(terms)})', fmt=YEN, border=True)
    put(h, f"H{r}", f"=SUM(B{r}:G{r})", fmt=YEN, border=True)
WK_SUM = WK_SEC + 5
put(h, f"A{WK_SUM}", "週計", bold=True, fill=TOTALFILL, border=True)
for i in range(6):
    col = get_column_letter(2 + i)
    put(h, f"{col}{WK_SUM}", f'=IF({col}${WK_HEAD}="","",SUM({col}{WK_SEC}:{col}{WK_SUM - 1}))',
        bold=True, fill=TOTALFILL, fmt=YEN, border=True)
put(h, f"H{WK_SUM}", f"=SUM(B{WK_SUM}:G{WK_SUM})", bold=True, fill=TOTALFILL, fmt=YEN, border=True)

# ---- 追加分 ----
ADD_HEAD = WK_SUM + 2
put(h, f"A{ADD_HEAD}", "■ 追加分（週・部門ごとに入力）", bold=True, fill=GRAY)
ADD_FIRST = ADD_HEAD + 1
for k, sec in enumerate(SECTIONS):
    r = ADD_FIRST + k
    put(h, f"A{r}", sec, border=True)
    for i in range(6):
        col = get_column_letter(2 + i)
        put(h, f"{col}{r}", None, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
    put(h, f"H{r}", f"=SUM(B{r}:G{r})", fmt=YEN, border=True)

# ---- 合計 ----
TOT_HEAD = ADD_FIRST + 6
put(h, f"A{TOT_HEAD}", "■ 合計（週別集計＋追加分）", bold=True, fill=GRAY)
TOT_FIRST = TOT_HEAD + 1
for k, sec in enumerate(SECTIONS):
    r = TOT_FIRST + k
    put(h, f"A{r}", sec, bold=True, border=True)
    for i in range(6):
        col = get_column_letter(2 + i)
        put(h, f"{col}{r}", f'=IF({col}${WK_HEAD}="","",{col}{WK_SEC + k}+{col}{ADD_FIRST + k})',
            fmt=YEN, border=True)
    put(h, f"H{r}", f"=SUM(B{r}:G{r})", fmt=YEN, border=True)
TOT_SUM = TOT_FIRST + 5
put(h, f"A{TOT_SUM}", "週合計", bold=True, fill=TOTALFILL, border=True)
for i in range(6):
    col = get_column_letter(2 + i)
    put(h, f"{col}{TOT_SUM}", f'=IF({col}${WK_HEAD}="","",SUM({col}{TOT_FIRST}:{col}{TOT_SUM - 1}))',
        bold=True, fill=TOTALFILL, fmt=YEN, border=True)
put(h, f"H{TOT_SUM}", f"=SUM(B{TOT_SUM}:G{TOT_SUM})", bold=True, fill=TOTALFILL,
    fmt=YEN, border=True)

# ---- 予算との比較 ----
CMP = TOT_SUM + 2
put(h, f"A{CMP}", "■ 予算との比較", bold=True, fill=GRAY)
put(h, f"A{CMP + 1}", "予算配分（予算割振シート）", bold=True, border=True)
for i in range(6):
    col = get_column_letter(2 + i)
    src = get_column_letter(4 + i)
    put(h, f"{col}{CMP + 1}", f'=IF({col}${WK_HEAD}="","",予算割振!{src}${ITEM_FIRST})',
        fmt=YEN, border=True)
put(h, f"H{CMP + 1}", f"=SUM(B{CMP + 1}:G{CMP + 1})", bold=True, fmt=YEN, border=True)
put(h, f"A{CMP + 2}", "差額（予算配分 − 合計）", bold=True, border=True)
for i in range(6):
    col = get_column_letter(2 + i)
    put(h, f"{col}{CMP + 2}", f'=IF({col}${WK_HEAD}="","",{col}{CMP + 1}-{col}{TOT_SUM})',
        fmt=YEN, border=True)
put(h, f"H{CMP + 2}", f"=SUM(B{CMP + 2}:G{CMP + 2})", bold=True, fmt=YEN, border=True)
put(h, f"A{CMP + 3}", "予算比（合計 ÷ 予算配分）", bold=True, border=True)
for i in range(6):
    col = get_column_letter(2 + i)
    put(h, f"{col}{CMP + 3}", f'=IF(OR({col}${WK_HEAD}="",{col}{CMP + 1}=0),"",{col}{TOT_SUM}/{col}{CMP + 1})',
        fmt=PCT2, border=True)
put(h, f"H{CMP + 3}", f'=IF($H{CMP + 1}=0,"",$H{TOT_SUM}/$H{CMP + 1})', bold=True,
    fmt=PCT2, border=True)

put(h, f"A{CMP + 5}",
    "※ 部門別の金額は「人件費マスター」シートの動員帯テーブルから、日別の予測動員に応じて引いています。",
    size=9, color="808080")
put(h, f"A{CMP + 6}",
    "※ 週別集計では、早朝はフロアに、ナイト分は各部門に足しています（現行の割振表と同じ扱い）。",
    size=9, color="808080")

widths(h, {"A": 22, "B": 8, "C": 6, "D": 12,
           **{get_column_letter(5 + i): 12 for i in range(5)},
           **{get_column_letter(10 + i): 12 for i in range(5)}, "O": 13})
h.freeze_panes = f"E{DAY_FIRST}"

# =====================================================================
# 6. 有休
# =====================================================================
lv = wb.create_sheet("有休")
title(lv, "有休計算表",
      "週ごとの人数を入れると金額が出ます。予算との差額は各週へ均等に振り分けます。")
put(lv, "A3", "有休単価", bold=True)
put(lv, "B3", 979, color=BLUE, fill=YELLOW, fmt=YEN, border=True)
put(lv, "A4", "時間", bold=True)
put(lv, "B4", 5, color=BLUE, fill=YELLOW, fmt=NUM, border=True)
put(lv, "A5", "予算", bold=True)
put(lv, "B5", 325789, color=BLUE, fill=YELLOW, fmt=YEN, border=True)

for col, label in zip("ABCDEF", ["週", "WEEK", "人数", "金額", "補正後", "1日あたり"]):
    put(lv, f"{col}7", label, bold=True, fill=HEAD, align="center", border=True)
for i in range(6):
    r = 8 + i
    put(lv, f"A{r}", i + 1, align="center", border=True)
    put(lv, f"B{r}", f"=設定!$B${21 + i}", align="center", border=True)
    put(lv, f"C{r}", None, color=BLUE, fill=YELLOW, fmt=NUM, border=True)
    put(lv, f"D{r}", f'=IF($B{r}="","",$B$3*$B$4*$C{r})', fmt=YEN, border=True)
    put(lv, f"E{r}", f'=IF($B{r}="","",$D{r}+$B$16)', fmt=YEN, border=True)
    put(lv, f"F{r}", f'=IF(OR($B{r}="",$C{r}=0),"",$E{r}/$C{r})', fmt=YEN, border=True)
put(lv, "A14", "合計", bold=True, fill=TOTALFILL, border=True)
put(lv, "B14", None, fill=TOTALFILL, border=True)
put(lv, "C14", "=SUM(C8:C13)", bold=True, fill=TOTALFILL, fmt=NUM, border=True)
put(lv, "D14", "=SUM(D8:D13)", bold=True, fill=TOTALFILL, fmt=YEN, border=True)
put(lv, "E14", "=SUM(E8:E13)", bold=True, fill=TOTALFILL, fmt=YEN, border=True)
put(lv, "F14", None, fill=TOTALFILL, border=True)
put(lv, "A15", "予算 − 単純計算額")
put(lv, "B15", "=$B$5-$D$14", fmt=YEN, border=True)
put(lv, "A16", "1週あたりの調整額")
put(lv, "B16", "=IF(設定!$B$17=0,0,$B$15/設定!$B$17)", fmt=YEN, border=True)
put(lv, "C16", "← 各週へ均等に振り分け、補正後の合計が予算と一致します", size=9, color="808080")
widths(lv, {"A": 20, "B": 14, "C": 10, "D": 14, "E": 14, "F": 14})

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

# シートの並び順
ORDER = ["設定", "動員予測", "予算割振", "接客部門", "明細", "有休", "人件費マスター"]
wb._sheets = [wb[n] for n in ORDER]

wb.calculation.fullCalcOnLoad = True  # 開いたときに必ず再計算させる

out = "/home/user/test/excel/月次予算割振.xlsx"
wb.save(out)
print("saved", out)
