"""入力が欠けた・崩れたときの挙動を実機（LibreOffice再計算）で確かめる。

    python3 excel/scenarios.py            # 全シナリオ
    python3 excel/scenarios.py numeric    # 名前の前方一致で絞る

サンプル入りブックを複製して壊し、再計算後に「エラー数」と「黙って発注が消えていないか」を見る。
"""
import shutil, subprocess, sys, warnings
from pathlib import Path
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

HERE = Path(__file__).resolve().parent
SAMPLE = HERE / '売店発注ツール_サンプルデータ入り.xlsx'
WORK = Path('/tmp/claude-0/-home-user-test/6ebcf1c1-1102-5af1-96f9-c94aebb0a201/scratchpad/scenarios')
RECALC = '/root/.claude/skills/xlsx/scripts/recalc.py'
ERRS = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
S_CUR = '①今日の在庫を貼る'; S_MID = '月1回_1ヶ月前の在庫を貼る'; S_PRV = '月1回_2ヶ月前の在庫を貼る'
S_SET = '設定（最初に1回）'; S_TT = '②発注数を決める'; S_CALC = '発注計算'; S_CONST = '定数マスタ'


def numeric_code(wb):
    """Excelに貼ると先頭ゼロが落ちて数値になる、をわざと起こす"""
    for n in (S_CUR, S_MID, S_PRV):
        ws = wb[n]
        for r in range(6, ws.max_row + 1):
            for c in (2, 9):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip().isdigit():
                    ws.cell(r, c).value = int(v)
                    ws.cell(r, c).number_format = 'General'


def cur_only(wb):
    """当日CSVだけ貼って、1ヶ月前・2ヶ月前が空"""
    for n in (S_MID, S_PRV):
        ws = wb[n]
        for r in range(6, ws.max_row + 1):
            for c in range(1, 20):
                ws.cell(r, c).value = None


def over400(wb):
    """取扱商品が計算行の上限(400)を超える"""
    ws = wb[S_CUR]
    rows = [[ws.cell(r, c).value for c in range(1, 20)]
            for r in range(6, ws.max_row + 1) if ws.cell(r, 2).value == '0761']
    last = max(r for r in range(6, ws.max_row + 1) if ws.cell(r, 2).value)
    k = 0
    while len(rows) + k < 420:
        base = list(rows[k % len(rows)]); k += 1
        base[8] = f'9{k:012d}'; base[9] = f'合成商品{k}'
        for c, v in enumerate(base, start=1):
            ws.cell(last + k, c).value = v


SCENARIOS = {'numeric_code': numeric_code, 'cur_only': cur_only, 'over400': over400}


def report(path):
    wb = load_workbook(path, data_only=True)
    errs = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and any(e in v for e in ERRS):
                    errs.append(f'{ws.title}!{c.coordinate}={v[:14]}')
    s, tt, c, th = wb[S_SET], wb[S_TT], wb[S_CALC], wb[S_CONST]
    items = sum(1 for r in range(6, 406) if c.cell(r, 3).value)
    rec = sum(1 for r in range(6, 406) if c.cell(r, 3).value
              and isinstance(c.cell(r, 25).value, (int, float)) and c.cell(r, 25).value > 0)
    par = sum(1 for r in range(6, 406) if c.cell(r, 3).value and c.cell(r, 46).value not in (None, ''))
    match = [th.cell(r, 13).value for r in range(6, 106) if th.cell(r, 2).value == '0761']
    return dict(errors=len(errs), err_cells=errs[:5], items=items, rec_gt0=rec, with_par=par,
                banner=str(wb[S_CUR]['H2'].value)[:22], h3=str(wb[S_CUR]['H3'].value)[:60],
                c17=s['C17'].value, c19=s['C19'].value, b7=str(tt['B7'].value)[:50],
                csv_match=match[0] if match else None)


def main():
    want = sys.argv[1] if len(sys.argv) > 1 else ''
    WORK.mkdir(parents=True, exist_ok=True)
    for name, mutate in SCENARIOS.items():
        if not name.startswith(want):
            continue
        dst = WORK / f'{name}.xlsx'
        shutil.copy(SAMPLE, dst)
        wb = load_workbook(dst); mutate(wb); wb.save(dst)
        subprocess.run([sys.executable, RECALC, str(dst)], capture_output=True)
        r = report(dst)
        print(f'== {name}: errors={r["errors"]} 商品={r["items"]} 定数あり={r["with_par"]} 推奨>0={r["rec_gt0"]}')
        print(f'   バナー: {r["banner"]} | H3: {r["h3"]}')
        print(f'   C17={r["c17"]} C19={r["c19"]} 照合(0761)={r["csv_match"]}')
        print(f'   ②B7: {r["b7"]}')
        if r['err_cells']:
            print('   ERR:', r['err_cells'])


if __name__ == '__main__':
    main()
