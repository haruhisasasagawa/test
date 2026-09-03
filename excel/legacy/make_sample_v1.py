"""サンプルデータ入りのブックを作る。3劇場・動員計画・発注履歴まで入れて、
開いた時点でひととおりの図が埋まっている状態にする。

    python3 excel/make_sample.py            # サンプル入りブックを作る
    python3 excel/make_sample.py <シート名>  # さらに、そのシートだけ印刷範囲を絞った描画用コピーを作る

在庫CSV(stock.csv)は実データのためリポジトリには入れず、scratchpadから読む。
"""
import csv
import datetime
import random
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl

random.seed(20260501)
HERE = Path(__file__).resolve().parent
S = '/tmp/claude-0/-home-user-test/6ebcf1c1-1102-5af1-96f9-c94aebb0a201/scratchpad'
SRC = str(HERE / '売店発注ツール.xlsx')
OUT = str(HERE / '売店発注ツール_サンプルデータ入り.xlsx')
RECALC = '/root/.claude/skills/xlsx/scripts/recalc.py'

H = ['対象日付', '劇場コード', '劇場名', '支払先コード', '支払先名', '大分類コード',
     '小分類コード', '商品分類名', '商品コード', '商品名', '入数', '在庫数ケース',
     '在庫数バラ', '総数', '規定数', '資産廃棄', '税抜単価', '仕入税区分', '仕入税率']

rows = list(csv.DictReader(open(S + '/stock.csv', encoding='cp932')))
BASE = datetime.date(2026, 5, 1)          # 当日基準日
MID = datetime.date(2026, 4, 1)           # 1ヶ月前の1日
PREV = datetime.date(2026, 3, 1)          # 2ヶ月前の1日

THEATERS = [
    # (コード, 名前, 規模, 提供方式, 取扱商品の割合, 在庫の潤沢さ)
    # 劇場マスタに登録済みの実在劇場から3館を選んでサンプルにしている
    ('0761', '新宿', '大規模', '1杯売り',      1.00, 1.0),
    ('0841', '池袋', '中規模', 'ドリンクバー', 0.72, 1.4),
    ('0681', '上田', '小規模', '1杯売り',      0.55, 2.2),
]


def put(ws, i, r):
    for j, h in enumerate(H):
        v = r[h]
        if h in ('入数', '在庫数ケース', '在庫数バラ', '総数', '資産廃棄', '規定数'):
            v = int(v) if str(v).strip() else None
        elif h in ('税抜単価', '仕入税率'):
            v = float(v) if str(v).strip() else None
        elif h == '対象日付':
            v = int(v)
        ws.cell(i, 1 + j, v)


def row_for(date, code, name, src, total):
    p = dict(src)
    ins = max(1, int(src['入数']))
    p['対象日付'] = date.strftime('%Y%m%d')
    p['劇場コード'] = code
    p['劇場名'] = name
    p['在庫数ケース'] = total // ins
    p['在庫数バラ'] = total - (total // ins) * ins
    p['総数'] = total
    return p


def snapshots(code, name, items, scale):
    """当日を先に決め、さかのぼって1ヶ月前・2ヶ月前を作る。
    こうすると期間消費が必ず正になり、実績消費が算出できる状態になる。
    取扱の入れ替わりも、3枚のどこに載せるかで作り分ける。"""
    cur, mid, prv = [], [], []
    for k, src in enumerate(items):
        base = max(1, int(src['総数']))
        today = int(base * random.choice([0.05, 0.2, 0.4, 0.7, 1.0, 1.5, 2.2]) * scale)
        if random.random() < 0.05:          # 理論値がマイナスの商品もたまにある
            today = -random.randint(1, 60)
        used1 = int(base * random.uniform(0.4, 1.3))      # 直近1ヶ月の消費
        used2 = int(base * random.uniform(0.4, 1.3))      # その前1ヶ月の消費
        m = max(0, today) + used1
        cur.append(row_for(BASE, code, name, src, today))
        mid.append(row_for(MID, code, name, src, m))
        prv.append(row_for(PREV, code, name, src, m + used2))
    return cur, mid, prv


def variant(items, idx, suffix, label, qty):
    src = dict(items[idx])
    src.update({'商品コード': suffix, '商品名': label, '入数': '1'})
    return src, qty


wb = openpyxl.load_workbook(SRC)
ws_cur = wb['①今日の在庫を貼る']
ws_mid = wb['月1回_1ヶ月前の在庫を貼る']
ws_prv = wb['月1回_2ヶ月前の在庫を貼る']

cur_rows, mid_rows, prv_rows = [], [], []
for code, name, size, style, share, scale in THEATERS:
    items = rows[:int(len(rows) * share)]
    c, m, p = snapshots(code, name, items, scale)
    # 取扱の入れ替わりを、3枚のどこに載せるかで作り分ける
    churn = [
        # (ラベル, 当日, 1ヶ月前, 2ヶ月前)
        ('（新商品）夏季限定フローズン',   True,  False, False),   # 今月から
        ('（新商品）コラボカップ',        True,  True,  False),   # 先月から
        ('（終売）季節限定ドリンク',       False, True,  True),    # 今月から終売
        ('（終売）旧パッケージカップ',      False, False, True),    # 先月から終売
        ('（復活）冬季限定ホットスナック',   True,  False, True),    # 1ヶ月だけ消えていた
    ]
    for n, (label, a, b, d) in enumerate(churn):
        src, qty = variant(items, 3 + n, f'9{n}{code}000001', label, 40 + n * 25)
        if a:
            c.append(row_for(BASE, code, name, src, qty))
        if b:
            m.append(row_for(MID, code, name, src, qty + 30))
        if d:
            p.append(row_for(PREV, code, name, src, qty + 60))
    cur_rows += c
    mid_rows += m
    prv_rows += p

for ws_, rows_ in ((ws_cur, cur_rows), (ws_mid, mid_rows), (ws_prv, prv_rows)):
    for i, r in enumerate(rows_):
        put(ws_, 6 + i, r)

# ---- 劇場マスタ：登録済みの行を探して、サンプル3館ぶんだけ手を入れる ----
th = wb['定数マスタ']
ROW = {}
for r in range(6, 106):
    code = th.cell(r, 2).value
    if code:
        ROW[str(code)] = r
for code, name, size, style, _s, _c in THEATERS:
    r = ROW[code]
    th.cell(r, 4, size)
    th.cell(r, 5, style)
th.cell(ROW['0841'], 6, 1800)     # 個別の標準来場者数で上書きする例（F列）
th.cell(ROW['0841'], 12, '中規模だが動員が多いため個別に上書き。ドリンクバー方式')

# ---- 商品マスタ：PI値・リードタイム・最大保管数を実務らしく埋める ----
im = wb['商品マスタ']
for i in range(len(rows)):
    r = 6 + i
    name = str(im.cell(r, 3).value or '')
    cat = str(im.cell(r, 4).value or '')
    base_pi = {
        'コンセ包材': 9.0, 'ポップコーン': 6.5, 'コールド': 5.5, 'ホット': 1.2,
        'その他ドリンク': 2.4, 'アルコール': 0.8, 'コーヒー': 1.1, 'ホットドッグ': 1.4,
        '軽食系フード': 1.0, '調理系スイーツ': 1.6, 'フード調味料': 2.0,
        'ＳＥＴ作品コンボ': 0.9, 'その他フード': 0.7,
    }.get(cat, 1.0)
    jitter = 0.6 + random.random() * 0.9
    pi_cup = round(base_pi * jitter, 1)
    # ドリンクバー店はカップ構成が変わる。Sカップ・Lカップは使わずMに寄る
    if 'Ｓ' in name or 'Ｌ' in name:
        pi_bar = 0
    elif 'Ｍ' in name or 'カップ' in name:
        pi_bar = round(pi_cup * 1.8, 1)
    else:
        pi_bar = round(pi_cup * 0.95, 1)
    # 保管区分は分類名から機械的に割り当てる（実運用では現場が選ぶ）
    kind = ('冷凍' if cat in ('調理系スイーツ', '軽食系フード', 'ホットドッグ')
            else '冷蔵' if cat in ('コールド', 'その他ドリンク', 'アルコール') else '常温')
    # 発注グループ＝便のまとまり。冷凍は週3、シロップは週3、包材は木曜のみ
    if kind == '冷凍':
        group = '冷凍品'
    elif cat in ('コールド', 'その他ドリンク'):
        group = 'ドリンクシロップ'
    elif cat == 'コンセ包材':
        group = '常温包材'
    else:
        group = 'その他常温'
    im.cell(r, 10, kind)     # J 保管区分
    im.cell(r, 11, group)    # K 発注グループ
    im.cell(r, 12, pi_cup)   # L PI値 1杯売り
    im.cell(r, 13, pi_bar)   # M PI値 ドリンクバー
    im.cell(r, 8, random.choice([2, 3, 3, 4, 5, 7]))    # H L/T日数
    im.cell(r, 9, random.choice([1, 1, 1, 2, 5]))       # I 最低ロット
    # N 最大保管数（ケース）。8割の商品だけ設定済みにして、未設定の見え方も確認する。
    # わざと小さすぎる定数も混ぜて「定数が小さい」の検出を見る
    if random.random() < 0.8:
        im.cell(r, 14, random.choice([2, 3, 4, 6, 8, 10, 12, 1]))

# ---- 季節係数マスタ：繁忙期・閑散期 ----
se = wb['定数マスタ']
for m, rate in ((1, 1.15), (2, 0.9), (3, 1.05), (4, 0.95), (5, 1.2), (6, 0.85),
                (7, 1.25), (8, 1.35), (9, 0.9), (10, 1.0), (11, 1.0), (12, 1.3)):
    se.cell(5 + m, 21, rate)          # U列
specials = [('ゴールデンウィーク', datetime.date(2026, 4, 29), datetime.date(2026, 5, 6), 1.45),
            ('大型作品公開週', datetime.date(2026, 5, 15), datetime.date(2026, 5, 21), 1.6),
            ('閑散期', datetime.date(2026, 6, 8), datetime.date(2026, 6, 19), 0.8)]
for i, (nm, s0, e0, rate) in enumerate(specials):
    se.cell(6 + i, 23, nm)            # W〜Z列
    se.cell(6 + i, 24, s0)
    se.cell(6 + i, 25, e0)
    se.cell(6 + i, 26, rate)

# ---- 保管設備：劇場ごとの冷蔵庫・ストッカー ----
# 実在庫（新宿：常温1,776／冷蔵577／冷凍258ケース）に見合う設備量にしてある
EQUIP = {
    '0761': [('常温', 'バックヤード棚', 8, 300), ('冷蔵', '業務用冷蔵庫', 4, 140),
             ('冷蔵', 'ドリンククーラー', 2, 80), ('冷凍', '冷凍ストッカー', 4, 80)],
    '0841': [('常温', 'バックヤード棚', 5, 260), ('冷蔵', '業務用冷蔵庫', 3, 140),
             ('冷凍', '冷凍ストッカー', 2, 80)],
    '0681': [('常温', 'バックヤード棚', 3, 220), ('冷蔵', '業務用冷蔵庫', 2, 120),
             ('冷凍', '冷凍ストッカー', 1, 80)],
}
# build_workbook.py の EQUIP_FIRST と揃える。見出し行を数えて位置を取る
EQUIP_ROW = next(r for r in range(100, 140) if th.cell(r, 2).value == '劇場コード' and th.cell(r, 4).value == '設備名') + 1
r = EQUIP_ROW
for code, items_ in EQUIP.items():
    for kind, nm, n, cap in items_:
        th.cell(r, 2, code); th.cell(r, 3, kind); th.cell(r, 4, nm)
        th.cell(r, 5, n); th.cell(r, 6, cap)
        r += 1

# ---- 動員：前回基準日(3/1)から実績、当日基準日(5/1)から先は予測 ----
plan = wb['随時_動員を入れる']
weekday_base = {0: 2600, 1: 1050, 2: 980, 3: 1150, 4: 1250, 5: 1750, 6: 3000}  # 月=0
titles = {
    datetime.date(2026, 5, 1): 'GW初日',
    datetime.date(2026, 5, 2): 'GW',
    datetime.date(2026, 5, 3): 'GW',
    datetime.date(2026, 5, 4): 'GW',
    datetime.date(2026, 5, 5): 'GWこどもの日／ファミリー作品',
    datetime.date(2026, 5, 15): '大型アニメ作品 公開初日',
    datetime.date(2026, 5, 16): '大型アニメ作品 公開2日目',
    datetime.date(2026, 5, 17): '大型アニメ作品 公開3日目',
    datetime.date(2026, 5, 22): '洋画大作 公開',
    datetime.date(2026, 6, 3): '映画の日',
    datetime.date(2026, 3, 20): '春休み作品 公開',
    datetime.date(2026, 4, 3): '話題作 公開',
}
plan.cell(11, 3, PREV)                      # 表の開始日
PLAN_TOP = 16
for i in range(220):
    d = PREV + datetime.timedelta(days=i)
    r = PLAN_TOP + i
    base = weekday_base[d.weekday()]
    if d in titles:
        base = int(base * random.uniform(1.5, 2.3))
    if d < BASE:
        # 過去：実績を入れる。予測も半分の日だけ入っていて、実績とはズレている
        actual = int(base * random.uniform(0.88, 1.16))
        plan.cell(r, 5, actual)             # E列＝実績来場者数
        if i % 2 == 0 or d in titles:
            plan.cell(r, 4, int(actual * random.uniform(0.85, 1.15)))   # D列＝予測
    elif i < 21 + 61 or d in titles:
        # これから：読みが立つ範囲だけ予測を入れる
        plan.cell(r, 4, int(base * random.uniform(0.9, 1.12)))
    if d in titles:
        plan.cell(r, 6, titles[d])          # F列＝主な作品・備考

# ---- 発注管理：入荷済みと未入荷を混ぜる ----
po = wb['発注管理']
picks = random.sample(range(len(rows)), 14)
for i, idx in enumerate(picks):
    r = 6 + i
    received = i < 8
    order_date = BASE - datetime.timedelta(days=random.randint(8, 40) if received else random.randint(1, 4))
    po.cell(r, 3, THEATERS[i % 3][0])
    po.cell(r, 5, order_date)
    po.cell(r, 6, rows[idx]['商品コード'])
    po.cell(r, 11, random.choice([2, 3, 4, 5, 6, 8, 10, 12]))
    po.cell(r, 17, '入荷済み' if received else '発注済')
    if received:
        po.cell(r, 18, order_date + datetime.timedelta(days=3))
    po.cell(r, 20, random.choice(['笹川', '田中', '佐藤']))
    po.cell(r, 21, '定期発注' if i % 3 else '公開週の追加発注')

# ---- 設定・発注数 ----
wb['設定（最初に1回）']['C16'] = '笹川'
# 期間の平均動員は動員シートの実績から自動で出るので、ここでは何も入れない
tt = wb['②発注数を決める']
for r, q in ((10, 6), (11, 3), (12, 10), (14, 4), (17, 8), (20, 2), (23, 5)):
    tt.cell(r, 12, q)   # L列＝発注数

wb.save(OUT)
print(f'sample built: 当日{len(cur_rows)}行 / 1ヶ月前{len(mid_rows)}行 / '
      f'2ヶ月前{len(prv_rows)}行 / {len(THEATERS)}劇場')

# ---- 実棚：理論値が要るので、一度再計算してから数えた数を書き込む ----
# 実運用でも「CSVを貼る → 理論値が出る → 現物を数える」の順なので、同じ順序で作る。
_r = subprocess.run([sys.executable, RECALC, OUT], capture_output=True, text=True)
vals = openpyxl.load_workbook(OUT, data_only=True)
calc = vals['発注計算']
theory = {r: calc.cell(r, 38).value for r in range(6, 406)}      # AL列＝理論値在庫
codes = {r: calc.cell(r, 3).value for r in range(6, 406)}

wb2 = openpyxl.load_workbook(OUT)
cnt = wb2['実棚を入れる']
counted = 0
for i, cr in enumerate(range(6, 406)):
    if not codes.get(cr):
        continue
    t = theory.get(cr)
    if not isinstance(t, (int, float)):
        continue
    row = 10 + (cr - 6)
    if t < 0:                       # マイナスは必ず数える対象
        actual = random.randint(0, 12)
    elif counted < 14 and i % 7 == 3:
        # 数えると理論値と少しずれる。ロス寄りに出るのが実際
        actual = max(0, int(t * random.uniform(0.86, 1.02)))
    else:
        continue
    cnt.cell(row, 6, actual)                    # F列＝数えた数
    cnt.cell(row, 7, BASE)                      # G列＝カウント日
    counted += 1
# 1件だけ、古い日付のまま残っている行を混ぜる（「日付が古い」の見え方の確認用）
cnt.cell(10 + 5, 6, 40)
cnt.cell(10 + 5, 7, BASE - datetime.timedelta(days=9))
wb2.save(OUT)
print(f'実棚サンプル: {counted}件（＋日付が古い行 1件）')

# ---- 描画用コピー：指定シートだけ印刷範囲を絞る ----
if len(sys.argv) > 1:
    shutil.copy(OUT, S + '/v4.xlsx')
    w = openpyxl.load_workbook(S + '/v4.xlsx')
    target = sys.argv[1]
    areas = {'ダッシュボード': 'B2:AC64', '③発注書': 'B2:J40', '発注のしくみ': 'B2:AL86',
             '②発注数を決める': 'B2:AA36', '随時_動員を入れる': 'B2:P46',
             '実棚を入れる': 'B2:R42', 'はじめに': 'B2:D62'}
    for ws in w.worksheets:
        if ws.title == target:
            ws.print_area = areas.get(target, 'B2:Z40')
            ws.page_setup.orientation = 'landscape'
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
        else:
            ws.print_area = 'A1:A1'
    w.active = w.sheetnames.index(target)
    w.save(S + '/v4.xlsx')
    print('render copy ready:', target)
