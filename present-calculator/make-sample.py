#!/usr/bin/env python3
"""
貼付式Excelに、実際のデータを貼り付け済みのサンプルを作る。

    python3 make-sample.py <PS.xlsx> <入場者プレゼント情報.xlsx> <起算日 YYYY-MM-DD> [出力.xlsx]

利用者が手でコピー＆ペーストする操作を Python で再現しているだけで、
テンプレート側の数式は一切変えていない。出来上がったファイルの動きは、
自分で貼り付けた場合とまったく同じになる。

PS のシートは並び順に PS1、PS2、… へ入れ、曜日との対応はシート名から推測する
（シート名に「金」が含まれていれば金曜、など）。推測できない場合は先頭のシートを既定にする。
"""
import sys
import datetime
import openpyxl

PSROWS, GROWS = 250, 150
WD_NAMES = "日月火水木金土"          # WEEKDAY(,1) の 1〜7 に対応

# 作品名から自動生成したキーワードでは当たらないものの手直し（利用者が手で直す作業の再現）
KEYWORD_FIXES = [("ファミリーシアター", "ファミリーシアター")]


def paste(src_ws, dst_ws, max_row, max_col):
    """結合セルは左上だけ値が入る＝Excelで貼り付けた直後と同じ状態にする。"""
    for r in range(1, min(src_ws.max_row, max_row) + 1):
        for c in range(1, max_col + 1):
            v = src_ws.cell(row=r, column=c).value
            if v is not None:
                dst_ws.cell(row=r, column=c, value=v)


def gift_data_rows(ws):
    """「取込」シートと同じ規則で、入場者プレゼント情報のデータ行を拾う。"""
    rows, seen_header = [], False
    for r in range(1, min(ws.max_row, GROWS) + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if a in (None, ""):
            continue
        if str(a).strip() == "作品名" or "納品数" in str(c or ""):
            seen_header = True
            continue
        if b in (None, "") and c in (None, ""):
            continue
        rows.append(str(a))
    return rows


def weekday_map(sheet_names):
    """シート名から曜日→PS番号の対応を作る。名前に曜日が含まれない分は先頭のPSを使う。"""
    mapping = {i: 1 for i in range(1, 8)}       # 1=日 … 7=土
    for ps_no, name in enumerate(sheet_names, start=1):
        for wd_index, wd in enumerate(WD_NAMES, start=1):
            if wd in name:
                mapping[wd_index] = ps_no
    return mapping


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        return 1
    ps_path, gift_path, base_str = sys.argv[1], sys.argv[2], sys.argv[3]
    out = sys.argv[4] if len(sys.argv) > 4 else "入場者プレゼント計算_サンプル入り.xlsx"
    base = datetime.date.fromisoformat(base_str)

    # ファイル名にハイフンが入っていて import できないため、パス指定で読み込む
    import importlib.util, os, tempfile
    spec = importlib.util.spec_from_file_location(
        "make_paste_excel", os.path.join(os.path.dirname(os.path.abspath(__file__)), "make-paste-excel.py"))
    builder = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(builder)
    tmp = os.path.join(tempfile.gettempdir(), "_template_for_sample.xlsx")
    builder.build(tmp)

    src = openpyxl.load_workbook(ps_path, data_only=True)
    gsrc = openpyxl.load_workbook(gift_path, data_only=True)
    dst = openpyxl.load_workbook(tmp)

    # PS を並び順に貼り付ける（データのあるシートだけ）
    used = []
    for name in src.sheetnames:
        ws = src[name]
        if ws.max_row < 10:
            continue
        has_perf = any(ws.cell(row=r, column=4).value == "A" for r in range(1, min(ws.max_row, PSROWS) + 1))
        if not has_perf:
            continue
        if len(used) >= builder.PSHEETS:
            print(f"  ※PSのシートが {builder.PSHEETS} 枚を超えたため「{name}」は省きました")
            break
        used.append(name)
        paste(ws, dst[f"PS{len(used)}"], PSROWS, 14)
    print(f"  PS を {len(used)} 枚貼り付け: {', '.join(used)}")

    # 入場者プレゼント情報を貼り付ける
    gws = gsrc[gsrc.sheetnames[0]]
    paste(gws, dst["プレゼント貼付"], GROWS, 5)
    titles = gift_data_rows(gws)
    print(f"  入場者プレゼント情報を {len(titles)} 件貼り付け（枠は {builder.GIFTS} 件）")

    # 設定
    st = dst["設定"]
    st["B2"] = base
    st["B3"] = 1
    wmap = weekday_map(used)
    for wd, ps_no in wmap.items():
        st.cell(row=3 + wd, column=7, value=ps_no)
    print("  曜日→PS: " + " ".join(f"{WD_NAMES[wd - 1]}=PS{ps}" for wd, ps in sorted(wmap.items())))

    # 自動生成のキーワードでは当たらない行を手直し（利用者が行う作業の再現）
    g = dst["プレゼント"]
    for i, title in enumerate(titles[:builder.GIFTS]):
        for needle, kw in KEYWORD_FIXES:
            if needle in title:
                g.cell(row=3 + i, column=3, value=kw)
                print(f"  キーワードを手直し: 「{title[:24]}」→ {kw}")

    dst.save(out)
    print(f"生成しました: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
