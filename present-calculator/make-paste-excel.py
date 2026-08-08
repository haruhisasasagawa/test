#!/usr/bin/env python3
"""
「パフォーマンススケジュールを貼り付けるだけ」で計算できる Excel テンプレートを生成する。

    python3 make-paste-excel.py [出力先.xlsx]

マクロ（VBA）は一切使わず、通常の数式だけで動く .xlsx を出力する。
生成物はテンプレート（空）で、利用者が PS1〜PS3 シートに PS を貼り付けて使う。
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.datavalidation import DataValidation

import os

# 検証用に小さく作りたいときは環境変数で上書きできる
DAYS = int(os.environ.get("PC_DAYS", 42))      # 計算する日数（6週）
KMAX = int(os.environ.get("PC_KMAX", 45))      # PS 1シートあたりに読み取る上映行の最大数
PSROWS = int(os.environ.get("PC_PSROWS", 250)) # PS 1シートあたりに読み取る行数
GIFTS = int(os.environ.get("PC_GIFTS", 12))    # 登録できる入場者プレゼントの数
PSHEETS = int(os.environ.get("PC_PSHEETS", 7)) # 貼り付けられる PS のシート数（1週間=最大7）
GROWS = int(os.environ.get("PC_GROWS", 150))   # 入場者プレゼント情報から読み取る行数
KEYLEN = int(os.environ.get("PC_KEYLEN", 8))   # 作品名から自動生成するキーワードの文字数
FIRST = 9                                      # 「計算」シートのデータ開始行

NAVY = "FF12233D"
LIGHT = "FFEAF1FC"
GRAY = "FFF2F4F8"

HEAD = dict(font=Font(bold=True, color="FFFFFFFF", size=11),
            fill=PatternFill("solid", fgColor=NAVY),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
INPUT = dict(font=Font(bold=True), fill=PatternFill("solid", fgColor=LIGHT))
THIN = Border(*[Side(style="thin", color="FFB8BFCC")] * 4)


def style(cell, spec=None, fmt=None, border=True):
    if spec:
        for k, v in spec.items():
            setattr(cell, k, v)
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = THIN
    return cell


def clean_text(ref):
    """括弧・空白・全角記号を取り除いた比較用の文字列を作る数式。"""
    expr = ref
    for a, b in [("『", ""), ("』", ""), ("「", ""), ("」", ""), ("【", ""), ("】", ""),
                 ("〈", ""), ("〉", ""), ("　", ""), (" ", "")]:
        expr = f'SUBSTITUTE({expr},"{a}","{b}")'
    return expr


# 突き合わせのときに落とす記号。全角・半角どちらで書かれていても同じ文字列になるようにする
#（実データに「スパイダーマン：」と「スパイダーマン:」の両方が出てくるため）
MATCH_NOISE = ["　", " ", "・", "･", "：", ":", "／", "/", "＆", "&", "！", "!",
               "？", "?", "、", "，", ",", "。", ".", "～", "〜", "＝", "=",
               "『", "』", "「", "」", "【", "】", "〈", "〉", "（", "(", "）", ")"]


def strip_noise(ref):
    """突き合わせ用に、空白と記号を取り除いた文字列を作る数式。"""
    expr = ref
    for ch in MATCH_NOISE:
        expr = f'SUBSTITUTE({expr},"{ch}","")'
    return expr


def to_half(ref, extra=()):
    """全角数字と指定した全角記号を半角にする数式。"""
    expr = ref
    for a, b in [("０", "0"), ("１", "1"), ("２", "2"), ("３", "3"), ("４", "4"),
                 ("５", "5"), ("６", "6"), ("７", "7"), ("８", "8"), ("９", "9")] + list(extra):
        expr = f'SUBSTITUTE({expr},"{a}","{b}")'
    return expr


def fullwidth_to_half(ref):
    """全角数字・全角スラッシュ・改行を半角スラッシュ区切りに正規化する数式を組み立てる。
    ASC() は環境差があるため使わず、SUBSTITUTE の入れ子で行う。"""
    expr = ref
    for a, b in [("０", "0"), ("１", "1"), ("２", "2"), ("３", "3"), ("４", "4"),
                 ("５", "5"), ("６", "6"), ("７", "7"), ("８", "8"), ("９", "9"),
                 ("／", "/"), ("　", "")]:
        expr = f'SUBSTITUTE({expr},"{a}","{b}")'
    return f'SUBSTITUTE({expr},CHAR(10),"/")'


def _date_formula(ref):
    """「8/7(金)」「7月24日(金)」「45678」「無くなるまで」等を日付に直す数式。
    年は書かれていないので、設定の起算日の年を使い、半年以上ずれる場合は前後の年に寄せる。"""
    t = to_half(f'{ref}&""', [("／", "/"), ("（", "("), ("　", ""), ("月", "/"), ("日", "")])
    t = f'TRIM(IFERROR(LEFT({t},FIND("(",{t})-1),{t}))'
    m = f'VALUE(LEFT({t},FIND("/",{t})-1))'
    dd = f'VALUE(MID({t},FIND("/",{t})+1,10))'
    y = "YEAR(設定!$B$2)"
    d0 = f'DATE({y},{m},{dd})'
    # 起算日から前後240日を超える場合は年をずらす
    fixed = (f'IF({d0}-設定!$B$2>240,DATE({y}-1,{m},{dd}),'
             f'IF(設定!$B$2-{d0}>240,DATE({y}+1,{m},{dd}),{d0}))')
    return f'IF({ref}="","",IF(ISNUMBER({ref}),{ref},IFERROR({fixed},"")))'


def build(path):
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ 使い方
    ws = wb.active
    ws.title = "使い方"
    ws.column_dimensions["A"].width = 118
    lines = [
        ("入場者プレゼント 配布可能期間 計算シート（貼り付けるだけ・マクロ不要）", True),
        ("", False),
        ("■ 手順", True),
        (f"1. パフォーマンススケジュールの各シートを、PS1〜PS{PSHEETS} シートの A1 に順に貼り付けます。", False),
        (f"   PS は週によってシート数が変わります（1枚〜7枚）。使う枚数だけ貼り付けてください。", False),
        ("   例1）3枚のとき：金曜 → PS1、土日月火木 → PS2、水曜 → PS3", False),
        ("   例2）1枚（1週間同じ編成）のとき：PS1 だけに貼り付け", False),
        ("   例3）7枚（毎日違う）のとき：PS1〜PS7 に1日ずつ貼り付け", False),
        ("   ※コピー元をそのまま貼り付けてください（結合セルのままで構いません）。", False),
        ("2. 「設定」シートで、計算起算日・想定稼働率・曜日ごとにどのPSを使うかを指定します。", False),
        ("   1枚しか貼っていない場合は、7曜日すべてに 1 を指定します。", False),
        ("   貼り付けが済むと「PS別の読み取り状況」に上映行数が出ます。0 のままなら貼り付け先が違います。", False),
        ("3. 入場者プレゼント情報（作品名／プレゼント内容／納品数／配布開始日／配布終了日）を、", False),
        ("   「プレゼント貼付」シートの A1 にそのまま貼り付けます。", False),
        ("   プレゼント名・納品数・配布開始日／終了日・作品キーワードが「プレゼント」シートに自動で入ります。", False),
        ("   （手入力でも使えます。貼り付けない場合は「プレゼント」シートに直接書いてください）", False),
        ("4. 「プレゼント」シートの『一致した作品数』を確認します。ここが最重要です。", True),
        ("   0 のものは PS の作品名と結びついていないので、作品キーワードを書き換えてください。", False),
        ("   キーワードは PS の作品名に含まれる文字列です（例：ミニオンズ、トイ・ストーリー）。", False),
        ("   （字）（吹）（IMAXレーザー）などの表記違いをまとめて拾うため、共通部分だけを入れます。", False),
        ("   0 でなくても、別の作品を拾っていることがあります。作品名が似ている週はご注意ください。", False),
        ("5. 「結果」シートに、プレゼントごとの配布可能最終日が表示されます。", False),
        ("", False),
        ("■ 計算のルール", True),
        ("1回の上映で必要な数 ＝ スクリーンのキャパ × 想定稼働率（切り上げ）× 1人あたりの配布数", False),
        ("その回のお客様全員に配布できない場合、その回は配布しません（在庫は減りません）。", False),
        ("在庫が大きいスクリーンに足りなくなっても、小さいスクリーンの回では配布を続けます。", False),
        ("", False),
        ("■ 必ずお読みください（この方式の割り切り）", True),
        (f"1. 計算できるのは起算日から {DAYS} 日（{DAYS // 7} 週）先までです。", False),
        ("   貼り付けた週の編成がそのまま続く前提のため、これ以上先は実務上あてになりません。", False),
        ("2. 1日の中の順序は、開始時刻順ではなく PS の並び順（スクリーン順）で計算します。", False),
        ("   配布可能『日』は変わりませんが、最終日の残数がわずかに違う場合があります。", False),
        ("3. PS のヘッダにある日本語の日付（例：2026年8月8日（土）～…）は読み取りません。", False),
        ("   起算日と曜日ごとのPS対応だけで日付を組み立てます。", False),
        (f"4. 読み取るのは PS 1シートにつき先頭 {PSROWS} 行、上映行は最大 {KMAX} 本までです。", False),
        (f"   貼り付けられる PS は最大 {PSHEETS} 枚です。", False),
        (f"5. 登録できるプレゼントは {GIFTS} 件までです。", False),
        ("6. 在庫欄には『起算日の時点で手元にある数』を入力してください。", False),
        ("   貼り付けから入るのは納品数です。起算日より前に配った分は差し引かれていません。", False),
        ("7. 作品キーワードは作品名の先頭から自動で作ります。必ず当たるわけではありません。", False),
        ("   『一致した作品数』が 0 のものは手で直してください（例：「映画クレヨンしんちゃん…」→「クレヨンしん」）。", False),
        ("8. 補助席は計上していません。キャパは PS の記載どおりです。", False),
        ("", False),
        ("■ 上映行の見分け方（PSの書式が変わったとき用）", True),
        ("D列が「A」で、E〜N列に「時:分」が1つ以上ある行を上映行として扱います。", False),
        ("A列の「スクリーン番号／キャパ」（例：７/391、1改行88）から、番号と座席数を取り出します。", False),
        ("A列が空の行は、上の行のスクリーン番号・キャパを引き継ぎます（結合セル対策）。", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.border = Border()
        if bold:
            c.font = Font(bold=True, size=13 if i == 1 else 11)

    # ------------------------------------------------------------------ 設定
    st = wb.create_sheet("設定")
    st.column_dimensions["A"].width = 20
    st.column_dimensions["B"].width = 16
    st.column_dimensions["C"].width = 44
    st.column_dimensions["E"].width = 10
    st.column_dimensions["F"].width = 12
    style(st.cell(row=1, column=1, value="設定"), dict(font=Font(bold=True, size=13)), border=False)
    style(st.cell(row=2, column=1, value="計算起算日"), dict(font=Font(bold=True)))
    style(st.cell(row=2, column=2), INPUT, fmt="yyyy/m/d")
    style(st.cell(row=2, column=3, value="←ここから計算します（貼り付けたPSの初日）"), border=False)
    style(st.cell(row=3, column=1, value="想定稼働率"), dict(font=Font(bold=True)))
    style(st.cell(row=3, column=2, value=1), INPUT, fmt="0%")
    style(st.cell(row=3, column=3, value="←満席想定なら 100%"), border=False)
    style(st.cell(row=4, column=1, value="計算日数"), dict(font=Font(bold=True)))
    style(st.cell(row=4, column=2, value=DAYS), fmt="0")
    style(st.cell(row=4, column=3, value="固定（このシートの作りで変えられません）"), border=False)

    # E列は WEEKDAY() が返す番号（1=日〜7=土）。VLOOKUP がこの列を引くので並べ替えないこと
    style(st.cell(row=3, column=5, value="番号"), HEAD)
    style(st.cell(row=3, column=6, value="曜日"), HEAD)
    style(st.cell(row=3, column=7, value="使うPS"), HEAD)
    st.column_dimensions["G"].width = 12
    for i, (num, name, default) in enumerate([(1, "日", 2), (2, "月", 2), (3, "火", 2), (4, "水", 3),
                                              (5, "木", 2), (6, "金", 1), (7, "土", 2)]):
        style(st.cell(row=4 + i, column=5, value=num), fmt="0")
        style(st.cell(row=4 + i, column=6, value=name))
        style(st.cell(row=4 + i, column=7, value=default), INPUT, fmt="0")
        st.cell(row=4 + i, column=5).alignment = Alignment(horizontal="center")
        st.cell(row=4 + i, column=6).alignment = Alignment(horizontal="center")
    style(st.cell(row=11, column=5, value="※「その曜日はPS1〜PS3のどれを使うか」を指定します"),
          dict(font=Font(size=9)), border=False)
    dv = DataValidation(type="list",
                        formula1='"' + ",".join(str(i) for i in range(1, PSHEETS + 1)) + '"',
                        allow_blank=False)
    st.add_data_validation(dv)
    dv.add("G4:G10")

    # 貼り付けが正しくできたか一目で分かるようにする（0 のままなら貼り付け先が違う）
    style(st.cell(row=13, column=1, value="PS別の読み取り状況"), dict(font=Font(bold=True)), border=False)
    for i, h in enumerate(["PS", "上映行数", "1日の座席数", "使う曜日"], start=1):
        style(st.cell(row=14, column=i, value=h), HEAD)
    for p in range(1, PSHEETS + 1):
        r = 14 + p
        style(st.cell(row=r, column=1, value=f"PS{p}"))
        style(st.cell(row=r, column=2, value=f"=SUM(解析{p}!$D$1:$D${PSROWS})"), fmt="0")
        style(st.cell(row=r, column=3,
                      value=f"=SUMPRODUCT(解析{p}!$C$1:$C${PSROWS},解析{p}!$E$1:$E${PSROWS})"), fmt="#,##0")
        style(st.cell(row=r, column=4, value=(
            f'=IF(COUNTIF($G$4:$G$10,{p})=0,"（未使用）",COUNTIF($G$4:$G$10,{p})&"日/週")')))

    # ------------------------------------------------------------------ PS貼付
    for p in range(1, PSHEETS + 1):
        s = wb.create_sheet(f"PS{p}")
        s.column_dimensions["B"].width = 34
        c = s.cell(row=1, column=16, value=f"← A1 を選んで、PS の {p} 番目のシートをそのまま貼り付けてください")
        c.font = Font(bold=True, color="FF1F5FBF")

    # ------------------------------------------- 入場者プレゼント情報の貼り付け先
    gp = wb.create_sheet("プレゼント貼付")
    for col, w in zip("ABCDE", [34, 34, 14, 14, 14]):
        gp.column_dimensions[col].width = w
    c = gp.cell(row=1, column=7,
                value="← A1 を選んで、入場者プレゼント情報（作品名／プレゼント内容／納品数／"
                      "配布開始日／配布終了日）をそのまま貼り付けてください")
    c.font = Font(bold=True, color="FF1F5FBF")

    # ------------------------------------------------------------------ 解析
    for p in range(1, PSHEETS + 1):
        a = wb.create_sheet(f"解析{p}")
        a.sheet_state = "visible"
        for col, w in zip("ABCDEFGIJKLM", [14, 8, 8, 6, 6, 34, 8, 5, 8, 34, 8, 6]):
            a.column_dimensions[col].width = w
        for r in range(1, PSROWS + 1):
            ps = f"PS{p}!"
            norm = fullwidth_to_half(f"{ps}$A{r}")
            # B: スクリーン番号（取れなければ上の行を引き継ぐ）
            prev_b = "" if r == 1 else f"B{r - 1}"
            prev_c = "" if r == 1 else f"C{r - 1}"
            a.cell(row=r, column=1, value=f'=IF({ps}$A{r}="","",{norm})')
            a.cell(row=r, column=2,
                   value=f'=IFERROR(VALUE(LEFT(A{r},FIND("/",A{r})-1)),{prev_b or 0})')
            a.cell(row=r, column=3,
                   value=f'=IFERROR(VALUE(MID(A{r},FIND("/",A{r})+1,10)),{prev_c or 0})')
            # D: 上映行フラグ（D列が"A" かつ E〜N に時刻がある）
            # 時刻の入った回数を数える。文字列の "10:35" と、時刻値で入っている場合の両方に対応
            times = (f'(COUNTIF({ps}$E{r}:$N{r},"*:*")'
                     f'+COUNTIFS({ps}$E{r}:$N{r},">0",{ps}$E{r}:$N{r},"<2"))')
            a.cell(row=r, column=4, value=f'=IF(AND({ps}$D{r}="A",{times}>0),1,0)')
            a.cell(row=r, column=5, value=f'=IF(D{r}=1,{times},0)')
            a.cell(row=r, column=6, value=f'=IF(D{r}=1,{ps}$B{r},"")')
            a.cell(row=r, column=7, value=f'=IF(D{r}=1,SUM($D$1:D{r}),"")')
        for k in range(1, KMAX + 1):
            a.cell(row=k, column=9, value=k)
            a.cell(row=k, column=10, value=f'=IFERROR(MATCH(I{k},$G$1:$G${PSROWS},0),0)')
            a.cell(row=k, column=11, value=f'=IF(J{k}=0,"",INDEX($F$1:$F${PSROWS},J{k}))')
            a.cell(row=k, column=12, value=f'=IF(J{k}=0,0,INDEX($C$1:$C${PSROWS},J{k}))')
            a.cell(row=k, column=13, value=f'=IF(J{k}=0,0,INDEX($E$1:$E${PSROWS},J{k}))')
            # N列: 空白を除いた作品名。プレゼントのキーワードはこの列と突き合わせる
            a.cell(row=k, column=14, value=f'=IF(K{k}="","",{strip_noise(f"K{k}")})')

    # ------------------------------------------------------- 取込（貼り付けたプレゼント情報の解析）
    im = wb.create_sheet("取込")
    for col, w in zip("ABCDEFGHI", [7, 7, 34, 34, 12, 12, 12, 20, 22]):
        im.column_dimensions[col].width = w
    for i, h in enumerate(["データ行", "通番", "作品名", "プレゼント名", "納品数",
                           "配布開始日", "配布終了日", "キーワード候補", "元の納品数表記"], start=1):
        style(im.cell(row=1, column=i), HEAD).value = h
    for r in range(2, GROWS + 2):
        src = r - 1                       # プレゼント貼付シートの行番号
        P = "プレゼント貼付!"
        a, b, cc, d, e = (f'{P}$A{src}', f'{P}$B{src}', f'{P}$C{src}',
                          f'{P}$D{src}', f'{P}$E{src}')
        # データ行の判定: 作品名があり、見出し行（「作品名」「納品数」）ではないこと
        im.cell(row=r, column=1, value=(
            f'=IF(AND({a}<>"",{a}<>"作品名",ISERROR(SEARCH("納品数",{cc}&"")),'
            f'OR({b}<>"",{cc}<>"")),1,0)'))
        im.cell(row=r, column=2, value=f'=IF(A{r}=1,SUM($A$2:A{r}),"")')
        im.cell(row=r, column=3, value=f'=IF(A{r}=1,{a},"")')
        im.cell(row=r, column=4, value=f'=IF(A{r}=1,{b},"")')
        # 納品数: 「2800（200）」の括弧内は一束の数なので落とし、「初回42000」等の語も除く
        qty = to_half(cc, [("（", "("), ("，", ""), (",", "")])
        qty = f'IFERROR(LEFT({qty},FIND("(",{qty})-1),{qty})'
        for w in ("初回", "約", "計", "枚", "個"):
            qty = f'SUBSTITUTE({qty},"{w}","")'
        im.cell(row=r, column=5, value=f'=IF(A{r}=0,"",IF({cc}="",0,IFERROR(VALUE(TRIM({qty})),0)))')
        for col, ref in ((6, d), (7, e)):
            im.cell(row=r, column=col, value=f'=IF(A{r}=0,"",{_date_formula(ref)})').number_format = "yyyy/m/d"
        # 「映画」「劇場版」はPSの作品名に付いていないことがあるので落としてからキーワードにする
        key = strip_noise(a)
        for w in ("映画", "劇場版"):
            key = f'SUBSTITUTE({key},"{w}","")'
        im.cell(row=r, column=8, value=f'=IF(A{r}=0,"",LEFT({key},{KEYLEN}))')
        im.cell(row=r, column=9, value=f'=IF(A{r}=0,"",{cc}&"")')

    # ------------------------------------------------------------------ プレゼント
    g = wb.create_sheet("プレゼント")
    headers = ["No", "プレゼント名", "作品キーワード", "除外キーワード", "在庫",
               "1人あたり", "配布開始日", "配布終了日", "一致した作品数", "取込元"]
    widths = [5, 30, 22, 16, 10, 9, 13, 13, 12, 30]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        g.column_dimensions[CL(i)].width = w
        style(g.cell(row=2, column=i, value=h), HEAD)
    style(g.cell(row=1, column=1,
                 value="「プレゼント貼付」に貼り付けると自動で入ります。直したいセルは上書きしてください"
                       "（数式が消えますが問題ありません）。"),
          dict(font=Font(bold=True)), border=False)
    for i in range(GIFTS):
        r = 3 + i
        n = i + 1
        style(g.cell(row=r, column=1, value=n), fmt="0")
        # 取込シートの n 件目を引く。貼り付けが無ければ空欄のまま手入力できる
        row_ref = f'MATCH({n},取込!$B$2:$B${GROWS + 1},0)+1'
        def pick(col, empty='""'):
            return (f'=IFERROR(IF(INDEX(取込!${col}$2:${col}${GROWS + 1},{row_ref}-1)="",{empty},'
                    f'INDEX(取込!${col}$2:${col}${GROWS + 1},{row_ref}-1)),{empty})')
        style(g.cell(row=r, column=2, value=pick("D")), INPUT)
        style(g.cell(row=r, column=3, value=pick("H")), INPUT)
        style(g.cell(row=r, column=4), INPUT)
        style(g.cell(row=r, column=5, value=pick("E", "0")), INPUT, fmt="#,##0")
        style(g.cell(row=r, column=6, value=1), INPUT, fmt="0")
        style(g.cell(row=r, column=7, value=pick("F")), INPUT, fmt="yyyy/m/d")
        style(g.cell(row=r, column=8, value=pick("G")), INPUT, fmt="yyyy/m/d")
        # キーワードが PS の作品名に何本当たったか。0 ならキーワードを直す必要がある
        hits = "+".join(f'COUNTIF(解析{p}!$N$1:$N${KMAX},"*"&$C{r}&"*")'
                        for p in range(1, PSHEETS + 1))
        style(g.cell(row=r, column=9, value=f'=IF($C{r}="","",{hits})'), fmt="0")
        style(g.cell(row=r, column=10, value=pick("C")))

    # ------------------------------------------------------------------ 計算
    c = wb.create_sheet("計算")
    last = FIRST + DAYS * KMAX - 1
    for col, w in zip("ABCDEF", [11, 5, 32, 7, 6, 8]):
        c.column_dimensions[col].width = w
    labels = ["プレゼント名", "作品キーワード", "除外キーワード", "在庫", "1人あたり", "配布開始日", "配布終了日"]
    for i, lab in enumerate(labels, start=1):
        style(c.cell(row=i, column=7, value=lab), dict(font=Font(bold=True, size=9)),
              border=False)
    for gi in range(GIFTS):
        gr = 3 + gi
        base = 8 + gi * 3      # A〜G を共通列に使うため、プレゼントの列は H から
        # 空欄のセルを参照すると "" ではなく 0 が返るため、必ず IF で空欄を空文字に戻す。
        # これを省くと「配布終了日が空欄＝無くなるまで」の判定が成立しない。
        def blank(ref, empty='""'):
            return f'=IF({ref}="",{empty},{ref})'
        c.cell(row=1, column=base, value=blank(f'プレゼント!$B${gr}')).font = Font(bold=True, size=9)
        c.cell(row=2, column=base, value=blank(f'プレゼント!$C${gr}')).font = Font(size=9)
        c.cell(row=3, column=base, value=blank(f'プレゼント!$D${gr}')).font = Font(size=9)
        c.cell(row=4, column=base, value=blank(f'プレゼント!$E${gr}', '0')).font = Font(size=9)
        c.cell(row=5, column=base, value=blank(f'プレゼント!$F${gr}', '1')).font = Font(size=9)
        c.cell(row=6, column=base, value=blank(f'プレゼント!$G${gr}', '0')).number_format = "yyyy/m/d"
        c.cell(row=7, column=base, value=blank(f'プレゼント!$H${gr}')).number_format = "yyyy/m/d"
        for off, name in enumerate(["対象回数", "配布回数", "残数"]):
            style(c.cell(row=8, column=base + off, value=name), HEAD)
            c.column_dimensions[CL(base + off)].width = 8
    for i, h in enumerate(["日付", "PS", "作品（PS表記）", "キャパ", "回数", "必要数", "照合用"], start=1):
        style(c.cell(row=8, column=i, value=h), HEAD)
    c.column_dimensions["G"].width = 32

    for d in range(DAYS):
        for k in range(1, KMAX + 1):
            r = FIRST + d * KMAX + (k - 1)
            style(c.cell(row=r, column=1, value=f"=設定!$B$2+{d}"), fmt="yyyy/m/d")
            style(c.cell(row=r, column=2,
                         value=f'=IFERROR(VLOOKUP(WEEKDAY($A{r},1),設定!$E$4:$G$10,3,FALSE),"")'), fmt="0")
            # 曜日ごとに使う PS が変わるため、貼り付け枚数分の CHOOSE で切り替える
            pick = ("CHOOSE($B{r},"
                    + ",".join(f"解析{p}!${{col}}${{k}}" for p in range(1, PSHEETS + 1)) + ")")
            style(c.cell(row=r, column=3,
                         value='=IFERROR(' + pick.format(r=r, col="K", k=k) + ',"")'))
            style(c.cell(row=r, column=4,
                         value="=IFERROR(" + pick.format(r=r, col="L", k=k) + ",0)"), fmt="#,##0")
            style(c.cell(row=r, column=5,
                         value="=IFERROR(" + pick.format(r=r, col="M", k=k) + ",0)"), fmt="0")
            style(c.cell(row=r, column=6,
                         value=f'=IF($D{r}=0,0,ROUNDUP($D{r}*設定!$B$3,0))'), fmt="#,##0")
            # 突き合わせは空白を除いた作品名で行う（PSは全角空白、プレゼント情報は半角空白のことがある）
            style(c.cell(row=r, column=7,
                         value='=IFERROR(' + pick.format(r=r, col="N", k=k) + ',"")'))
            for gi in range(GIFTS):
                base = 8 + gi * 3
                B = CL(base)          # 対象回数
                C_ = CL(base + 1)     # 配布回数
                D_ = CL(base + 2)     # 残数
                prev = f"{D_}{r - 1}" if k > 1 or d > 0 else f"{B}$4"
                style(c.cell(row=r, column=base, value=(
                    f'=IF(OR({B}$2="",$G{r}=""),0,'
                    f'IF(AND(ISNUMBER(SEARCH({B}$2,$G{r})),'
                    f'OR({B}$3="",NOT(ISNUMBER(SEARCH({B}$3,$G{r})))),'
                    f'$A{r}>={B}$6,OR({B}$7="",$A{r}<={B}$7)),$E{r},0))')), fmt="0")
                style(c.cell(row=r, column=base + 1, value=(
                    f'=IF(OR({B}{r}=0,$F{r}<=0),0,MIN({B}{r},INT({prev}/($F{r}*{B}$5))))')), fmt="0")
                style(c.cell(row=r, column=base + 2,
                             value=f"={prev}-{C_}{r}*$F{r}*{B}$5"), fmt="#,##0")

    c.freeze_panes = "A9"

    # ------------------------------------------------------------------ 結果
    res = wb.create_sheet("結果")
    for col, w in zip("ABCDEFGHIJ", [5, 30, 20, 10, 10, 10, 11, 10, 15, 22]):
        res.column_dimensions[col].width = w
    style(res.cell(row=1, column=1, value="入場者プレゼント 配布可能期間"),
          dict(font=Font(bold=True, size=14)), border=False)
    res.cell(row=2, column=1, value="=\"計算起算日 \"&TEXT(設定!$B$2,\"yyyy/m/d\")&\"　想定稼働率 \""
                                    f"&TEXT(設定!$B$3,\"0%\")&\"　計算期間 {DAYS}日\"").border = Border()
    heads = ["No", "プレゼント名", "作品キーワード", "在庫", "対象回数", "配布回数",
             "配布見込", "残数", "配布可能最終日", "判定"]
    for i, h in enumerate(heads, start=1):
        style(res.cell(row=3, column=i, value=h), HEAD)
    for gi in range(GIFTS):
        r = 4 + gi
        gr = 3 + gi
        base = 8 + gi * 3
        B, C_, D_ = CL(base), CL(base + 1), CL(base + 2)
        rngB = f"計算!${B}${FIRST}:${B}${last}"
        rngC = f"計算!${C_}${FIRST}:${C_}${last}"
        style(res.cell(row=r, column=1, value=gi + 1), fmt="0")
        style(res.cell(row=r, column=2, value=f"=プレゼント!$B${gr}"))
        style(res.cell(row=r, column=3, value=f"=プレゼント!$C${gr}"))
        style(res.cell(row=r, column=4, value=f"=プレゼント!$E${gr}"), fmt="#,##0")
        style(res.cell(row=r, column=5, value=f"=SUM({rngB})"), fmt="#,##0")
        style(res.cell(row=r, column=6, value=f"=SUM({rngC})"), fmt="#,##0")
        style(res.cell(row=r, column=7, value=f"=D{r}-H{r}"), fmt="#,##0")
        style(res.cell(row=r, column=8, value=f"=計算!${D_}${last}"), fmt="#,##0")
        style(res.cell(row=r, column=9, value=(
            f'=IF(F{r}=0,"-",SUMPRODUCT(MAX(({rngC}>0)*計算!$A${FIRST}:$A${last})))')),
            fmt="yyyy/m/d")
        # 判定は「対象の回すべてに配れたか」で行う。計算期間の末日と比べる方法では、
        # 配布終了日が先に来るプレゼント（週替わりなど）を在庫切れと誤判定する。
        style(res.cell(row=r, column=10, value=(
            f'=IF(プレゼント!$B${gr}="","",'
            f'IF(E{r}=0,"対象上映なし（キーワードを確認）",'
            f'IF(F{r}=0,"在庫不足で1回も配布できません",'
            f'IF(F{r}>=E{r},"対象の回すべてに配布できます（残 "&TEXT(H{r},"#,##0")&"）",'
            f'"この日で在庫切れ（配れない回 "&TEXT(E{r}-F{r},"#,##0")&"回）"))))')))
    res.freeze_panes = "A4"

    wb.save(path)
    print(f"生成しました: {path}")
    print(f"  計算行数: {DAYS * KMAX:,} 行 / プレゼント枠: {GIFTS} 件 / 期間: {DAYS}日")


if __name__ == "__main__":
    build(sys.argv[1] if len(sys.argv) > 1 else "入場者プレゼント計算_貼付式.xlsx")
